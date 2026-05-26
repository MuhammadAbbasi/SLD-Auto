# -*- coding: utf-8 -*-
"""
generate_sld.py
===============
Generate a DC Single Line Diagram DXF from an Excel cable list.

HOW TO RUN
----------
1. Install dependencies (once):
       pip install ezdxf openpyxl

2. Edit the three paths in the "FILE PATHS" section below.

3. Run:
       python generate_sld.py

WHAT YOU NEED TO PROVIDE
-------------------------
  DXF_PATH   – The source DXF file that contains Inverter 1.1 as a visual
                template (geometry is copied from it, NOT its labels).
  XLSX_PATH  – Excel cable list with sheet "2E802-3":
                  Col 1 = Inverter ID  (e.g. "1.2")
                  Col 3 = String name  (e.g. "1.2.3")
                  Col 4 = MPPT number  (e.g. 4)
  OUTPUT_PATH – Where the new DXF will be written (created from scratch).

OUTPUT LAYOUT
-------------
  - Each transformer = one horizontal row.
  - Inverters within a transformer = side-by-side columns.
  - One paper-space layout per transformer (Tx1 … Tx9), A3 landscape.
  - Variable MPPTs per inverter handled automatically (unused → "reserve").
"""

# ──────────────────────────────────────────────────────────────────────────────
#  FILE PATHS  ← change these three lines
# ──────────────────────────────────────────────────────────────────────────────
DXF_PATH    = r'C:\Users\user\Desktop\SLD Diagram\YANEL\26S001_2E103 - DC Single Line Diagram.dxf'
XLSX_PATH   = r'C:\Users\user\Desktop\SLD Diagram\YANEL\Yanel - Lista Cavi - Cavi LV-DC.xlsx'
OUTPUT_PATH = r'C:\Users\user\Desktop\SLD Diagram\YANEL\YANEL_SLD_Generated.dxf'
# ──────────────────────────────────────────────────────────────────────────────

import ezdxf
import openpyxl
import re
from collections import defaultdict

# ── TEMPLATE SETTINGS ─────────────────────────────────────────────────────────
# Y band that locates Inverter 1.1 in the source DXF (used as the visual stamp)
TEMPLATE_Y_MIN = 159400
TEMPLATE_Y_MAX = 168000

# Horizontal step between inverter columns (measured from original DXF: ~11 740)
COL_STEP = 11740

# Vertical step between transformer rows (template height ~8 600 + gap)
ROW_STEP = 10200

# Only use the narrow inverter-box portion of the template.
# The source DXF also contains a wide cable bus (x up to 257 000) that is
# row-level geometry; limiting to one COL_STEP prevents column overlap.
X_CUTOFF_RELATIVE = COL_STEP


# ══════════════════════════════════════════════════════════════════════════════
# 1.  READ EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def read_excel():
    wb   = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws   = wb['2E802-3']
    data = {}
    cur  = None
    for i in range(1, ws.max_row + 1):
        inv   = ws.cell(row=i, column=1).value
        sname = ws.cell(row=i, column=3).value
        mppt  = ws.cell(row=i, column=4).value
        if inv is not None and '.' in str(inv):
            try:
                t, v = str(inv).split('.')[:2]
                cur  = (int(t), int(v))
                data.setdefault(cur, defaultdict(list))
            except Exception:
                pass
        if sname and cur and mppt:
            try:
                data[cur][int(float(mppt))].append(str(sname))
            except Exception:
                pass
    return data


excel    = read_excel()
inv_list = sorted(excel.keys())

transformers = {}
for (T, I) in inv_list:
    transformers.setdefault(T, []).append(I)
for T in transformers:
    transformers[T] = sorted(transformers[T])
transformer_list = sorted(transformers.keys())

print(f"Inverters  : {len(inv_list)}")
print(f"Transformers: {len(transformer_list)}")
for T in transformer_list:
    inv_ids     = transformers[T]
    mppt_counts = [len(excel.get((T, I), {})) for I in inv_ids]
    str_counts  = [sum(len(v) for v in excel.get((T, I), {}).values()) for I in inv_ids]
    print(f"  Tx{T:>2}: {len(inv_ids):>3} inverters  "
          f"MPPTs {min(mppt_counts)}-{max(mppt_counts)}  "
          f"Strings {min(str_counts)}-{max(str_counts)}")


# ══════════════════════════════════════════════════════════════════════════════
# 2.  LOAD SOURCE DXF & FIND TEMPLATE ENTITIES
# ══════════════════════════════════════════════════════════════════════════════
doc = ezdxf.readfile(DXF_PATH)
msp = doc.modelspace()


def entity_y(e):
    try:
        t = e.dxftype()
        if t in ('TEXT', 'MTEXT', 'INSERT'):  return e.dxf.insert.y
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return e.dxf.center.y
        if t == 'LINE':                         return e.dxf.start.y
        if t == 'LWPOLYLINE':
            pts = list(e.get_points())
            return pts[0][1] if pts else None
        if t == 'POLYLINE':
            vs = list(e.vertices)
            return vs[0].dxf.location.y if vs else None
    except Exception:
        pass
    return None


tmpl_live = [e for e in msp
             if (y := entity_y(e)) is not None
             and TEMPLATE_Y_MIN <= y <= TEMPLATE_Y_MAX]
print(f"\nTemplate entities found : {len(tmpl_live)}")


# ══════════════════════════════════════════════════════════════════════════════
# 3.  EXTRACT INTO PLAIN DICTS  (before clearing model space)
# ══════════════════════════════════════════════════════════════════════════════
def common_attrs(e):
    d = {}
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if e.dxf.hasattr(a):
                d[a] = getattr(e.dxf, a)
        except Exception:
            pass
    return d


def extract_entity(e):
    t = e.dxftype()
    d = {'type': t}
    d.update(common_attrs(e))
    try:
        if t == 'LWPOLYLINE':
            d['pts']    = list(e.get_points())
            d['closed'] = e.closed
            if e.dxf.hasattr('const_width'):
                d['const_width'] = e.dxf.const_width
        elif t == 'MTEXT':
            pos = e.dxf.insert
            d['x'], d['y'], d['text'] = pos.x, pos.y, e.text
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'style'):
                try:
                    if e.dxf.hasattr(a): d[a] = getattr(e.dxf, a)
                except Exception: pass
        elif t == 'ARC':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'radius': e.dxf.radius,
                      'start_angle': e.dxf.start_angle,
                      'end_angle': e.dxf.end_angle})
        elif t == 'CIRCLE':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z, 'radius': e.dxf.radius})
        elif t == 'LINE':
            s, en = e.dxf.start, e.dxf.end
            d.update({'sx': s.x, 'sy': s.y, 'sz': s.z,
                      'ex': en.x, 'ey': en.y, 'ez': en.z})
        elif t == 'ELLIPSE':
            c, ma = e.dxf.center, e.dxf.major_axis
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'major_axis': (ma.x, ma.y, ma.z),
                      'ratio': e.dxf.ratio,
                      'start_param': e.dxf.start_param,
                      'end_param': e.dxf.end_param})
        elif t == 'INSERT':
            ins = e.dxf.insert
            d.update({'name': e.dxf.name,
                      'ix': ins.x, 'iy': ins.y,
                      'iz': ins.z if hasattr(ins, 'z') else 0})
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if e.dxf.hasattr(a): d[a] = getattr(e.dxf, a)
                except Exception: pass
        elif t == 'POLYLINE':
            pts3d = [v.dxf.location for v in e.vertices]
            d['pts3d'] = [(p.x, p.y, p.z) for p in pts3d]
        else:
            return None
    except Exception as ex:
        print(f"  [warn] extract {t}: {ex}")
        return None
    return d


def entity_min_x(d):
    t = d['type']
    try:
        if t == 'LWPOLYLINE':  return min(p[0] for p in d['pts'])
        if t == 'MTEXT':       return d['x']
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return d['cx'] - d.get('radius', 0)
        if t == 'LINE':        return min(d['sx'], d['ex'])
        if t == 'INSERT':      return d['ix']
        if t == 'POLYLINE':    return min(p[0] for p in d['pts3d'])
    except Exception: pass
    return 0


raw_dicts  = [d for e in tmpl_live if (d := extract_entity(e)) is not None]
TMPL_X_MIN = min((entity_min_x(d) for d in raw_dicts), default=0)
X_CUTOFF   = TMPL_X_MIN + X_CUTOFF_RELATIVE
tmpl_data  = [d for d in raw_dicts if entity_min_x(d) <= X_CUTOFF]

print(f"Template X left edge : {TMPL_X_MIN:.0f}")
print(f"X cutoff             : {X_CUTOFF:.0f}  "
      f"(kept {len(tmpl_data)}/{len(raw_dicts)} entities)")


# ══════════════════════════════════════════════════════════════════════════════
# 4.  CLASSIFY MTEXT
# ══════════════════════════════════════════════════════════════════════════════
PORT_RE   = re.compile(r'^\d+-\d+$')
STRING_RE = re.compile(r'String \d+\.\d+\.\d+')


def strip_mtext_codes(txt):
    """Remove DXF MTEXT formatting codes, keep visible text."""
    c = re.sub(r'\\[A-Za-z][^;]*;', '', txt)   # remove \fontcode; sequences
    c = re.sub(r'[{}]', '', c)                   # remove bare braces (NOT content)
    return c.strip().replace('\n', ' ')

def classify(txt):
    c = strip_mtext_codes(txt)
    if re.search(r'INVERTER 1\.1', txt, re.I) and 'P=' in txt: return 'title'
    if re.search(r'Cabin Tx\.\d+.*Inverter', txt):              return 'cabin_label'
    if re.match(r'^CABIN \d+$', c.strip()):                     return 'cabin_header'
    if re.search(r'String \d+\.\d+\.\d+', c) or c.strip() == 'reserve':
        return 'string_label'
    return 'fixed'


tmpl_texts = []
for d in tmpl_data:
    if d['type'] == 'MTEXT':
        d['cls'] = classify(d['text'])
        tmpl_texts.append(d)

print(f"MTEXT: {len(tmpl_texts)} total  |  "
      + "  ".join(f"{k}={sum(1 for m in tmpl_texts if m['cls']==k)}"
                  for k in ('title','cabin_header','cabin_label','string_label','fixed')))


# ══════════════════════════════════════════════════════════════════════════════
# 5.  BUILD MPPT-PORT → STRING-LABEL POSITION MAP
# ══════════════════════════════════════════════════════════════════════════════
port_labels   = [(m['x'], m['y'], m['text'].strip())
                 for m in tmpl_texts if PORT_RE.match(m['text'].strip())]
string_labels = [m for m in tmpl_texts if m['cls'] == 'string_label']

mppt_port_sl = {}
for px, py, ptxt in port_labels:
    best, best_d = None, 9999
    for sl in string_labels:
        if abs(sl['y'] - py) < 80 and sl['x'] > px:
            dd = abs(sl['y'] - py)
            if dd < best_d:
                best_d, best = dd, sl
    if best:
        m2 = re.match(r'^(\d+)-(\d+)$', ptxt)
        if m2:
            mppt_port_sl[(int(m2.group(1)), int(m2.group(2)))] = best

panel_sl = next(
    (sl for sl in string_labels
     if not any(abs(sl['y'] - py) < 80 for _, py, _ in port_labels)), None)
if panel_sl:
    mppt_port_sl[(1, 1)] = panel_sl

print(f"MPPT/port slots: {len(mppt_port_sl)}")


# ══════════════════════════════════════════════════════════════════════════════
# 6.  STRING LOOKUP
#     Returns "String T.I.N" or "reserve" (handles variable MPPT counts)
# ══════════════════════════════════════════════════════════════════════════════
def get_string(T, I, mppt, port):
    mppt_list = excel.get((T, I), {}).get(mppt, [])
    idx       = port - 1
    return f"String {mppt_list[idx]}" if idx < len(mppt_list) else "reserve"


# ══════════════════════════════════════════════════════════════════════════════
# 7.  POSITION HELPER
# ══════════════════════════════════════════════════════════════════════════════
def inv_offsets(T, I):
    row = transformer_list.index(T)
    col = transformers[T].index(I)
    return col * COL_STEP, -row * ROW_STEP


# ══════════════════════════════════════════════════════════════════════════════
# 8.  ENTITY PLACEMENT HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def apply_common(ne, d):
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if a in d: setattr(ne.dxf, a, d[a])
        except Exception: pass


def add_entity(layout, d, dx, dy):
    t = d['type']
    try:
        if t == 'LWPOLYLINE':
            pts = [(p[0]+dx, p[1]+dy) + tuple(p[2:]) for p in d['pts']]
            ne  = layout.add_lwpolyline(pts)
            ne.closed = d.get('closed', False)
            if 'const_width' in d: ne.dxf.const_width = d['const_width']
            apply_common(ne, d)
        elif t == 'MTEXT':
            attribs = {'insert': (d['x']+dx, d['y']+dy)}
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
                if a in d: attribs[a] = d[a]
            if 'color' in d: attribs['color'] = d['color']
            ne = layout.add_mtext(d['text'], dxfattribs=attribs)
            apply_common(ne, d)
        elif t == 'ARC':
            ne = layout.add_arc((d['cx']+dx, d['cy']+dy, d['cz']),
                                d['radius'], d['start_angle'], d['end_angle'])
            apply_common(ne, d)
        elif t == 'CIRCLE':
            ne = layout.add_circle((d['cx']+dx, d['cy']+dy, d['cz']), d['radius'])
            apply_common(ne, d)
        elif t == 'LINE':
            ne = layout.add_line((d['sx']+dx, d['sy']+dy, d['sz']),
                                 (d['ex']+dx, d['ey']+dy, d['ez']))
            apply_common(ne, d)
        elif t == 'ELLIPSE':
            ne = layout.add_ellipse(
                center=(d['cx']+dx, d['cy']+dy, d['cz']),
                major_axis=d['major_axis'], ratio=d['ratio'],
                start_param=d['start_param'], end_param=d['end_param'])
            apply_common(ne, d)
        elif t == 'INSERT':
            ne = layout.add_blockref(d['name'], (d['ix']+dx, d['iy']+dy, d['iz']))
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d: setattr(ne.dxf, a, d[a])
                except Exception: pass
            apply_common(ne, d)
        elif t == 'POLYLINE':
            pts = [(p[0]+dx, p[1]+dy, p[2]) for p in d['pts3d']]
            if pts:
                ne = layout.add_polyline3d(pts)
                apply_common(ne, d)
    except Exception as ex:
        print(f"  [warn] add_entity {t}: {ex}")


GREY = 8   # ACI colour 8 = medium grey in AutoCAD/BricsCAD

def add_mtext_var(layout, d, dx, dy, text, color=None):
    try:
        attribs = {'insert': (d['x']+dx, d['y']+dy)}
        for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                  'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
            if a in d: attribs[a] = d[a]
        # Copy colour from template first, then let caller override
        if 'color' in d:
            attribs['color'] = d['color']
        if color is not None:
            attribs['color'] = color   # caller colour wins (e.g. grey for "reserve")
        ne = layout.add_mtext(text, dxfattribs=attribs)
        # Apply remaining common attrs (linetype, lineweight, ltscale)
        for a in ('linetype', 'lineweight', 'ltscale'):
            try:
                if a in d: setattr(ne.dxf, a, d[a])
            except Exception: pass
    except Exception as ex:
        print(f"  [warn] add_mtext_var: {ex}")


# ══════════════════════════════════════════════════════════════════════════════
# 9.  CLEAR MODEL SPACE
# ══════════════════════════════════════════════════════════════════════════════
print("\nClearing model space ...")
for e in list(msp):
    try: msp.delete_entity(e)
    except Exception: pass


# ══════════════════════════════════════════════════════════════════════════════
# 10. GENERATE ALL INVERTER SECTIONS
# ══════════════════════════════════════════════════════════════════════════════
title_d     = next((m for m in tmpl_texts if m['cls'] == 'title'),        None)
cabin_hdr_d = next((m for m in tmpl_texts if m['cls'] == 'cabin_header'), None)
cabin_lbl_d = next((m for m in tmpl_texts if m['cls'] == 'cabin_label'),  None)

print(f"Generating {len(inv_list)} inverter sections ...")
for idx, (T, I) in enumerate(inv_list):
    dx, dy = inv_offsets(T, I)

    for d in tmpl_data:
        if d['type'] == 'MTEXT':
            if d.get('cls') == 'fixed':
                add_entity(msp, d, dx, dy)
        else:
            add_entity(msp, d, dx, dy)

    if title_d:
        add_mtext_var(msp, title_d, dx, dy,
                      f"INVERTER {T}.{I} - P= 350 KWp - P= 320 KWac @40°C")
    if cabin_hdr_d:
        add_mtext_var(msp, cabin_hdr_d, dx, dy, f"CABIN {T}")
    if cabin_lbl_d:
        add_mtext_var(msp, cabin_lbl_d, dx, dy,
                      f"\\pxqc;Cabin Tx.{T}\\PInverter {T}.{I}")

    for (mppt, port), sl in mppt_port_sl.items():
        label = get_string(T, I, mppt, port)
        # Embed grey colour inline in the text for "reserve" labels.
        # Using DXF MTEXT inline colour code \C8; is more reliable than
        # setting entity-level colour, which ezdxf doesn't always persist.
        text_out = f"\\C8;{label}" if label == "reserve" else label
        add_mtext_var(msp, sl, dx, dy, text_out)

    if (idx + 1) % 10 == 0:
        print(f"  {idx+1:>3}/{len(inv_list)}")

print("All sections generated.")


# ══════════════════════════════════════════════════════════════════════════════
# 11. PAPER SPACE LAYOUTS  (one per transformer, A3 landscape)
# ══════════════════════════════════════════════════════════════════════════════
for lname in [l.name for l in doc.layouts if l.name != 'Model']:
    try: doc.layouts.delete(lname)
    except Exception: pass

TMPL_X_CENTER = TMPL_X_MIN + COL_STEP / 2
TMPL_HEIGHT   = TEMPLATE_Y_MAX - TEMPLATE_Y_MIN
TMPL_Y_CENTER = (TEMPLATE_Y_MIN + TEMPLATE_Y_MAX) / 2

for T in transformer_list:
    try:    layout = doc.layouts.new(f"Tx{T}")
    except: layout = doc.layouts.get(f"Tx{T}")

    n_inv         = len(transformers[T])
    row           = transformer_list.index(T)
    row_center_x  = TMPL_X_CENTER + (n_inv - 1) * COL_STEP / 2
    row_center_y  = TMPL_Y_CENTER - row * ROW_STEP
    row_width     = COL_STEP * n_inv
    view_height   = max(TMPL_HEIGHT * 1.05,
                        row_width / (420.0 / 297.0) * 1.05)

    layout.add_viewport(
        center=(210, 148.5), size=(420, 297),
        view_center_point=(row_center_x, row_center_y),
        view_height=view_height)

print(f"Created {len(transformer_list)} layouts.")


# ================================================================================
# 12. SAVE
# ================================================================================
print(f"\nSaving -> {OUTPUT_PATH}")
doc.saveas(OUTPUT_PATH)
print("Done!")
