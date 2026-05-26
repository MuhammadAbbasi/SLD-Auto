# -*- coding: utf-8 -*-
"""
sld_gui.py  -  GUI front-end for the DC Single Line Diagram generator.
Run:  python sld_gui.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
import os
import re
import json
from collections import defaultdict
from datetime import datetime

_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logoA176LAB.jpg')


def _load_logo_tk(size=(90, 90)):
    """Return a tkinter PhotoImage of the logo, or None if PIL is unavailable."""
    try:
        from PIL import Image, ImageTk
        img = Image.open(_LOGO_PATH).convert('RGBA')
        img.thumbnail(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  GENERATION CORE
# ─────────────────────────────────────────────────────────────────────────────

def _generate(cfg, log):
    """Run the full SLD generation with a config dict. log(str) sends text to UI."""
    try:
        import ezdxf as _ez
    except ImportError:
        raise RuntimeError("ezdxf not installed.  Run:  pip install ezdxf")
    try:
        import openpyxl as _xl
    except ImportError:
        raise RuntimeError("openpyxl not installed.  Run:  pip install openpyxl")

    TEMPLATE_DXF = cfg['template_dxf']
    XLSX_PATH    = cfg['xlsx_path']
    OUTPUT_PATH  = cfg['output_path']

    panel_model       = cfg.get('panel_model', '').strip()
    panel_power_wp    = float(cfg.get('panel_power_wp') or 0)
    panels_per_string = int(float(cfg.get('panels_per_string') or 0))
    inverter_model    = cfg.get('inverter_model', '').strip()
    dc_power_kwp      = float(cfg.get('dc_power_kwp') or 0)
    ac_power_kwac     = float(cfg.get('ac_power_kwac') or 0)
    temp_rating       = float(cfg.get('temp_rating') or 40)
    transformer_power = cfg.get('transformer_power', '').strip()

    PORT_RE   = re.compile(r'^\d+-\d+$')
    STRING_RE = re.compile(r'String \d+\.\d+\.\d+')

    # ── 1. Read Excel ─────────────────────────────────────────────────────────
    log("Reading Excel …")
    wb = _xl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['2E802-3']
    excel = {}
    cur   = None
    for i in range(1, ws.max_row + 1):
        inv   = ws.cell(row=i, column=1).value
        sname = ws.cell(row=i, column=3).value
        mppt  = ws.cell(row=i, column=4).value
        if inv is not None and '.' in str(inv):
            try:
                t, v = str(inv).split('.')[:2]
                cur  = (int(t), int(v))
                excel.setdefault(cur, defaultdict(list))
            except Exception:
                pass
        if sname and cur and mppt:
            try:
                excel[cur][int(float(mppt))].append(str(sname))
            except Exception:
                pass

    inv_list = sorted(excel.keys())
    transformers = {}
    for (T, I) in inv_list:
        transformers.setdefault(T, []).append(I)
    for T in transformers:
        transformers[T] = sorted(transformers[T])
    transformer_list = sorted(transformers.keys())

    log(f"Inverters  : {len(inv_list)}")
    log(f"Transformers: {len(transformer_list)}")
    for T in transformer_list:
        ii = transformers[T]
        mc = [len(excel.get((T, i), {})) for i in ii]
        sc = [sum(len(v) for v in excel.get((T, i), {}).values()) for i in ii]
        log(f"  Tx{T}: {len(ii)} inverters  "
            f"MPPTs {min(mc)}-{max(mc)}  Strings {min(sc)}-{max(sc)}")

    # ── 2. Load template DXF and extract template-band entities ──────────────
    log(f"Loading template DXF …")
    doc = _ez.readfile(TEMPLATE_DXF)
    msp = doc.modelspace()

    TEMPLATE_Y_MIN = 159_400
    TEMPLATE_Y_MAX = 168_000
    COL_STEP       = 11_740
    ROW_STEP       = 10_200

    def _ent_y(e):
        try:
            t = e.dxftype()
            if t in ('TEXT', 'MTEXT', 'INSERT'):  return e.dxf.insert.y
            if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return e.dxf.center.y
            if t == 'LINE':                        return e.dxf.start.y
            if t == 'LWPOLYLINE':
                pts = list(e.get_points())
                return pts[0][1] if pts else None
        except Exception:
            pass
        return None

    def _common_attrs(e):
        d = {}
        for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
            try:
                if e.dxf.hasattr(a):
                    d[a] = getattr(e.dxf, a)
            except Exception:
                pass
        return d

    def _extract(e):
        t = e.dxftype()
        d = {'type': t}
        d.update(_common_attrs(e))
        try:
            if t == 'LWPOLYLINE':
                d['pts']    = [list(p) for p in e.get_points()]
                d['closed'] = e.closed
                if e.dxf.hasattr('const_width'):
                    d['const_width'] = e.dxf.const_width
            elif t == 'MTEXT':
                pos = e.dxf.insert
                d.update({'x': pos.x, 'y': pos.y, 'text': e.text})
                for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                          'line_spacing_style', 'line_spacing_factor', 'style'):
                    try:
                        if e.dxf.hasattr(a):
                            d[a] = getattr(e.dxf, a)
                    except Exception:
                        pass
            elif t == 'ARC':
                c = e.dxf.center
                d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                          'radius': e.dxf.radius,
                          'start_angle': e.dxf.start_angle,
                          'end_angle':   e.dxf.end_angle})
            elif t == 'CIRCLE':
                c = e.dxf.center
                d.update({'cx': c.x, 'cy': c.y, 'cz': c.z, 'radius': e.dxf.radius})
            elif t == 'LINE':
                s, en = e.dxf.start, e.dxf.end
                d.update({'sx': s.x, 'sy': s.y, 'sz': s.z,
                          'ex': en.x, 'ey': en.y, 'ez': en.z})
            elif t == 'ELLIPSE':
                c, ma = e.dxf.center, e.dxf.major_axis
                d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                          'major_axis': [ma.x, ma.y, ma.z],
                          'ratio': e.dxf.ratio,
                          'start_param': e.dxf.start_param,
                          'end_param':   e.dxf.end_param})
            elif t == 'INSERT':
                ins = e.dxf.insert
                d.update({'name': e.dxf.name,
                          'ix': ins.x, 'iy': ins.y,
                          'iz': ins.z if hasattr(ins, 'z') else 0.0})
                for a in ('xscale', 'yscale', 'rotation'):
                    try:
                        if e.dxf.hasattr(a):
                            d[a] = getattr(e.dxf, a)
                    except Exception:
                        pass
            elif t == 'POLYLINE':
                d['pts3d'] = [[v.dxf.location.x, v.dxf.location.y, v.dxf.location.z]
                              for v in e.vertices]
            else:
                return None
        except Exception as ex:
            log(f"  [warn] extract {t}: {ex}")
            return None
        return d

    raw_ents = [e for e in msp
                if (y := _ent_y(e)) is not None
                and TEMPLATE_Y_MIN <= y <= TEMPLATE_Y_MAX]
    log(f"Template band: {len(raw_ents)} entities found")

    all_dicts = [d for e in raw_ents if (d := _extract(e)) is not None]

    # Compute xmin from extracted entities
    xs = []
    for d in all_dicts:
        t = d['type']
        try:
            if t == 'LWPOLYLINE': xs += [p[0] for p in d['pts']]
            elif t == 'MTEXT':    xs.append(d['x'])
            elif t in ('ARC', 'CIRCLE', 'ELLIPSE'): xs.append(d['cx'])
            elif t == 'LINE':     xs += [d['sx'], d['ex']]
            elif t == 'INSERT':   xs.append(d['ix'])
            elif t == 'POLYLINE': xs += [p[0] for p in d['pts3d']]
        except Exception:
            pass
    xmin = min(xs) if xs else 0

    def min_x(d):
        t = d['type']
        try:
            if t == 'LWPOLYLINE': return min(p[0] for p in d['pts'])
            if t == 'MTEXT':      return d['x']
            if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return d['cx'] - d.get('radius', 0)
            if t == 'LINE':       return min(d['sx'], d['ex'])
            if t == 'INSERT':     return d['ix']
            if t == 'POLYLINE':   return min(p[0] for p in d['pts3d'])
        except Exception:
            pass
        return 0

    xcut = xmin + COL_STEP

    def _is_placeholder_rect(d):
        if d.get('type') != 'LWPOLYLINE' or not d.get('closed'):
            return False
        pts = d.get('pts', [])
        if len(pts) != 4:
            return False
        xs2 = [p[0] for p in pts]
        return min(xs2) > 23500 and (max(xs2) - min(xs2)) > 500

    tmpl = [d for d in all_dicts
            if min_x(d) <= xcut and not _is_placeholder_rect(d)]
    log(f"Template: {len(tmpl)}/{len(all_dicts)} entities in column slice")

    # ── 4. Classify MTEXT ─────────────────────────────────────────────────────
    def classify(txt):
        c = re.sub(r'\{[^}]*\}', '', re.sub(r'\\[A-Za-z][^;]*;', '', txt))
        c = c.strip().replace('\n', ' ')
        if re.search(r'INVERTER 1\.1', txt, re.I) and 'P=' in txt:
            return 'title'
        if re.search(r'Cabin Tx\.\d+.*Inverter', txt):
            return 'cabin_label'
        if re.match(r'^CABIN \d+$', c):
            return 'cabin_header'
        # "28 PV modules in series - Model Xwp" is the wide top-slot label
        if STRING_RE.search(txt) or c == 'reserve' or re.search(r'\bPV modules?\b', txt, re.I):
            return 'string_label'
        return 'fixed'

    tmpl_texts = []
    for d in tmpl:
        if d['type'] == 'MTEXT':
            d['cls'] = classify(d['text'])
            tmpl_texts.append(d)

    # ── 5. MPPT-port → string label position map ──────────────────────────────
    port_lbl = [(m['x'], m['y'], m['text'].strip())
                for m in tmpl_texts if PORT_RE.match(m['text'].strip())]
    str_lbl  = [m for m in tmpl_texts if m['cls'] == 'string_label']
    mppt_map = {}
    for px, py, ptxt in port_lbl:
        best, bd = None, 9999
        for sl in str_lbl:
            if sl['x'] > px:
                dd = abs(sl['y'] - py)
                if dd < 400 and dd < bd:   # raised from 80 → 400 for top slot
                    bd, best = dd, sl
        if best:
            m2 = re.match(r'^(\d+)-(\d+)$', ptxt)
            if m2:
                mppt_map[(int(m2.group(1)), int(m2.group(2)))] = best

    # Catch any orphan string_label not claimed by a port (fallback for port 1-1)
    panel_sl = next(
        (sl for sl in str_lbl
         if not any(abs(sl['y'] - py) < 80 for _, py, _ in port_lbl)),
        None)
    if panel_sl and (1, 1) not in mppt_map:
        mppt_map[(1, 1)] = panel_sl
    log(f"MPPT/port slots: {len(mppt_map)}")

    # ── 6. Label builders ─────────────────────────────────────────────────────
    def make_string_label(T, I, mppt, port):
        lst = excel.get((T, I), {}).get(mppt, [])
        if port - 1 < len(lst):
            label = f"String {lst[port - 1]}"
            if panels_per_string > 0:
                if panel_model:
                    suffix = f" {panel_power_wp:.0f}Wp" if panel_power_wp > 0 else ""
                    label += f" - {panels_per_string}× {panel_model}{suffix}"
                else:
                    label += f" ({panels_per_string}P)"
            return label
        return "reserve"

    def make_inv_title(T, I):
        dc = dc_power_kwp
        if dc <= 0 and panels_per_string > 0 and panel_power_wp > 0:
            n_str = sum(len(v) for v in excel.get((T, I), {}).values())
            dc = n_str * panels_per_string * panel_power_wp / 1000.0
        parts = [f"INVERTER {T}.{I}"]
        if inverter_model:
            parts.append(inverter_model)
        if dc > 0:
            parts.append(f"P= {dc:.0f} KWp")
        if ac_power_kwac > 0:
            parts.append(f"P= {ac_power_kwac:.0f} KWac @{temp_rating:.0f}°C")
        return " - ".join(parts)

    def make_cabin_hdr(T):
        return f"CABIN {T}\n{transformer_power}" if transformer_power else f"CABIN {T}"

    COLUMN_SPACING = int(COL_STEP * 1.22)   # ~22% wider gap between inverter columns
    ROW_SPACING    = int(ROW_STEP * 1.18)   # ~18% taller gap between transformer rows

    def inv_offset(T, I):
        return (transformers[T].index(I) * COLUMN_SPACING,
                -transformer_list.index(T) * ROW_SPACING)

    # ── 7. Entity placement ───────────────────────────────────────────────────
    def apply_common(ne, d):
        for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
            try:
                if a in d:
                    setattr(ne.dxf, a, d[a])
            except Exception:
                pass

    def place(layout, d, dx, dy):
        t = d['type']
        try:
            if t == 'LWPOLYLINE':
                pts = [(p[0] + dx, p[1] + dy) + tuple(p[2:]) for p in d['pts']]
                ne  = layout.add_lwpolyline(pts)
                ne.closed = d.get('closed', False)
                if 'const_width' in d:
                    ne.dxf.const_width = d['const_width']
                apply_common(ne, d)
            elif t == 'MTEXT':
                att = {'insert': (d['x'] + dx, d['y'] + dy)}
                for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                          'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
                    if a in d:
                        att[a] = d[a]
                ne = layout.add_mtext(d['text'], dxfattribs=att)
                apply_common(ne, d)
            elif t == 'ARC':
                ne = layout.add_arc(
                    (d['cx'] + dx, d['cy'] + dy, d['cz']),
                    d['radius'], d['start_angle'], d['end_angle'])
                apply_common(ne, d)
            elif t == 'CIRCLE':
                ne = layout.add_circle(
                    (d['cx'] + dx, d['cy'] + dy, d['cz']), d['radius'])
                apply_common(ne, d)
            elif t == 'LINE':
                ne = layout.add_line(
                    (d['sx'] + dx, d['sy'] + dy, d['sz']),
                    (d['ex'] + dx, d['ey'] + dy, d['ez']))
                apply_common(ne, d)
            elif t == 'ELLIPSE':
                ne = layout.add_ellipse(
                    center=(d['cx'] + dx, d['cy'] + dy, d['cz']),
                    major_axis=d['major_axis'], ratio=d['ratio'],
                    start_param=d['start_param'], end_param=d['end_param'])
                apply_common(ne, d)
            elif t == 'INSERT':
                ne = layout.add_blockref(
                    d['name'], (d['ix'] + dx, d['iy'] + dy, d['iz']))
                for a in ('xscale', 'yscale', 'rotation'):
                    try:
                        if a in d:
                            setattr(ne.dxf, a, d[a])
                    except Exception:
                        pass
                apply_common(ne, d)
            elif t == 'POLYLINE':
                pts = [(p[0] + dx, p[1] + dy, p[2]) for p in d['pts3d']]
                if pts:
                    ne = layout.add_polyline3d(pts)
                    apply_common(ne, d)
        except Exception as ex:
            log(f"  [warn] place {t}: {ex}")

    def place_mtext(layout, d, dx, dy, text):
        try:
            att = {'insert': (d['x'] + dx, d['y'] + dy)}
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
                if a in d:
                    att[a] = d[a]
            ne = layout.add_mtext(text, dxfattribs=att)
            apply_common(ne, d)
        except Exception as ex:
            log(f"  [warn] place_mtext: {ex}")

    # ── 8. Clear modelspace and paper-space layouts (keep all DXF resources) ──
    log("Preparing output DXF …")
    msp.delete_all_entities()
    paper_layouts = [l.name for l in doc.layouts if not l.is_modelspace]
    for name in paper_layouts:
        try:
            doc.layouts.delete(name)
        except Exception:
            pass
    if paper_layouts:
        log(f"  Removed {len(paper_layouts)} paper-space layout(s) from template")

    # ── 9. Generate all inverter sections ─────────────────────────────────────
    td  = next((m for m in tmpl_texts if m['cls'] == 'title'),        None)
    chd = next((m for m in tmpl_texts if m['cls'] == 'cabin_header'), None)
    cld = next((m for m in tmpl_texts if m['cls'] == 'cabin_label'),  None)

    log(f"Generating {len(inv_list)} inverter sections …")
    for idx, (T, I) in enumerate(inv_list):
        dx, dy = inv_offset(T, I)
        for d in tmpl:
            if d['type'] == 'MTEXT':
                if d.get('cls') == 'fixed':
                    place(msp, d, dx, dy)
            else:
                place(msp, d, dx, dy)
        if td:
            place_mtext(msp, td,  dx, dy, make_inv_title(T, I))
        if chd:
            place_mtext(msp, chd, dx, dy, make_cabin_hdr(T))
        if cld:
            place_mtext(msp, cld, dx, dy, f"\\pxqc;Cabin Tx.{T}\\PInverter {T}.{I}")
        for (mppt, port), sl in mppt_map.items():
            # Widen the MTEXT box so long labels never wrap to a second line
            sl_wide = dict(sl, width=max(sl.get('width', 0), 3000))
            place_mtext(msp, sl_wide, dx, dy, make_string_label(T, I, mppt, port))
        if (idx + 1) % 10 == 0 or (idx + 1) == len(inv_list):
            log(f"  {idx + 1}/{len(inv_list)} inverters done")

    # ── 10. Save ──────────────────────────────────────────────────────────────
    log(f"Saving → {OUTPUT_PATH}")
    doc.saveas(OUTPUT_PATH)
    log("Done!  SLD generated successfully.")


# ─────────────────────────────────────────────────────────────────────────────
#  HISTORY STORE
# ─────────────────────────────────────────────────────────────────────────────

_HISTORY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'history.json')

_DEFAULTS = {
    'panel_model': [
        'JA Solar JAM72D42-625/LB',
        'JA Solar JAM72S20-460/MR',
        'Longi Solar LR5-72HBD-580M',
        'Longi Solar LR5-72HBD-545M',
        'Canadian Solar HiKu7 CS7N-655MB',
        'Jinko Solar JKM660M-78HL4-V',
        'Trina Solar TSM-670NEG21C.20',
        'REC Alpha Pure-R 430AA',
    ],
    'panel_power_wp': [
        '430', '460', '500', '540', '545',
        '570', '580', '600', '625', '640', '655', '660',
    ],
    'panels_per_string': ['16', '18', '20', '22', '24', '26', '28', '30'],
    'inverter_model': [
        'Sungrow SG350HX',
        'Sungrow SG250HX',
        'Sungrow SG125HX',
        'Huawei SUN2000-330KTL',
        'Huawei SUN2000-275KTL-H1',
        'Huawei SUN2000-215KTL-H3',
        'ABB PVS-250-TL',
        'SMA Sunny Tripower Core2 150',
        'Fronius Symo GEN24 25.0 Plus',
        'Power Electronics FS5000CU17',
    ],
    'dc_power_kwp': ['250', '275', '300', '320', '330', '350', '375', '400', '450', '500'],
    'ac_power_kwac': ['200', '225', '250', '275', '300', '320', '330', '350'],
}


class _HistoryStore:
    """Persists combo-box histories to a JSON file keyed by field name."""

    def __init__(self, path):
        self._path = path
        self._data: dict = {}
        self._load()

    def _load(self):
        try:
            if os.path.isfile(self._path):
                with open(self._path, encoding='utf-8') as f:
                    self._data = json.load(f)
        except Exception:
            self._data = {}

    def _save(self):
        try:
            with open(self._path, 'w', encoding='utf-8') as f:
                json.dump(self._data, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def values(self, key):
        """Return values for key: user history (newest first) then built-in defaults."""
        entries = sorted(self._data.get(key, []),
                         key=lambda e: e.get('ts', ''), reverse=True)
        history_vals = [e['value'] for e in entries]
        result = list(history_vals)
        for v in _DEFAULTS.get(key, []):
            if v not in result:
                result.append(v)
        return result

    def record(self, key, value):
        """Mark value as used now; update timestamp if already present."""
        value = value.strip()
        if not value:
            return
        ts = datetime.now().isoformat(timespec='seconds')
        entries = self._data.setdefault(key, [])
        for e in entries:
            if e['value'] == value:
                e['ts'] = ts
                break
        else:
            entries.append({'value': value, 'ts': ts})
        self._save()

    def all_entries(self, key):
        """Return list of user-saved entry dicts for key (unsorted)."""
        return list(self._data.get(key, []))

    def add_manual(self, key, value, note=''):
        """Add or update a manual entry; preserves existing timestamp."""
        value = value.strip()
        if not value:
            return
        entries = self._data.setdefault(key, [])
        for e in entries:
            if e['value'] == value:
                if note.strip():
                    e['note'] = note.strip()
                break
        else:
            entry = {'value': value}
            if note.strip():
                entry['note'] = note.strip()
            entries.append(entry)
        self._save()

    def delete_entry(self, key, value):
        """Remove a user entry by value."""
        entries = self._data.get(key, [])
        self._data[key] = [e for e in entries if e['value'] != value]
        self._save()


_history = _HistoryStore(_HISTORY_PATH)


_FIELD_LABELS: dict[str, str] = {
    'panel_model':       'Panel Model',
    'panel_power_wp':    'Power per Panel (Wp)',
    'panels_per_string': 'Panels per String',
    'inverter_model':    'Inverter Model',
    'dc_power_kwp':      'DC Power (KWp)',
    'ac_power_kwac':     'AC Power (KWac)',
}


# ─────────────────────────────────────────────────────────────────────────────
#  GUI WIDGETS
# ─────────────────────────────────────────────────────────────────────────────

_PAD = {'padx': 6, 'pady': 3}


class _HistoryCombo(ttk.Frame):
    """Label + editable Combobox + optional unit, backed by _HistoryStore."""

    def __init__(self, parent, label, history_key, default='', unit='', width=28, **kw):
        super().__init__(parent, **kw)
        self._key = history_key
        self._app_ref = None
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=default)
        self._combo = ttk.Combobox(self, textvariable=self.var, width=width,
                                   values=_history.values(history_key))
        self._combo.pack(side='left', **_PAD)
        if unit:
            ttk.Label(self, text=unit, foreground='gray').pack(side='left')
        ttk.Button(self, text='...', width=3,
                   command=self._open_manager).pack(side='left', padx=(6, 0))

    def _open_manager(self):
        root = self.winfo_toplevel()
        def _on_close():
            if self._app_ref:
                self._app_ref._refresh_all_combos()
        _PresetManagerDialog(root, initial_key=self._key, on_close=_on_close)

    def set_app(self, app):
        self._app_ref = app

    def get(self):
        return self.var.get().strip()

    def refresh(self):
        """Reload dropdown values from the history store."""
        self._combo['values'] = _history.values(self._key)

    def record(self):
        """Save current value to history and refresh dropdown list."""
        v = self.get()
        if v:
            _history.record(self._key, v)
            self._combo['values'] = _history.values(self._key)


class _FileRow(ttk.Frame):
    """Label + Entry + Browse button for a single file path."""

    def __init__(self, parent, label, default='', save=False,
                 filetypes=None, **kw):
        super().__init__(parent, **kw)
        self._save      = save
        self._filetypes = filetypes or [('All files', '*.*')]
        ttk.Label(self, text=label, width=18, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=default)
        ttk.Entry(self, textvariable=self.var, width=52).pack(
            side='left', padx=(0, 4), fill='x', expand=True)
        ttk.Button(self, text='Browse…', width=9,
                   command=self._browse).pack(side='left')

    def _browse(self):
        cur = self.var.get()
        init_dir = os.path.dirname(cur) if cur else os.path.expanduser('~')
        if self._save:
            p = filedialog.asksaveasfilename(
                initialdir=init_dir,
                defaultextension='.dxf',
                filetypes=self._filetypes)
        else:
            p = filedialog.askopenfilename(
                initialdir=init_dir,
                filetypes=self._filetypes)
        if p:
            self.var.set(p)

    def get(self):
        return self.var.get().strip()


class _FieldRow(ttk.Frame):
    """Label + Entry + optional unit label in one grid row."""

    def __init__(self, parent, label, default='', unit='', width=20, **kw):
        super().__init__(parent, **kw)
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=str(default))
        ttk.Entry(self, textvariable=self.var, width=width).pack(side='left', **_PAD)
        if unit:
            ttk.Label(self, text=unit, foreground='gray').pack(side='left')

    def get(self):
        return self.var.get().strip()


class _PresetManagerDialog(tk.Toplevel):
    """Modal dialog to view, add, and delete preset entries for any combo field."""

    def __init__(self, parent, initial_key=None, on_close=None):
        super().__init__(parent)
        self.title("Manage Presets")
        self.geometry("720x520")
        self.minsize(580, 380)
        self.resizable(True, True)
        self._on_close = on_close
        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._close)
        self._build(initial_key)

    def _build(self, initial_key):
        # Field selector row
        top = ttk.Frame(self, padding=(12, 10, 12, 4))
        top.pack(fill='x')
        ttk.Label(top, text="Field:").pack(side='left', padx=(0, 6))
        self._field_var = tk.StringVar()
        self._keys = list(_FIELD_LABELS.keys())
        labels = [_FIELD_LABELS[k] for k in self._keys]
        self._field_cb = ttk.Combobox(top, textvariable=self._field_var,
                                      values=labels, state='readonly', width=32)
        self._field_cb.pack(side='left')
        self._field_cb.bind('<<ComboboxSelected>>', lambda _: self._refresh_tree())

        # Treeview
        mid = ttk.Frame(self, padding=(12, 4))
        mid.pack(fill='both', expand=True)
        cols = ('value', 'note', 'last_used', 'kind')
        self._tree = ttk.Treeview(mid, columns=cols, show='headings', selectmode='browse')
        self._tree.heading('value',     text='Value')
        self._tree.heading('note',      text='Note / Details')
        self._tree.heading('last_used', text='Last Used')
        self._tree.heading('kind',      text='Type')
        self._tree.column('value',     width=210, minwidth=100)
        self._tree.column('note',      width=210, minwidth=100)
        self._tree.column('last_used', width=140, minwidth=90)
        self._tree.column('kind',      width=70,  minwidth=55)
        vsb = ttk.Scrollbar(mid, orient='vertical', command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='left', fill='y')
        self._tree.bind('<<TreeviewSelect>>', self._on_select)

        # Add / Update form
        form = ttk.LabelFrame(self, text='Add / Update Entry', padding=(10, 6))
        form.pack(fill='x', padx=12, pady=(4, 0))
        r1 = ttk.Frame(form)
        r1.pack(fill='x', pady=2)
        ttk.Label(r1, text='Value:', width=8, anchor='e').pack(side='left')
        self._val_var = tk.StringVar()
        ttk.Entry(r1, textvariable=self._val_var, width=40).pack(side='left', padx=4)
        r2 = ttk.Frame(form)
        r2.pack(fill='x', pady=2)
        ttk.Label(r2, text='Note:', width=8, anchor='e').pack(side='left')
        self._note_var = tk.StringVar()
        ttk.Entry(r2, textvariable=self._note_var, width=56).pack(side='left', padx=4)
        btns = ttk.Frame(form)
        btns.pack(fill='x', pady=(6, 2))
        ttk.Button(btns, text='Add / Update',
                   command=self._add_entry).pack(side='left', padx=(0, 8))
        self._del_btn = ttk.Button(btns, text='Delete Selected',
                                   command=self._delete_entry, state='disabled')
        self._del_btn.pack(side='left')

        # Close button
        foot = ttk.Frame(self, padding=(12, 6))
        foot.pack(fill='x')
        ttk.Button(foot, text='Close', command=self._close).pack(side='right')

        # Select initial field
        if initial_key and initial_key in self._keys:
            self._field_cb.current(self._keys.index(initial_key))
        else:
            self._field_cb.current(0)
        self._refresh_tree()

    def _current_key(self):
        label = self._field_var.get()
        for k, v in _FIELD_LABELS.items():
            if v == label:
                return k
        return None

    def _refresh_tree(self):
        self._tree.delete(*self._tree.get_children())
        key = self._current_key()
        if not key:
            return
        user_entries = _history.all_entries(key)
        user_values  = {e['value'] for e in user_entries}
        for e in sorted(user_entries, key=lambda x: x.get('ts', ''), reverse=True):
            self._tree.insert('', 'end', iid=f'u|{e["value"]}',
                              values=(e['value'], e.get('note', ''),
                                      e.get('ts', ''), 'User'),
                              tags=('user',))
        for v in _DEFAULTS.get(key, []):
            if v not in user_values:
                self._tree.insert('', 'end', iid=f'd|{v}',
                                  values=(v, '', '', 'Default'),
                                  tags=('default',))
        self._tree.tag_configure('default', foreground='gray')
        self._del_btn.configure(state='disabled')

    def _on_select(self, _):
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        row = self._tree.item(iid, 'values')
        self._val_var.set(row[0])
        self._note_var.set(row[1])
        self._del_btn.configure(state='normal' if iid.startswith('u|') else 'disabled')

    def _add_entry(self):
        key   = self._current_key()
        value = self._val_var.get().strip()
        if not key or not value:
            messagebox.showwarning("Input Required", "Please enter a value.", parent=self)
            return
        _history.add_manual(key, value, self._note_var.get())
        self._val_var.set('')
        self._note_var.set('')
        self._refresh_tree()

    def _delete_entry(self):
        sel = self._tree.selection()
        if not sel or not sel[0].startswith('u|'):
            return
        value = self._tree.item(sel[0], 'values')[0]
        if messagebox.askyesno("Confirm Delete",
                               f"Delete preset:\n\n  {value}\n\nThis cannot be undone.",
                               parent=self):
            _history.delete_entry(self._current_key(), value)
            self._refresh_tree()

    def _close(self):
        self.grab_release()
        self.destroy()
        if self._on_close:
            self._on_close()


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN APP
# ─────────────────────────────────────────────────────────────────────────────

class SLDApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("A176LAB - DC Single Line Diagram Generator")
        self.geometry("860x800")
        self.minsize(700, 640)
        self.resizable(True, True)

        style = ttk.Style(self)
        for theme in ('vista', 'winnative', 'clam', 'default'):
            try:
                style.theme_use(theme)
                break
            except Exception:
                pass

        self._build_ui()
        self._set_icon()
        self.bind('<F5>', lambda _: self._on_generate())

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self._add_header()
        ttk.Separator(self, orient='horizontal').pack(fill='x', padx=8, pady=(0, 4))

        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        self._build_files_tab(nb)
        self._build_equip_tab(nb)
        self._build_run_tab(nb)

    def _add_header(self):
        header = ttk.Frame(self, padding=(10, 8, 10, 6))
        header.pack(fill='x')
        logo = _load_logo_tk((72, 72))
        if logo:
            self._logo_ref = logo
            ttk.Label(header, image=logo).pack(side='left', padx=(0, 14))
        txt = ttk.Frame(header)
        txt.pack(side='left', fill='y', pady=2)
        ttk.Label(txt, text="DC Single Line Diagram Generator",
                  font=('Segoe UI', 14, 'bold')).pack(anchor='w')
        ttk.Label(txt, text="A176 LAB  –  Think different project",
                  foreground='gray', font=('Segoe UI', 9)).pack(anchor='w')

    def _set_icon(self):
        try:
            from PIL import Image, ImageTk
            img = Image.open(_LOGO_PATH).resize((32, 32), Image.LANCZOS)
            self._icon_ref = ImageTk.PhotoImage(img)
            self.iconphoto(True, self._icon_ref)
        except Exception:
            pass

    # Files tab ----------------------------------------------------------------

    def _build_files_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Files  ')

        ttk.Label(tab, text="File Paths",
                  font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))

        self.fe_tmpl = _FileRow(
            tab, "Template DXF:",
            filetypes=[('DXF files', '*.dxf'), ('All files', '*.*')])
        self.fe_tmpl.pack(fill='x', pady=2)

        self.fe_xlsx = _FileRow(
            tab, "Excel Cable List:",
            filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')])
        self.fe_xlsx.pack(fill='x', pady=2)
        self.fe_xlsx.var.trace_add('write', self._sync_output_path)

        self.fe_out = _FileRow(
            tab, "Output DXF:", save=True,
            filetypes=[('DXF files', '*.dxf'), ('All files', '*.*')])
        self.fe_out.pack(fill='x', pady=2)

        ttk.Separator(tab).pack(fill='x', pady=12)

        hint = (
            "Template DXF  – source file containing Inverter 1.1 in Y-band 159 400 - 168 000.\n"
            "Excel File       – must contain sheet '2E802-3' with:\n"
            "   Column 1 = Inverter ID (e.g. '1.2')   "
            "Column 3 = String name   Column 4 = MPPT number\n\n"
            "Output DXF is auto-filled to match the Excel file location.\n\n"
            "Press F5 or switch to the Generate tab to run."
        )
        ttk.Label(tab, text=hint, foreground='gray',
                  wraplength=720, justify='left').pack(anchor='w')

    # Equipment tab ------------------------------------------------------------

    def _build_equip_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Equipment  ')

        # Solar Panel section
        self._section(tab, "Solar Panel")
        self.f_panel_model = _HistoryCombo(tab, "Panel Model:", 'panel_model')
        self.f_panel_model.pack(fill='x')

        self.f_panel_power = _HistoryCombo(tab, "Power per Panel:", 'panel_power_wp',
                                            default='460', unit='Wp', width=14)
        self.f_panel_power.pack(fill='x')

        self.f_panels_str = _HistoryCombo(tab, "Panels per String:", 'panels_per_string',
                                           default='20', unit='panels', width=14)
        self.f_panels_str.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=10)

        # Inverter section
        self._section(tab, "Inverter")
        self.f_inv_model = _HistoryCombo(tab, "Inverter Model:", 'inverter_model')
        self.f_inv_model.pack(fill='x')

        self.f_dc_power = _HistoryCombo(tab, "DC Power per Inverter:", 'dc_power_kwp',
                                         default='350', unit='KWp  (0 = auto-calculate)', width=14)
        self.f_dc_power.pack(fill='x')

        self.f_ac_power = _HistoryCombo(tab, "AC Power:", 'ac_power_kwac',
                                         default='320', unit='KWac', width=14)
        self.f_ac_power.pack(fill='x')

        self.f_temp = _FieldRow(tab, "Temperature Rating:", '40', unit='°C')
        self.f_temp.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=10)

        # Transformer section
        self._section(tab, "Transformer")
        self.f_tx_power = _FieldRow(tab, "Transformer Power:", '',
                                     unit='e.g. 2.5 MVA  (shown in cabin header)', width=20)
        self.f_tx_power.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=10)

        note = (
            "Auto-calculate DC power: set 'DC Power per Inverter' to 0 and fill in "
            "Panel Power + Panels per String.  The script counts strings per inverter "
            "from the Excel file and computes: strings × panels × Wp ÷ 1000."
        )
        ttk.Label(tab, text=note, foreground='gray',
                  wraplength=720, justify='left').pack(anchor='w')

        # Wire app reference so each combo's '...' button can refresh all combos
        for combo in (self.f_panel_model, self.f_panel_power, self.f_panels_str,
                      self.f_inv_model, self.f_dc_power, self.f_ac_power):
            combo.set_app(self)

    def _section(self, parent, title):
        ttk.Label(parent, text=title,
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(4, 2))

    # Generate / Log tab -------------------------------------------------------

    def _build_run_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Generate  ')

        # Top controls
        ctrl = ttk.Frame(tab)
        ctrl.pack(fill='x', pady=(0, 6))

        self.gen_btn = ttk.Button(ctrl, text='▶  Generate SLD  (F5)',
                                   command=self._on_generate)
        self.gen_btn.pack(side='left', ipadx=10, ipady=4)

        ttk.Button(ctrl, text='Clear Log',
                   command=self._clear_log).pack(side='left', padx=8)

        self.progress = ttk.Progressbar(ctrl, mode='indeterminate', length=160)
        self.progress.pack(side='left', padx=8)

        self.status_var = tk.StringVar(value='Ready.')
        ttk.Label(tab, textvariable=self.status_var,
                  foreground='gray').pack(anchor='w', pady=(0, 4))

        # Log box
        ttk.Label(tab, text='Output Log',
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(4, 2))

        self.log_box = scrolledtext.ScrolledText(
            tab, state='disabled',
            font=('Consolas', 9),
            bg='#1e1e1e', fg='#d4d4d4',
            insertbackground='white',
            relief='flat', borderwidth=1)
        self.log_box.pack(fill='both', expand=True)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _log(self, msg):
        def _do():
            self.log_box.configure(state='normal')
            self.log_box.insert('end', msg + '\n')
            self.log_box.see('end')
            self.log_box.configure(state='disabled')
        self.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state='normal')
        self.log_box.delete('1.0', 'end')
        self.log_box.configure(state='disabled')

    def _collect(self):
        return {
            'template_dxf':      self.fe_tmpl.get(),
            'xlsx_path':         self.fe_xlsx.get(),
            'output_path':       self.fe_out.get(),
            'panel_model':       self.f_panel_model.get(),
            'panel_power_wp':    self.f_panel_power.get(),
            'panels_per_string': self.f_panels_str.get(),
            'inverter_model':    self.f_inv_model.get(),
            'dc_power_kwp':      self.f_dc_power.get(),
            'ac_power_kwac':     self.f_ac_power.get(),
            'temp_rating':       self.f_temp.get(),
            'transformer_power': self.f_tx_power.get(),
        }

    def _validate(self, cfg):
        errs = []
        if not cfg.get('template_dxf'):
            errs.append("Template DXF path is required.")
        elif not os.path.isfile(cfg['template_dxf']):
            errs.append(f"Template DXF not found:\n  {cfg['template_dxf']}")
        if not cfg['xlsx_path']:
            errs.append("Excel file path is required.")
        elif not os.path.isfile(cfg['xlsx_path']):
            errs.append(f"Excel file not found:\n  {cfg['xlsx_path']}")
        if not cfg['output_path']:
            errs.append("Output DXF path is required.")
        else:
            out_dir = os.path.dirname(cfg['output_path'])
            if out_dir and not os.path.isdir(out_dir):
                errs.append(f"Output directory does not exist:\n  {out_dir}")
        for key, label in (('panel_power_wp', 'Panel Power'),
                           ('panels_per_string', 'Panels per String'),
                           ('dc_power_kwp', 'DC Power'),
                           ('ac_power_kwac', 'AC Power'),
                           ('temp_rating', 'Temperature Rating')):
            val = cfg.get(key, '')
            if val:
                try:
                    float(val)
                except ValueError:
                    errs.append(f"{label}: must be a number (got '{val}').")
        return errs

    def _sync_output_path(self, *_):
        xlsx = self.fe_xlsx.get()
        if xlsx:
            self.fe_out.var.set(os.path.splitext(xlsx)[0] + '.dxf')

    # ── Generate action ───────────────────────────────────────────────────────

    def _on_generate(self):
        cfg  = self._collect()
        errs = self._validate(cfg)
        if errs:
            messagebox.showerror("Input Error", "\n\n".join(errs))
            return

        self.gen_btn.configure(state='disabled')
        self.progress.start(12)
        self.status_var.set("Generating …")
        self._log("=" * 64)
        self._log("Starting SLD generation …")

        def worker():
            try:
                _generate(cfg, self._log)
                self.after(0, self._on_success, cfg['output_path'])
            except Exception as ex:
                self.after(0, self._on_error, str(ex))

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self, path):
        self.progress.stop()
        self.gen_btn.configure(state='normal')
        self.status_var.set("Done — file saved.")
        for combo in (self.f_panel_model, self.f_panel_power, self.f_panels_str,
                      self.f_inv_model, self.f_dc_power, self.f_ac_power):
            combo.record()
        messagebox.showinfo("Success",
                            f"SLD generated successfully!\n\n{path}")

    def _refresh_all_combos(self):
        for combo in (self.f_panel_model, self.f_panel_power, self.f_panels_str,
                      self.f_inv_model, self.f_dc_power, self.f_ac_power):
            combo.refresh()

    def _on_error(self, msg):
        self.progress.stop()
        self.gen_btn.configure(state='normal')
        self.status_var.set("Error — see log.")
        self._log(f"\nERROR: {msg}")
        messagebox.showerror("Generation Error", msg)


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app = SLDApp()
    app.mainloop()
