# -*- coding: utf-8 -*-
"""
smart_sld_generator.py  -  Smart Headless CLI & GUI Generator for DC Single Line Diagrams
=======================================================================================
Dynamically generates a DC Single Line Diagram (DXF) from an Excel cable schedule,
automatically grouping by Cabin, resizing inverter geometry (stretching/shrinking)
on a per-inverter basis, and creating correctly sized A3 Paper Space layouts.

GUI Launch:
  python smart_sld_generator.py

CLI Launch:
  python smart_sld_generator.py --excel <excel_path> --template <template_path> [options]
"""

import os
import sys
import re
import argparse
import json
import threading
from collections import defaultdict

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, scrolledtext, messagebox
    HAS_TK = True
except ImportError:
    HAS_TK = False

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS & DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────

# Original steps from the template geometry
COL_STEP = 11_740
ROW_STEP = 10_200

# Spacing default calculations
COL_SPACING_DEFAULT = int(COL_STEP * 1.22)
ROW_SPACING_DEFAULT = int(ROW_STEP * 1.18)

# Y-proximity for matching port labels to string-label slots.
PORT_Y_TOL = 400

# Template bottom-right placeholder rectangle filter.
PLACEHOLDER_X_MIN = 23_500
PLACEHOLDER_MIN_W = 500

# MTEXT text-box width override so long string labels never wrap
STRING_LABEL_MIN_WIDTH = 4_500

PORT_RE = re.compile(r'^\d+-\d+$')
STRING_RE = re.compile(r'String \d+\.\d+\.\d+')

# Inverter model → (DC KWp, AC KWac) for autocompletion autofill
_INVERTER_POWERS = {
    'Sungrow SG100HX': ('100', '100'),
    'Sungrow SG125HX': ('125', '125'),
    'Sungrow SG250HX': ('250', '250'),
    'Sungrow SG350HX': ('350', '350'),
    'Sungrow SG500HX': ('500', '500'),
    'Huawei SUN2000-100KTL-M3': ('100', '100'),
    'Huawei SUN2000-215KTL-H3': ('215', '160'),
    'Huawei SUN2000-275KTL-H1': ('275', '220'),
    'Huawei SUN2000-330KTL': ('330', '275'),
    'Huawei SUN2000-450KTL-H1': ('450', '400'),
    'ABB PVS-120-TL': ('120', '120'),
    'ABB PVS-250-TL': ('250', '220'),
    'ABB PVS-350-TL': ('350', '280'),
    'ABB PVS-500-TL': ('500', '400'),
    'SMA Sunny Tripower Core1 25': ('25', '25'),
    'SMA Sunny Tripower Core2 150': ('150', '150'),
    'SMA Sunny Tripower 25000TL': ('25', '25'),
    'SMA Sunny Tripower 60000TL': ('60', '60'),
    'SMA Sunny Tripower 100000TL': ('100', '100'),
    'Fronius Symo GEN24 25.0 Plus': ('25', '25'),
    'Fronius Symo GEN24 50.0 Plus': ('50', '50'),
    'Fronius Symo GEN24 60.0 Plus': ('60', '60'),
    'Fronius Symo GEN24 100.0 Plus': ('100', '100'),
    'KACO blueplanet 100.0 TL3': ('100', '100'),
    'KACO blueplanet 125.0 TL3': ('125', '125'),
    'KACO blueplanet 250.0 TL3': ('250', '250'),
    'Growatt 50000MT': ('50', '50'),
    'Growatt 60000MT': ('60', '60'),
    'Growatt 100000MT': ('100', '100'),
}

_PANELS_PRESETS = [
    'JA Solar JAM72D42-625/LB',
    'JA Solar JAM72S20-460/MR',
    'Longi Solar LR5-72HBD-580M',
    'Longi Solar LR5-72HBD-545M',
    'Canadian Solar HiKu7 CS7N-655MB',
    'Jinko Solar JKM660M-78HL4-V',
    'Trina Solar TSM-670NEG21C.20',
    'REC Alpha Pure-R 430AA',
]

# ─────────────────────────────────────────────────────────────────────────────
#  DRAWING EXTRACTION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _strip_mtext_fmt(txt):
    """Strip AutoCAD inline MTEXT formatting codes to get plain text."""
    txt = re.sub(r'\\[A-Za-z][^;]*;', '', txt)
    return re.sub(r'\{[^}]*\}', '', txt).strip()


def _ent_y(e):
    """Return a representative Y coordinate for an entity, or None."""
    try:
        t = e.dxftype()
        if t in ('TEXT', 'MTEXT', 'INSERT'):  return e.dxf.insert.y
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return e.dxf.center.y
        if t == 'LINE':                        return e.dxf.start.y
        if t == 'LWPOLYLINE':
            pts = list(e.get_points())
            return pts[0][1] if pts else None
    except Exception:
        pass
    return None


def _ent_y_dict(d):
    """Return a representative Y coordinate for a dictionary entity representation."""
    t = d['type']
    if t in ('TEXT', 'MTEXT', 'INSERT'):  return d.get('y') or d.get('iy')
    if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return d.get('cy')
    if t == 'LINE':                        return d.get('sy')
    if t == 'LWPOLYLINE':
        pts = d.get('pts', [])
        return pts[0][1] if pts else None
    if t == 'POLYLINE':
        pts = d.get('pts3d', [])
        return pts[0][1] if pts else None
    return None


def _is_cable_line(d, sl_y):
    """True if d is the horizontal connection line at the Y-level of sl_y."""
    t = d['type']
    if t == 'LINE':
        length = abs(d['sx'] - d['ex'])
        is_horiz = abs(d['sy'] - d['ey']) < 5
        is_near_y = abs(d['sy'] - sl_y) < 80
        return is_horiz and is_near_y and length > 200
    elif t == 'LWPOLYLINE':
        xs = [p[0] for p in d['pts']]
        ys = [p[1] for p in d['pts']]
        if not xs or not ys:
            return False
        length = max(xs) - min(xs)
        is_horiz = (max(ys) - min(ys)) < 5
        is_near_y = abs(ys[0] - sl_y) < 80
        return is_horiz and is_near_y and length > 200
    elif t == 'POLYLINE':
        xs = [p[0] for p in d['pts3d']]
        ys = [p[1] for p in d['pts3d']]
        if not xs or not ys:
            return False
        length = max(xs) - min(xs)
        is_horiz = (max(ys) - min(ys)) < 5
        is_near_y = abs(ys[0] - sl_y) < 80
        return is_horiz and is_near_y and length > 200
    return False


def _common_attrs(e):
    d = {}
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if e.dxf.hasattr(a):
                d[a] = getattr(e.dxf, a)
        except Exception:
            pass
    return d


def _extract_entity(e):
    """Return a plain dict with all drawing data, or None for unsupported types."""
    t = e.dxftype()
    d = {'type': t}
    d.update(_common_attrs(e))
    try:
        if t == 'LWPOLYLINE':
            d['pts']    = [list(p) for p in e.get_points()]
            d['closed'] = e.closed
            if e.dxf.hasattr('const_width'):
                d['const_width'] = e.dxf.const_width
        elif t == 'MTEXT':
            pos = e.dxf.insert
            d.update({'x': pos.x, 'y': pos.y, 'text': e.text})
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'style'):
                try:
                    if e.dxf.hasattr(a):
                        d[a] = getattr(e.dxf, a)
                except Exception:
                    pass
        elif t == 'ARC':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'radius': e.dxf.radius,
                      'start_angle': e.dxf.start_angle,
                      'end_angle':   e.dxf.end_angle})
        elif t == 'CIRCLE':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z, 'radius': e.dxf.radius})
        elif t == 'LINE':
            s, en = e.dxf.start, e.dxf.end
            d.update({'sx': s.x, 'sy': s.y, 'sz': s.z,
                      'ex': en.x, 'ey': en.y, 'ez': en.z})
        elif t == 'ELLIPSE':
            c, ma = e.dxf.center, e.dxf.major_axis
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'major_axis': [ma.x, ma.y, ma.z],
                      'ratio': e.dxf.ratio,
                      'start_param': e.dxf.start_param,
                      'end_param':   e.dxf.end_param})
        elif t == 'INSERT':
            ins = e.dxf.insert
            d.update({'name': e.dxf.name,
                      'ix': ins.x, 'iy': ins.y,
                      'iz': ins.z if hasattr(ins, 'z') else 0.0})
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d:
                        setattr(e.dxf, a, d[a])
                except Exception:
                    pass
        elif t == 'POLYLINE':
            d['pts3d'] = [[v.dxf.location.x, v.dxf.location.y, v.dxf.location.z]
                          for v in e.vertices]
        else:
            return None
    except Exception as ex:
        print(f"  [warn] extract {t}: {ex}", file=sys.stderr)
        return None
    return d


def _entity_min_x(d):
    """Leftmost X coordinate for an extracted entity dict."""
    t = d['type']
    try:
        if t == 'LWPOLYLINE':                  return min(p[0] for p in d['pts'])
        if t == 'MTEXT':                       return d['x']
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return d['cx'] - d.get('radius', 0)
        if t == 'LINE':                        return min(d['sx'], d['ex'])
        if t == 'INSERT':                      return d['ix']
        if t == 'POLYLINE':                    return min(p[0] for p in d['pts3d'])
    except Exception:
        pass
    return 0


def _is_placeholder_rect(d):
    """True if d is the template's bottom-right placeholder rectangle artifact."""
    if d.get('type') != 'LWPOLYLINE' or not d.get('closed'):
        return False
    pts = d.get('pts', [])
    if len(pts) != 4:
        return False
    xs = [p[0] for p in pts]
    return min(xs) > PLACEHOLDER_X_MIN and (max(xs) - min(xs)) > PLACEHOLDER_MIN_W


def _classify_mtext(txt):
    c = _strip_mtext_fmt(txt).replace('\n', ' ')
    if re.search(r'INVERTER 1\.1', txt, re.I) and 'P=' in txt:
        return 'title'
    if re.search(r'Cabin Tx\.\d+.*Inverter', txt):
        return 'cabin_label'
    if re.match(r'^CABIN \d+$', c):
        return 'cabin_header'
    if re.search(r'PV modules?', txt, re.I):
        return 'panel_count'
    if STRING_RE.search(txt) or c == 'reserve':
        return 'string_label'
    return 'fixed'


def _update_panel_count_label(text_val, panels_per_string, panel_model):
    updated = text_val
    if panels_per_string > 0:
        new = re.sub(
            r'\d+(\s*PV modules?)',
            lambda m: f'{panels_per_string}{m.group(1)}',
            updated, flags=re.I)
        if new != updated:
            updated = new
        else:
            # Fallback: MTEXT codes (e.g. \P paragraph break, closing brace) may sit
            # between the panel count number and the "PV modules" text.
            # Find "PV modules" then replace the last number that precedes it.
            pv_m = re.search(r'PV modules?', updated, re.I)
            if pv_m:
                before = updated[:pv_m.start()]
                num_matches = list(re.finditer(r'\d+', before))
                if num_matches:
                    last = num_matches[-1]
                    updated = (before[:last.start()] + str(panels_per_string)
                               + before[last.end():] + updated[pv_m.start():])
    
    if panel_model:
        idx_series = updated.lower().find('series')
        if idx_series != -1:
            dash_idx = updated.find('-', idx_series)
            if dash_idx != -1:
                base = updated[:dash_idx].strip()
                updated = f"{base} - {panel_model}"
            else:
                updated = f"{updated.strip()} - {panel_model}"
        else:
            if '-' in updated:
                base = updated.split('-')[0].strip()
                updated = f"{base} - {panel_model}"
            else:
                updated = f"{updated.strip()} - {panel_model}"
    return updated


def _mtext_attribs(d, dx, dy):
    """Build a dxfattribs dict for placing an MTEXT entity at an offset."""
    att = {'insert': (d['x'] + dx, d['y'] + dy)}
    for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
              'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
        if a in d:
            att[a] = d[a]
    return att


def _apply_common(ne, d):
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if a in d:
                setattr(ne.dxf, a, d[a])
        except Exception:
            pass


def _place_entity(layout, d, dx, dy):
    """Stamp one entity from its dict, shifted by (dx, dy)."""
    t = d['type']
    try:
        if t == 'LWPOLYLINE':
            pts = [(p[0] + dx, p[1] + dy) + tuple(p[2:]) for p in d['pts']]
            ne  = layout.add_lwpolyline(pts)
            ne.closed = d.get('closed', False)
            if 'const_width' in d:
                ne.dxf.const_width = d['const_width']
            _apply_common(ne, d)
        elif t == 'MTEXT':
            ne = layout.add_mtext(d['text'], dxfattribs=_mtext_attribs(d, dx, dy))
            _apply_common(ne, d)
        elif t == 'ARC':
            ne = layout.add_arc(
                (d['cx'] + dx, d['cy'] + dy, d['cz']),
                d['radius'], d['start_angle'], d['end_angle'])
            _apply_common(ne, d)
        elif t == 'CIRCLE':
            ne = layout.add_circle(
                (d['cx'] + dx, d['cy'] + dy, d['cz']), d['radius'])
            _apply_common(ne, d)
        elif t == 'LINE':
            ne = layout.add_line(
                (d['sx'] + dx, d['sy'] + dy, d['sz']),
                (d['ex'] + dx, d['ey'] + dy, d['ez']))
            _apply_common(ne, d)
        elif t == 'ELLIPSE':
            ne = layout.add_ellipse(
                center=(d['cx'] + dx, d['cy'] + dy, d['cz']),
                major_axis=d['major_axis'], ratio=d['ratio'],
                start_param=d['start_param'], end_param=d['end_param'])
            _apply_common(ne, d)
        elif t == 'INSERT':
            ne = layout.add_blockref(
                d['name'], (d['ix'] + dx, d['iy'] + dy, d['iz']))
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d:
                        setattr(ne.dxf, a, d[a])
                except Exception:
                    pass
            _apply_common(ne, d)
        elif t == 'POLYLINE':
            pts = [(p[0] + dx, p[1] + dy, p[2]) for p in d['pts3d']]
            if pts:
                ne = layout.add_polyline3d(pts)
                _apply_common(ne, d)
    except Exception as ex:
        print(f"  [warn] place {t}: {ex}", file=sys.stderr)


def _place_entity_stretched(layout, d, dx, dy, split_y, extra_h):
    """Stamps an entity, shifting any coordinate below split_y by extra_h (downward if positive, upward if negative)."""
    if extra_h == 0:
        _place_entity(layout, d, dx, dy)
        return

    def _sy(y):
        return y - extra_h if y < split_y else y

    t = d['type']
    try:
        if t == 'LWPOLYLINE':
            pts = [(p[0] + dx, _sy(p[1]) + dy) + tuple(p[2:]) for p in d['pts']]
            ne  = layout.add_lwpolyline(pts)
            ne.closed = d.get('closed', False)
            if 'const_width' in d:
                ne.dxf.const_width = d['const_width']
            _apply_common(ne, d)
        elif t == 'MTEXT':
            ne = layout.add_mtext(d['text'],
                                  dxfattribs=_mtext_attribs(dict(d, y=_sy(d['y'])), dx, dy))
            _apply_common(ne, d)
        elif t == 'ARC':
            ne = layout.add_arc(
                (d['cx'] + dx, _sy(d['cy']) + dy, d['cz']),
                d['radius'], d['start_angle'], d['end_angle'])
            _apply_common(ne, d)
        elif t == 'CIRCLE':
            ne = layout.add_circle(
                (d['cx'] + dx, _sy(d['cy']) + dy, d['cz']), d['radius'])
            _apply_common(ne, d)
        elif t == 'LINE':
            ne = layout.add_line(
                (d['sx'] + dx, _sy(d['sy']) + dy, d['sz']),
                (d['ex'] + dx, _sy(d['ey']) + dy, d['ez']))
            _apply_common(ne, d)
        elif t == 'ELLIPSE':
            ne = layout.add_ellipse(
                center=(d['cx'] + dx, _sy(d['cy']) + dy, d['cz']),
                major_axis=d['major_axis'], ratio=d['ratio'],
                start_param=d['start_param'], end_param=d['end_param'])
            _apply_common(ne, d)
        elif t == 'INSERT':
            ne = layout.add_blockref(
                d['name'], (d['ix'] + dx, _sy(d['iy']) + dy, d['iz']))
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d:
                        setattr(ne.dxf, a, d[a])
                except Exception:
                    pass
            _apply_common(ne, d)
        elif t == 'POLYLINE':
            pts = [(p[0] + dx, _sy(p[1]) + dy, p[2]) for p in d['pts3d']]
            if pts:
                ne = layout.add_polyline3d(pts)
                _apply_common(ne, d)
    except Exception as ex:
        print(f"  [warn] place_stretched {t}: {ex}", file=sys.stderr)


def _place_mtext(layout, d, dx, dy, text):
    """Stamp an MTEXT at (dx, dy) offset with a custom text string."""
    try:
        ne = layout.add_mtext(text, dxfattribs=_mtext_attribs(d, dx, dy))
        _apply_common(ne, d)
    except Exception as ex:
        print(f"  [warn] place_mtext: {ex}", file=sys.stderr)

# ─────────────────────────────────────────────────────────────────────────────
#  GENERATION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def generate(cfg, log_cb=print):
    """Run the headless SLD generation with the configuration dictionary."""
    def print(*args, **kwargs):
        sep = kwargs.get('sep', ' ')
        msg = sep.join(str(a) for a in args)
        log_cb(msg)

    try:
        import ezdxf as _ez
    except ImportError:
        print("[ERROR] ezdxf is not installed. Run: pip install ezdxf")
        sys.exit(1)
    try:
        import openpyxl as _xl
    except ImportError:
        print("[ERROR] openpyxl is not installed. Run: pip install openpyxl")
        sys.exit(1)

    TEMPLATE_DXF = cfg['template_dxf']
    XLSX_PATH    = cfg['xlsx_path']
    OUTPUT_PATH  = cfg['output_path']

    panel_model       = cfg.get('panel_model', '').strip()
    panels_per_string = int(float(cfg.get('panels_per_string') or 0))
    # Optional per-watt panel model override: "720=Model A; 725=Model B". Lets a single
    # inverter that mixes panel power classes label each string with the right model.
    panel_model_map = {}
    for _pair in re.split(r'[;\n]', cfg.get('panel_model_map', '') or ''):
        if '=' in _pair:
            _w, _mdl = _pair.split('=', 1)
            _w, _mdl = _w.strip(), _mdl.strip()
            if _w and _mdl:
                try:
                    panel_model_map[int(float(_w))] = _mdl
                except ValueError:
                    pass
    if panel_model_map:
        print(f"Panel model map (Wp -> model): {panel_model_map}")
    inverter_model    = cfg.get('inverter_model', '').strip()
    dc_power_kwp      = float(cfg.get('dc_power_kwp') or 0)
    ac_power_kwac     = float(cfg.get('ac_power_kwac') or 0)
    ac_power_30c      = float(cfg.get('ac_power_30c') or 0)
    max_ac_current    = cfg.get('max_ac_current', '').strip()
    max_pv_current_per_mppt    = cfg.get('max_pv_current_per_mppt', '').strip()
    max_dc_sc_current_per_mppt = cfg.get('max_dc_sc_current_per_mppt', '').strip()
    mppt_voltage_range = (cfg.get('mppt_voltage_range', '') or "500...1'500V").strip()
    max_vdc            = (cfg.get('max_vdc', '') or "1'500V").strip()
    temp_rating       = float(cfg.get('temp_rating') or 40)
    transformer_power = cfg.get('transformer_power', '').strip()
    show_cable_info     = cfg.get('show_cable_info', False)
    hide_string_details = cfg.get('hide_string_details', False)
    show_annot_circle   = cfg.get('show_annot_circle', False)   # large red annotation circle
    
    col_spacing       = float(cfg.get('col_spacing') or COL_SPACING_DEFAULT)
    row_spacing       = float(cfg.get('row_spacing') or ROW_SPACING_DEFAULT)
    circle_radius_cfg = float(cfg.get('circle_radius') or 24.59)
    text_size_cfg     = float(cfg.get('text_size') or 60.44)
    heavy_section     = cfg.get('heavy_section', '1x10').strip()
    heavy_linetype    = cfg.get('heavy_linetype', 'TRATTEGGIATA').strip()
    heavy_color       = int(float(cfg.get('heavy_color') or 40))
    heavy_layer       = cfg.get('heavy_layer', 'TRATTEGGIATA').strip()
    # Number of string rows drawn per MPPT (template default = 2 inputs/MPPT) and how
    # many MPPTs are grouped under each DC-switch label.
    strings_per_mppt  = max(1, int(float(cfg.get('strings_per_mppt') or 2)))
    mppts_per_switch  = max(1, int(float(cfg.get('mppts_per_switch') or 4)))

    # Per-MPPT slot layout: {"1":[1,2,3,4,"R"], "2":[1,2,"R",3,"R"], ...}
    # Integer = active slot (priority order), "R" = always reserve, null = auto-fill.
    mppt_layout: dict = {}
    _layout_raw = (cfg.get('mppt_layout') or '').strip()
    if _layout_raw:
        try:
            _parsed = json.loads(_layout_raw)
            mppt_layout = {int(k): list(v) for k, v in _parsed.items()}
            print(f"MPPT layout loaded: {len(mppt_layout)} MPPT row(s) configured")
        except Exception as _e:
            print(f"[WARNING] Could not parse mppt_layout JSON: {_e}")

    # ── 1. Read Excel sheets ──────────────────────────────────────────────────
    print(f"Opening Excel workbook: {XLSX_PATH}")
    wb = _xl.load_workbook(XLSX_PATH, data_only=True)
    
    # Locate sheet 'Inverter To String' (case-insensitive)
    its_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == 'inverter to string':
            its_sheet_name = name
            break
            
    its_data = {}
    if its_sheet_name:
        print(f"Loading string routing lengths from sheet '{its_sheet_name}'...")
        ws_its = wb[its_sheet_name]
        for r in range(2, ws_its.max_row + 1):
            s_id = ws_its.cell(r, 1).value
            l_p = ws_its.cell(r, 2).value
            l_m = ws_its.cell(r, 3).value
            tbl = ws_its.cell(r, 5).value
            if s_id:
                its_data[str(s_id).strip()] = {
                    'l_plus': float(l_p) if l_p is not None else 0.0,
                    'l_minus': float(l_m) if l_m is not None else 0.0,
                    'table_num': str(tbl).strip() if tbl else 'Default'
                }
        print(f"  Loaded {len(its_data)} routing records.")
    else:
        print("[WARNING] Sheet 'Inverter To String' not found. Routing lengths will default to 0.0.")

    # Locate sheet '2E802-3' (case-insensitive)
    master_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == '2e802-3':
            master_sheet_name = name
            break
            
    if not master_sheet_name:
        print(f"[ERROR] Master sheet '2E802-3' not found in workbook.")
        sys.exit(1)
        
    ws = wb[master_sheet_name]
    excel = {}
    cur = None

    # Dynamic column mapping based on header detection
    col_inverter = 1
    col_str_name = 3
    col_mppt = 4
    col_pos = 12
    col_wp = 21   # default = column U (panel power in Wp)
    col_sec = 25

    header_row = 30
    found_headers = False
    for r_idx in (30, 29, 28, 31, 32, 27, 26):
        row_vals = [str(ws.cell(row=r_idx, column=c).value or '').strip().lower() for c in range(1, 35)]
        if any('string name' in val for val in row_vals):
            header_row = r_idx
            found_headers = True
            break

    if found_headers:
        print(f"Found headers row at: {header_row}")
        # Watt-peak detection is two-tier: a "strong" match (an explicit module/panel
        # power header) wins outright; a "weak" match (a bare 'wp'/'potenza') is only
        # used if no strong column was found. This stops 'String Power (kWp)' from
        # hijacking the panel watt-peak via the 'wp' substring inside 'kwp'.
        _wp_strong = False
        for c in range(1, 35):
            val = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
            if val == 'inverter':
                col_inverter = c
            elif val == 'string name':
                col_str_name = c
            elif val == 'mppt':
                col_mppt = c
            elif 'posizione stringa' in val:
                col_pos = c
            elif any(kw in val for kw in ('module type', 'module power', 'watt peak')):
                col_wp = c
                _wp_strong = True
            elif not _wp_strong and ('potenza' in val or re.search(r'(?<![a-z])wp(?![a-z])', val)):
                col_wp = c
            elif 'section' in val:
                col_sec = c
    else:
        print("[WARNING] Header row containing 'string name' not found. Using default column offsets.")

    print(f"Mapped headers -> Inverter: col {col_inverter}, StringName: col {col_str_name}, MPPT: col {col_mppt}, Position: col {col_pos}, Wp: col {col_wp}, Section: col {col_sec}")

    print("Parsing master cable schedule...")
    for r in range(header_row + 1, ws.max_row + 1):
        inv = ws.cell(row=r, column=col_inverter).value
        sname = ws.cell(row=r, column=col_str_name).value
        mppt = ws.cell(row=r, column=col_mppt).value

        if inv is not None and '.' in str(inv):
            try:
                parts = str(inv).split('.')
                t = int(parts[0])
                v = int(parts[1])
                cur = (t, v)
                excel.setdefault(cur, defaultdict(list))
            except Exception:
                pass

        if sname and cur and mppt:
            sname_str = str(sname).strip()
            route = its_data.get(sname_str, {'l_plus': 0.0, 'l_minus': 0.0, 'table_num': 'Default'})

            # Read Section
            section = ws.cell(row=r, column=col_sec).value
            section_str = str(section).strip() if section else '1x6'

            # Read Tracker position
            tracker_pos = ws.cell(row=r, column=col_pos).value
            if not tracker_pos:
                tracker_pos = route['table_num']
            if not tracker_pos or str(tracker_pos).strip() == 'No piling information':
                tracker_pos = 'Default'

            # Wp modules
            wp = 0
            module_wp_raw = ws.cell(row=r, column=col_wp).value
            try:
                wp = int(float(module_wp_raw))
            except Exception:
                pass

            try:
                mppt_val = int(float(mppt))
                excel[cur][mppt_val].append({
                    'name': sname_str,
                    'wp': wp,
                    'l_plus': route['l_plus'],
                    'l_minus': route['l_minus'],
                    'section': section_str,
                    'tracker_pos': str(tracker_pos).strip()
                })
            except Exception:
                pass

    inv_list = sorted(excel.keys())
    transformers = {}
    for (T, I) in inv_list:
        transformers.setdefault(T, []).append(I)
    for T in transformers:
        transformers[T] = sorted(transformers[T])
    transformer_list = sorted(transformers.keys())

    print(f"Parsed {len(inv_list)} inverters across {len(transformer_list)} Cabin Transformer(s):")
    for T in transformer_list:
        ii = transformers[T]
        mppt_counts = [len(excel.get((T, i), {})) for i in ii]
        string_counts = [sum(len(v) for v in excel.get((T, i), {}).values()) for i in ii]
        print(f"  Cabin {T}: {len(ii)} inverters (MPPTs: {min(mppt_counts)}-{max(mppt_counts)}, Strings: {min(string_counts)}-{max(string_counts)})")

    # ── 2. Load template DXF ──────────────────────────────────────────────────
    print(f"Loading template DXF: {TEMPLATE_DXF}")
    doc = _ez.readfile(TEMPLATE_DXF)

    # Read the template's original module count BEFORE any block modifications.
    # The "N PV modules in series" text is inside a block definition (not the
    # modelspace), so it must be read here - after the block update it shows the
    # new count and can no longer be used to recover the original.
    _template_panels_orig = 0
    for _blk in doc.blocks:
        for _e in _blk:
            if _e.dxftype() in ('TEXT', 'MTEXT'):
                _raw = _e.dxf.text if _e.dxftype() == 'TEXT' else _e.text
                if re.search(r'PV modules?', _raw, re.I):
                    _tm = re.search(r'(\d+)', _raw) or re.search(r'(\d+)', _strip_mtext_fmt(_raw))
                    if _tm:
                        _template_panels_orig = int(_tm.group(1))
        if _template_panels_orig > 0:
            break
    if _template_panels_orig == 0:
        for _e in doc.modelspace():
            if _e.dxftype() == 'MTEXT' and re.search(r'PV modules?', _e.text, re.I):
                _tm = re.search(r'(\d+)', _e.text) or re.search(r'(\d+)', _strip_mtext_fmt(_e.text))
                if _tm:
                    _template_panels_orig = int(_tm.group(1))
                    break
    if _template_panels_orig > 0:
        print(f"  Template module count detected: {_template_panels_orig}")

    # Update block definitions with the custom panel model/count
    replaced_blocks_count = 0
    for blk in doc.blocks:
        for e in blk:
            if e.dxftype() == 'MTEXT':
                text_val = e.text
                if re.search(r'PV modules?', text_val, re.I):
                    e.text = _update_panel_count_label(text_val, panels_per_string, panel_model)
                    replaced_blocks_count += 1
            elif e.dxftype() == 'TEXT':
                text_val = e.dxf.text
                if re.search(r'PV modules?', text_val, re.I):
                    e.dxf.text = _update_panel_count_label(text_val, panels_per_string, panel_model)
                    replaced_blocks_count += 1
    if replaced_blocks_count > 0:
        print(f"  Updated {replaced_blocks_count} panel count labels inside block definitions.")

    msp = doc.modelspace()

    # Ensure linetype & layer are defined
    if heavy_linetype not in doc.linetypes:
        try:
            doc.linetypes.new(heavy_linetype, dxfattribs={
                'description': f'Custom dash linetype {heavy_linetype}',
                'pattern': [20.0, -10.0]
            })
            print(f"  Defined missing linetype '{heavy_linetype}' in DXF output")
        except Exception as ex:
            print(f"  [WARNING] Could not define linetype '{heavy_linetype}': {ex}")

    if heavy_layer not in doc.layers:
        try:
            doc.layers.new(heavy_layer, dxfattribs={'color': heavy_color, 'linetype': heavy_linetype})
            print(f"  Created missing layer '{heavy_layer}' in DXF output")
        except Exception as ex:
            print(f"  [WARNING] Could not define layer '{heavy_layer}': {ex}")

    # Auto-detect base template Y-band using "INVERTER 1.1" text anchor
    anchor_y = None
    for _e in msp:
        if _e.dxftype() == 'MTEXT':
            if re.search(r'INVERTER\s+1\.1\b', _e.text, re.I) and 'P=' in _e.text:
                anchor_y = _e.dxf.insert.y
                break

    if anchor_y is not None:
        tmpl_y_min = anchor_y - 10000
        tmpl_y_max = anchor_y + 5000
        print(f"Template anchor detected at Y={anchor_y:.0f}. Window band: {tmpl_y_min:.0f} to {tmpl_y_max:.0f}")
    else:
        tmpl_y_min, tmpl_y_max = 159_400, 168_000
        print(f"[WARNING] 'INVERTER 1.1' not found. Defaulting window band to {tmpl_y_min} - {tmpl_y_max}")

    # Extract all template column slice entities
    raw_ents = [e for e in msp if (y := _ent_y(e)) is not None and tmpl_y_min <= y <= tmpl_y_max]
    all_dicts = [d for e in raw_ents if (d := _extract_entity(e)) is not None]

    # Find the leftmost column coordinates to crop the template slice
    xs = []
    for d in all_dicts:
        t = d['type']
        try:
            if t == 'LWPOLYLINE':                  xs += [p[0] for p in d['pts']]
            elif t == 'MTEXT':                     xs.append(d['x'])
            elif t in ('ARC', 'CIRCLE', 'ELLIPSE'): xs.append(d['cx'])
            elif t == 'LINE':                      xs += [d['sx'], d['ex']]
            elif t == 'INSERT':                    xs.append(d['ix'])
            elif t == 'POLYLINE':                  xs += [p[0] for p in d['pts3d']]
        except Exception:
            pass
    xmin = min(xs) if xs else 0
    xcut = xmin + COL_STEP

    # Template column entities crop.
    # Apply X filter to both geometry and MTEXT so that multi-column templates
    # (e.g. a completed 9-inverter SLD used as the reference) do not bleed
    # texts from columns 2-N into the single-frame slice. A 20% padding over
    # COL_STEP ensures right-aligned text anchored near the column boundary
    # is still included.
    _mtext_xcut = xmin + COL_STEP * 1.20
    def _in_column(d):
        if d['type'] == 'MTEXT':
            return d.get('x', 0) <= _mtext_xcut
        return _entity_min_x(d) <= xcut

    tmpl = [d for d in all_dicts if _in_column(d) and not _is_placeholder_rect(d)]
    print(f"Extracted template geometry: {len(tmpl)} entities in column slice")


    # Classify MTEXT elements inside template
    tmpl_texts = []
    for d in tmpl:
        if d['type'] == 'MTEXT':
            d['cls'] = _classify_mtext(d['text'])
            tmpl_texts.append(d)

    # Use the original module count captured before block definitions were updated.
    # Fallback to tmpl_texts only if the panel_count entity was in the modelspace.
    _template_panels = _template_panels_orig
    if _template_panels == 0:
        _pc_d = next((m for m in tmpl_texts if m['cls'] == 'panel_count'), None)
        if _pc_d:
            for _src in (_pc_d['text'], _strip_mtext_fmt(_pc_d['text'])):
                _tm = re.search(r'(\d+)', _src)
                if _tm:
                    _template_panels = int(_tm.group(1))
                    break

    # Renumber right-side module box integer labels when panels_per_string differs
    # from the template default.  Box numbers may live in modelspace MTEXT *or*
    # inside INSERT block definitions (TEXT/MTEXT) - both are updated here.
    # Only the right-side group (beyond the consecutive left run 1..N_left) is
    # shifted so the last visible box always shows panels_per_string.
    # Example: template=28, target=26, delta=2 -> "28"->"26", "27"->"25", "22"->"20"
    if panels_per_string > 0 and _template_panels > 0 and _template_panels != panels_per_string:
        _delta = _template_panels - panels_per_string
        # Collect all integer labels in range [1, _template_panels] ONLY from
        # modelspace MTEXT to find the left/right boundary.
        _all_int_lbl = set()
        for _d in tmpl_texts:
            if _d.get('cls') == 'fixed':
                _s = _strip_mtext_fmt(_d['text']).strip()
                if _s.isdigit() and 1 <= int(_s) <= _template_panels:
                    _all_int_lbl.add(int(_s))
        # Left group = consecutive run from 1; first gap = start of right group
        _left_max = 0
        for _n in sorted(_all_int_lbl):
            if _n == _left_max + 1:
                _left_max = _n
            else:
                break

        if _delta != 0 and _left_max > 0:
            # Shift right-group MTEXT labels extracted into tmpl (modelspace)
            for _d in tmpl_texts:
                if _d.get('cls') == 'fixed':
                    _s = _strip_mtext_fmt(_d['text']).strip()
                    if _s.isdigit() and _left_max < int(_s) <= _template_panels:
                        _new_n = int(_s) - _delta
                        _d['text'] = re.sub(r'(?<!\d)' + re.escape(_s) + r'(?!\d)', str(_new_n), _d['text'])
            # Shift right-group TEXT/MTEXT labels inside block definitions
            _blk_count = 0
            for _blk in doc.blocks:
                for _e in _blk:
                    if _e.dxftype() in ('TEXT', 'MTEXT'):
                        _raw = _e.dxf.text if _e.dxftype() == 'TEXT' else _e.text
                        _rs = _raw.strip()
                        if _rs.isdigit() and _left_max < int(_rs) <= _template_panels:
                            _new_n = int(_rs) - _delta
                            if _e.dxftype() == 'TEXT':
                                _e.dxf.text = str(_new_n)
                            else:
                                _e.text = re.sub(r'(?<!\d)' + re.escape(_rs) + r'(?!\d)', str(_new_n), _e.text)
                            _blk_count += 1
            print(f"  Renumbered right-side box labels by {-_delta:+d} "
                  f"(template={_template_panels}, target={panels_per_string}, "
                  f"left_group=1-{_left_max}, block_entities={_blk_count})")

    # ── 2b. Build the per-string-row layout model from the template ───────────
    # The template models a 16-MPPT inverter with exactly 2 inputs (ports) per MPPT.
    # Real inverters can have a different (usually larger) number of strings per MPPT,
    # so we rebuild the connection region procedurally: capture the geometry of one
    # string "row" (terminal circle, cables, port/string labels, small symbols), the
    # per-MPPT combiner-bracket X anchors, and the MPP / DC-switch / string-label
    # prototypes; then redraw `strings_per_mppt` rows per MPPT at a uniform pitch.
    PORT_LBL_RE = re.compile(r'^(\d+)-(\d+)$')
    port_rows = sorted(
        ((int(_m2.group(1)), int(_m2.group(2)), d)
         for d in tmpl_texts
         if (_m2 := PORT_LBL_RE.match(_strip_mtext_fmt(d['text'])))),
        key=lambda r: -r[2]['y'])
    max_tmpl_m = max((mp for mp, _, _ in port_rows), default=16)

    # Uniform single-row pitch = median gap between consecutive port-label rows.
    _port_ys = sorted((d['y'] for _, _, d in port_rows), reverse=True)
    _gaps = sorted(_port_ys[i] - _port_ys[i + 1] for i in range(len(_port_ys) - 1))
    row_pitch = _gaps[len(_gaps) // 2] if _gaps else 108.84

    # Terminal circles in the string column (small radius, right of the inverter box).
    _term_circ = [d for d in tmpl if d['type'] == 'CIRCLE'
                  and d.get('radius', 99) < 50 and d['cx'] > 20800]
    circ_x        = _term_circ[0]['cx'] if _term_circ else 20945.2
    first_circle_y = max((d['cy'] for d in _term_circ), default=tmpl_y_max)
    tmpl_bottom_y  = min((d['cy'] for d in _term_circ), default=tmpl_y_min)

    # Reference row (port 8-1) anchors: connection (circle) Y, port-label Y, cable end.
    _ref_pl_y = next((d['y'] for mp, pt, d in port_rows if mp == 8 and pt == 1),
                     port_rows[0][2]['y'] if port_rows else first_circle_y)
    _ref_conn = min((d['cy'] for d in _term_circ),
                    key=lambda y: abs(y - (_ref_pl_y - 36.0)), default=_ref_pl_y)
    lbl_dy = _ref_pl_y - _ref_conn          # label sits this far above the circle/conn

    PORT_X = next((d['x'] for mp, pt, d in port_rows if mp == 8 and pt == 1), circ_x - 137.7)
    _strlbls = [m for m in tmpl_texts if m['cls'] == 'string_label']
    STR_X = min((d['x'] for d in _strlbls), default=circ_x + 1237.0)

    cable_x_end = STR_X - 20.0

    # ── Auto-detect layout parameters from template geometry ──────────────────
    # strings_per_mppt: template port labels define the actual slot count; always
    # override the configured value so the output matches the template exactly.
    _auto_max_slot = max((pt for _, pt, _ in port_rows), default=strings_per_mppt)
    if _auto_max_slot != strings_per_mppt:
        print(f"  Template defines {_auto_max_slot} slots/MPPT "
              f"(overriding configured strings_per_mppt={strings_per_mppt})")
        strings_per_mppt = _auto_max_slot

    # MPPT slot layout: auto-detect which slots are reserved vs active by scanning
    # the first inverter column's string labels in the template. Only runs when the
    # template has 3+ slots/MPPT (2-slot templates have no meaningful reserve pattern)
    # and the user has not already provided an explicit mppt_layout.
    if not mppt_layout and strings_per_mppt >= 3 and _strlbls:
        _left_port_x = min(d['x'] for _, _, d in port_rows) if port_rows else PORT_X
        _auto_layout: dict = {}
        _seen_mp: set = set()
        for _am, _ap, _apr_d in sorted(port_rows, key=lambda r: (r[0], r[1])):
            if abs(_apr_d['x'] - _left_port_x) > 300:
                continue  # skip duplicate columns from multi-inverter template
            if (_am, _ap) in _seen_mp:
                continue
            _seen_mp.add((_am, _ap))
            _apy = _apr_d['y']
            # Prefer exact Y match (string label on same line as port label)
            _exact = [sl for sl in _strlbls if abs(sl['y'] - _apy) < 10]
            if _exact:
                _asl = min(_exact, key=lambda d: abs(d['x'] - STR_X))
            else:
                # Fuzzy match: prefer highest Y so the panel-detail string label
                # (which sits above its port row) wins over the label one row below.
                _fuzzy = [sl for sl in _strlbls if abs(sl['y'] - _apy) < PORT_Y_TOL]
                _asl = (min(_fuzzy, key=lambda d: (abs(d['x'] - STR_X), -d['y']))
                        if _fuzzy else None)
            if _asl is not None:
                _atxt = _strip_mtext_fmt(_asl['text']).lower()
                _auto_layout.setdefault(_am, []).append('R' if _atxt == 'reserve' else _ap)
        if any('R' in v for v in _auto_layout.values()):
            mppt_layout = _auto_layout
            _nres = sum(v.count('R') for v in _auto_layout.values())
            print(f"  Auto-detected MPPT layout from template: {len(_auto_layout)} MPPTs, "
                  f"{_nres} reserved slot(s)")

    # Combiner bus X anchors = vertical bracket polylines inside the box, left of circle.
    _bus_xs = []
    for d in tmpl:
        if d['type'] == 'LWPOLYLINE':
            xs = [p[0] for p in d['pts']]
            ys = [p[1] for p in d['pts']]
            if xs and (max(ys) - min(ys)) > row_pitch * 0.3 and (max(xs) - min(xs)) < 5 \
               and 20100 < xs[0] < 20650:
                _bus_xs.append(round(xs[0], 1))
    _bus_xs = sorted(set(_bus_xs))
    BUS_X   = _bus_xs[-1] if _bus_xs else 20471.2   # inner bus (closer to the circle)
    OUTER_X = _bus_xs[0]  if _bus_xs else 20243.0   # outer bus (toward the inverter box)

    # Inverter-box internal vertical bus that the MPPT feeds tie back into.
    BOX_BUS_X = OUTER_X - 510.0
    for d in tmpl:
        if d['type'] == 'LWPOLYLINE':
            xs = [p[0] for p in d['pts']]
            ys = [p[1] for p in d['pts']]
            if xs and (max(xs) - min(xs)) < 5 and (max(ys) - min(ys)) > row_pitch * 5 \
               and 19300 < xs[0] < 19950:
                BOX_BUS_X = xs[0]
                break

    # MTEXT prototypes whose style we reuse when re-stamping labels.
    strlbl_proto  = _strlbls[0] if _strlbls else None
    portlbl_proto = port_rows[0][2] if port_rows else strlbl_proto
    mpp_proto = next((d for d in tmpl_texts
                      if re.match(r'^MPP\s*\d+$', _strip_mtext_fmt(d['text']))), None)
    dc_proto  = next((d for d in tmpl_texts
                      if re.search(r'DC SWITCH', _strip_mtext_fmt(d['text']), re.I)), None)

    # Capture small decorations of the reference row (diode/connector ticks + INSERT),
    # normalised so their connection Y == 0, to re-stamp on every active row.
    row_deco = []
    for d in tmpl:
        yc = _ent_y_dict(d)
        if yc is None or abs(yc - _ref_conn) > row_pitch * 0.55:
            continue
        if d['type'] == 'LWPOLYLINE':
            xs = [p[0] for p in d['pts']]
            ys = [p[1] for p in d['pts']]
            if xs and (max(ys) - min(ys)) <= row_pitch * 0.8 \
               and min(xs) >= 20880 and max(xs) <= 21160:
                row_deco.append(dict(d, pts=[[p[0], p[1] - _ref_conn] + list(p[2:])
                                             for p in d['pts']]))
        elif d['type'] == 'INSERT' and 20600 < d.get('ix', 0) < 20850:
            row_deco.append(dict(d, iy=d['iy'] - _ref_conn))

    # Split the slice into the STATIC FRAME (everything that is NOT part of the
    # repeating connection region). Repeating = terminal circles, color-40 cables,
    # port/string/MPP/DC labels, combiner bracket polylines and row decorations.
    def _is_repeating(d):
        t = d['type']
        if t == 'CIRCLE':
            return d.get('radius', 99) < 50 and d['cx'] > 20800
        if t == 'MTEXT':
            s = _strip_mtext_fmt(d['text'])
            return bool(PORT_LBL_RE.match(s) or d.get('cls') == 'string_label'
                        or re.match(r'^MPP\s*\d+$', s)
                        or re.search(r'DC SWITCH', s, re.I))
        if t == 'LWPOLYLINE':
            if d.get('color') == 40:
                return True
            xs = [p[0] for p in d['pts']]
            ys = [p[1] for p in d['pts']]
            if not xs:
                return False
            # Any vertex in the combiner-bus band is connection-region geometry (the
            # template's own 2-port brackets, which we redraw) - drop it regardless of
            # span so it can't overlap the procedural combiner.
            if any(20200 <= x <= 20520 for x in xs):
                return True
            # Short stubs / decorations just right of the terminals.
            return bool((max(ys) - min(ys)) < row_pitch * 2.4
                        and min(xs) >= 20520 and max(xs) <= 21160)
        if t == 'INSERT':
            return 20600 < d.get('ix', 0) < 20850
        return False

    frame = [d for d in tmpl if not _is_repeating(d)]
    TOP_Y = first_circle_y           # connection Y of the first (top) string row

    # Panel-detail strip: the first string's cable must extend to the numbered-box edge.
    _panel_cable_right_x = cable_x_end
    for _pd in frame:
        if _pd['type'] == 'LWPOLYLINE':
            _pd_xs = [p[0] for p in _pd.get('pts', [])]
            _pd_ys = [p[1] for p in _pd.get('pts', [])]
            if (_pd_xs and max(_pd_xs) > 22000
                    and any(abs(y - first_circle_y) < row_pitch for y in _pd_ys)):
                _panel_cable_right_x = max(_panel_cable_right_x, max(_pd_xs))

    # Panel-detail string label offset: capture Y distance from first circle to the
    # "String T.I.1" label, which sits above the panel box strip (not at lbl_dy).
    _panel_lbl_y_offset = lbl_dy
    _panel_lbl_x        = STR_X
    if _strlbls:
        _topmost_sl = max(_strlbls, key=lambda d: d['y'])
        _toff = _topmost_sl['y'] - first_circle_y
        if _toff > row_pitch * 0.6:
            _panel_lbl_y_offset = _toff
            _panel_lbl_x        = _topmost_sl['x']
    # The panel-detail label has 2 lines; enforce a minimum so it never overlaps
    # the next row's string label regardless of what the template reported.
    # Derivation: 2*char_h*line_spacing (label height) - row_pitch + lbl_dy + margin.
    _min_panel_lbl_offset = 2.0 * text_size_cfg * 1.5 - row_pitch + lbl_dy + 20.0
    _panel_lbl_y_offset = max(_panel_lbl_y_offset, _min_panel_lbl_offset)

    template_rows = len(port_rows)   # 32 for the 16-MPPT / 2-port template
    print(f"Row layout: pitch={row_pitch:.1f} units, template MPPTs={max_tmpl_m}, "
          f"strings/MPPT={strings_per_mppt}, MPPTs/switch={mppts_per_switch}, "
          f"frame entities={len(frame)} (repeating removed={len(tmpl) - len(frame)})")

    # ── Drawing helpers for the procedural connection region ──────────────────
    def _wire_style(active, section_str):
        if active and section_str == heavy_section:
            return {'linetype': heavy_linetype, 'color': heavy_color,
                    'layer': heavy_layer, 'ltscale': 0.5}
        if active:
            return {'linetype': 'Continuous', 'color': 40, 'layer': '0'}
        return {'linetype': 'Continuous', 'color': 8, 'layer': '0'}

    def _add_poly(pts, style=None):
        ne = msp.add_lwpolyline(pts)
        if style:
            for a, v in style.items():
                try:
                    setattr(ne.dxf, a, v)
                except Exception:
                    pass
        return ne

    def _place_lbl(proto, x, y, text, color, dxo, dyo, width=None, char_h=None):
        if not proto:
            return
        d = dict(proto, x=x, y=y, color=color)
        if width:
            d['width'] = width
        if char_h:
            d['char_height'] = char_h
        _place_mtext(msp, d, dxo, dyo, text)

    def _draw_combiner(dxo, dyo, grp_ys, mpp_text):
        # Vertical bus joining the group's terminals; horizontal feed to OUTER_X bus.
        # The DC switch bracket (OUTER_X vertical + feed to BOX_BUS_X) is drawn in section C.
        yb0, yb1 = grp_ys[0], grp_ys[-1]
        yc = (yb0 + yb1) / 2.0
        _add_poly([(BUS_X + dxo, yb0 + dyo), (BUS_X + dxo, yb1 + dyo)])
        _add_poly([(OUTER_X + dxo, yc + dyo), (BUS_X + dxo, yc + dyo)])
        if mpp_proto:
            _place_lbl(mpp_proto, mpp_proto['x'], yc + lbl_dy, mpp_text, 7, dxo, dyo)

    def _inv_layout(T, I):
        """(num_mppts, eff_k, total_rows) for one inverter."""
        inv_data = excel.get((T, I), {})
        n_mppts = max(inv_data.keys()) if inv_data else max_tmpl_m
        max_strings = max((len(v) for v in inv_data.values()), default=0)
        eff_k = max(strings_per_mppt, max_strings, 1)
        if mppt_layout:
            layout_k = max((len(v) for v in mppt_layout.values()), default=eff_k)
            eff_k = max(eff_k, layout_k)
        return n_mppts, eff_k, n_mppts * eff_k

    # ── 3. Clean Model Space & Paper Layouts ──────────────────────────────────
    print("Clearing templates from Model Space and existing Viewports...")
    msp.delete_all_entities()
    paper_layouts = [l.name for l in doc.layouts if not l.is_modelspace]
    for name in paper_layouts:
        try:
            doc.layouts.delete(name)
        except Exception:
            pass

    # Precalculate layout offsets per Cabin (height now driven by rows = MPPTs x strings).
    template_height = (template_rows - 1) * row_pitch
    cabin_y_offset = {}
    current_y = 0.0
    for T in transformer_list:
        cabin_y_offset[T] = current_y
        max_stretch = 0.0
        for I in transformers[T]:
            _, _, total_rows = _inv_layout(T, I)
            stretch = (total_rows - 1) * row_pitch - template_height
            if stretch > max_stretch:
                max_stretch = stretch
        current_y -= (row_spacing + max_stretch)

    # ── 4. Place Inverters and Group into Cabins ──────────────────────────────
    print(f"Generating DXF drawing sheets...")
    
    # Pre-select reference tags
    td  = next((m for m in tmpl_texts if m['cls'] == 'title'),        None)
    chd = next((m for m in tmpl_texts if m['cls'] == 'cabin_header'), None)
    cld = next((m for m in tmpl_texts if m['cls'] == 'cabin_label'),  None)

    # Extract the inverter model string embedded in the template title
    # (format: "INVERTER 1.1 - MODEL - P= ...") so we can replace it in the
    # inner-box model label regardless of what brand the template was built with.
    _tmpl_inv_model = None
    if td:
        _title_stripped = _strip_mtext_fmt(td['text'])
        _tparts = re.split(r'\s*-\s*', _title_stripped)
        if len(_tparts) >= 2:
            _candidate = _tparts[1].strip()
            if _candidate and 'P=' not in _candidate and len(_candidate) > 3:
                _tmpl_inv_model = _candidate

    def make_string_label(T, I, mppt, port, is_panel_row=False):
        lst = excel.get((T, I), {}).get(mppt, [])
        if port - 1 < len(lst):
            sdata = lst[port - 1]
            name = sdata['name']
            wp = sdata['wp']
            l_plus = sdata['l_plus']
            l_minus = sdata['l_minus']
            sec = sdata['section']

            label = f"String {name}"
            # Panel-detail row: label only - the panel strip below already shows the full model.
            if not is_panel_row and not hide_string_details:
                if panels_per_string > 0:
                    # Per-watt model override (mixed-panel inverters) falls back to the
                    # global panel model when this watt-peak is not mapped.
                    mdl = panel_model_map.get(wp, panel_model)
                    if mdl:
                        label += f" - {panels_per_string}x {mdl}"
                    else:
                        suffix = f" {wp}Wp" if wp > 0 else ""
                        label += f" - {panels_per_string}P{suffix}"
            if show_cable_info:
                label += f" (L+={l_plus:.3f}m, L-={l_minus:.3f}m, {sec})"
            return label
        return "reserve"

    def make_inv_title(T, I):
        dc = dc_power_kwp
        if panels_per_string > 0:
            inv_data = excel.get((T, I), {})
            total_kwp = sum(
                panels_per_string * sd['wp'] / 1000.0
                for strings in inv_data.values()
                for sd in strings
                if sd['wp'] > 0
            )
            if total_kwp > 0:
                dc = total_kwp
        parts = [f"INVERTER {T}.{I}"]
        if inverter_model:
            parts.append(inverter_model)
        if dc > 0:
            parts.append("P= " + f"{dc:.2f}".replace('.', ',') + " KWp")
        if ac_power_kwac > 0:
            parts.append(f"P= {ac_power_kwac:.0f} KWac @{temp_rating:.0f}°C")
        return " - ".join(parts)

    def make_cabin_hdr(T):
        return f"CABIN {T}\n{transformer_power}" if transformer_power else f"CABIN {T}"

    for T in transformer_list:
        print(f"Stamping Cabin {T} (Y Offset: {cabin_y_offset[T]:.0f}) ...")
        for idx, I in enumerate(transformers[T]):
            dxo = idx * col_spacing
            dyo = cabin_y_offset[T]
            inv_data = excel.get((T, I), {})
            num_mppts, eff_k, total_rows = _inv_layout(T, I)
            inv_max_strings = max((len(v) for v in inv_data.values()), default=0)
            if inv_max_strings > strings_per_mppt:
                print(f"  [WARNING] Inverter {T}.{I}: an MPPT carries {inv_max_strings} strings "
                      f"(> 'Strings per MPPT'={strings_per_mppt}); drawing {eff_k} rows/MPPT so no "
                      f"string is dropped. Set 'Strings per MPPT' to {inv_max_strings} for a uniform sheet.")

            # Vertical frame stretch: box top + upper content stay fixed; the box bottom
            # and bottom annotations close up just under the last drawn row.
            bottom_y    = TOP_Y - (total_rows - 1) * row_pitch
            frame_shift = bottom_y - tmpl_bottom_y
            split_y     = max(bottom_y, tmpl_bottom_y) - row_pitch * 0.5
            EXTRA       = -frame_shift           # helper shifts y < split_y by -EXTRA

            # Panel-detail box (top) illustrates the first string, so use that string's
            # model when the inverter mixes panel power classes.
            _first_lst = inv_data.get(1) or next((v for v in inv_data.values() if v), [])
            _first_wp = _first_lst[0]['wp'] if _first_lst else 0
            inv_panel_model = panel_model_map.get(_first_wp, panel_model)

            # A. Static frame (stretched)
            for d in frame:
                if (d['type'] == 'CIRCLE' and not show_annot_circle
                        and d.get('color') == 1 and d.get('radius', 0) > 50):
                    continue
                if d['type'] == 'MTEXT':
                    cls = d.get('cls', 'fixed')
                    if cls in ('title', 'cabin_header', 'cabin_label'):
                        continue                 # placed separately in section D
                    if cls == 'panel_count':
                        updated = _update_panel_count_label(d['text'], panels_per_string, inv_panel_model)
                        _place_entity_stretched(msp, dict(d, text=updated), dxo, dyo, split_y, EXTRA)
                        continue
                    if 'mmq' in d['text']:
                        sections_used = {sdata['section']
                                         for mppt_s in inv_data.values() for sdata in mppt_s}
                        if len(sections_used) == 1:
                            new_txt = f"2/ {list(sections_used)[0]} mmq - Cu - H1Z2Z2k"
                        else:
                            new_txt = "2/ (1x6/10)mmq - Cu - H1Z2Z2k"
                        _place_entity_stretched(msp, dict(d, text=new_txt), dxo, dyo, split_y, EXTRA)
                        continue
                    if ('SUNGROW' in d['text'] or 'inverter' in d['text'].lower()
                            or (_tmpl_inv_model and _tmpl_inv_model in d['text'])):
                        d2 = dict(d)
                        if inverter_model:
                            d2['text'] = inverter_model.upper()
                        _place_entity_stretched(msp, d2, dxo, dyo, split_y, EXTRA)
                        continue
                    _stripped_d = _strip_mtext_fmt(d['text'])
                    if 'AC output power' in _stripped_d or 'Nominal AC voltage' in _stripped_d:
                        _ac_parts = []
                        if ac_power_30c > 0:
                            _ac_parts.append(f"{ac_power_30c:.0f} kVA @30°C")
                        if ac_power_kwac > 0:
                            _ac_parts.append(f"{ac_power_kwac:.0f} kVA @40°C")
                        _ac_line1 = ("AC output power " + " / ".join(_ac_parts)) if _ac_parts else "AC output power"
                        _ac_curr  = f"Max AC output current: {max_ac_current}" if max_ac_current else "Max AC output current:"
                        _ac_txt   = f"{_ac_line1}\\PNominal AC voltage : 800V±10%, 3F + PE \\P{_ac_curr}"
                        _place_entity_stretched(msp, dict(d, text=_ac_txt), dxo, dyo, split_y, EXTRA)
                        continue
                    if 'CC side' in _stripped_d or 'MPPT range' in _stripped_d:
                        _n_inputs = num_mppts * eff_k
                        _cc_pv  = (f"Max PV input current per MPPT: {max_pv_current_per_mppt}"
                                   if max_pv_current_per_mppt else "Max PV input current per MPPT:")
                        _cc_sc  = (f"Max DC short circuit current per MPPT: {max_dc_sc_current_per_mppt}"
                                   if max_dc_sc_current_per_mppt else "Max DC short circuit current per MPPT:")
                        _cc_txt = (f"CC side\\P{_n_inputs} input - {num_mppts} MPPT"
                                   f"\\PMax Vdc : {max_vdc}\\P{_cc_pv}\\P{_cc_sc}"
                                   f"\\PMPPT range :{mppt_voltage_range}")
                        # Shift CC text up so it stays inside the shrinking box.
                        # floor: insertion point must be high enough that ~7 lines of
                        # text (including possible wrapping) clear the new box bottom.
                        _cc_y = d.get('y', d.get('cy', 0)) + max(0.0, frame_shift)
                        _cc_floor = (tmpl_bottom_y - row_pitch + frame_shift
                                     + 7.0 * text_size_cfg * 1.6 + 30.0)
                        _cc_y = max(_cc_y, _cc_floor)
                        _cc_d = dict(d, x=BOX_BUS_X - 2800.0, y=_cc_y, text=_cc_txt)
                        _place_entity_stretched(msp, _cc_d, dxo, dyo, split_y, EXTRA)
                        continue
                    _place_entity_stretched(msp, d, dxo, dyo, split_y, EXTRA)
                else:
                    _place_entity_stretched(msp, d, dxo, dyo, split_y, EXTRA)

            # B. String rows + per-MPPT combiner brackets (procedural, k rows per MPPT)
            gr = 0
            for m in range(1, num_mppts + 1):
                lst = inv_data.get(m, [])
                grp_ys = []

                # Build slot list: [(slot_p, is_active, lookup_p), ...]
                # slot_p   = physical slot label (1-indexed position in this MPPT)
                # is_active= True if a real string is assigned
                # lookup_p = 1-based index into lst for make_string_label (port arg)
                layout_row = mppt_layout.get(m)
                if layout_row:
                    slots = []
                    active_count = 0
                    for _si, _state in enumerate(layout_row):
                        _sp = _si + 1
                        if _state == 'R':
                            slots.append((_sp, False, _sp))
                        else:
                            _lp = active_count + 1
                            slots.append((_sp, active_count < len(lst), _lp))
                            active_count += 1
                else:
                    slots = [(p, (p - 1) < len(lst), p) for p in range(1, eff_k + 1)]

                for slot_p, active, lookup_p in slots:
                    yc = TOP_Y - gr * row_pitch
                    _is_panel_row = (m == 1 and slot_p == slots[0][0])
                    grp_ys.append(yc)
                    gr += 1
                    section_str = lst[lookup_p - 1]['section'] if active and (lookup_p - 1) < len(lst) else ''
                    style = _wire_style(active, section_str)
                    cc = msp.add_circle((circ_x + dxo, yc + dyo), circle_radius_cfg)
                    cc.dxf.color = 7 if active else 8
                    _add_poly([(BUS_X + dxo, yc + dyo),
                               (circ_x - circle_radius_cfg + dxo, yc + dyo)], style)
                    _ce = _panel_cable_right_x if _is_panel_row else cable_x_end
                    _add_poly([(circ_x + circle_radius_cfg + dxo, yc + dyo),
                               (_ce + dxo, yc + dyo)], style)
                    if active:
                        for dd in row_deco:
                            _place_entity(msp, dd, dxo, yc + dyo)
                    _place_lbl(portlbl_proto, PORT_X, yc + lbl_dy, f"{m}-{slot_p}",
                               7 if active else 8, dxo, dyo)
                    _sl_y = yc + (_panel_lbl_y_offset if _is_panel_row else lbl_dy)
                    _sl_x = _panel_lbl_x if _is_panel_row else STR_X
                    _place_lbl(strlbl_proto, _sl_x, _sl_y,
                               make_string_label(T, I, m, lookup_p, is_panel_row=_is_panel_row) if active else "reserve",
                               7 if active else 8, dxo, dyo,
                               width=STRING_LABEL_MIN_WIDTH, char_h=text_size_cfg)
                _draw_combiner(dxo, dyo, grp_ys, f"MPP{m}")

            # C. DC-switch brackets + labels, grouping MPPTs by `mppts_per_switch`.
            # OUTER_X vertical spans the center-Y of the first MPPT to the center-Y of
            # the last MPPT in the switch (matches the webapp _draw_dc_switch geometry).
            n_sw = (num_mppts + mppts_per_switch - 1) // mppts_per_switch
            for s in range(n_sw):
                first_m = s * mppts_per_switch + 1
                last_m  = min((s + 1) * mppts_per_switch, num_mppts)
                y_first_mid = TOP_Y - ((first_m - 1) * eff_k + (eff_k - 1) / 2.0) * row_pitch
                y_last_mid  = TOP_Y - ((last_m  - 1) * eff_k + (eff_k - 1) / 2.0) * row_pitch
                y_sw_mid = (y_first_mid + y_last_mid) / 2.0
                if mppts_per_switch > 1:
                    _add_poly([(OUTER_X + dxo, y_first_mid + dyo), (OUTER_X + dxo, y_last_mid + dyo)])
                _add_poly([(BOX_BUS_X + dxo, y_sw_mid + dyo), (OUTER_X + dxo, y_sw_mid + dyo)])
                _place_lbl(dc_proto, (dc_proto['x'] if dc_proto else OUTER_X - 100.0),
                           y_sw_mid + lbl_dy, f"DC SWITCH {s + 1}", 7, dxo, dyo)

            # D. Headers
            if td:
                _place_mtext(msp, td,  dxo, dyo, make_inv_title(T, I))
            if chd:
                _place_mtext(msp, chd, dxo, dyo, make_cabin_hdr(T))
            if cld:
                _place_mtext(msp, cld, dxo, dyo, f"\\pxqc;Cabin Tx.{T}\\PInverter {T}.{I}")

    # ── 5. Generate A3 Paper Space Viewports ──────────────────────────────────
    print("Setting up Paper Space layouts...")
    tmpl_x_center = xmin + COL_STEP / 2
    
    for T in transformer_list:
        lname = f"Tx{T}"
        try:
            layout = doc.layouts.new(lname)
        except Exception:
            layout = doc.layouts.get(lname)

        n_inv  = len(transformers[T])
        row_cx = tmpl_x_center + (n_inv - 1) * col_spacing / 2
        
        # Calculate row height bounding box from the procedural row stack
        min_y_val = tmpl_y_min
        for I in transformers[T]:
            _, _, total_rows = _inv_layout(T, I)
            inv_bottom = TOP_Y - (total_rows - 1) * row_pitch - row_pitch
            if inv_bottom < min_y_val:
                min_y_val = inv_bottom

        row_cy = cabin_y_offset[T] + (tmpl_y_max + min_y_val) / 2
        row_w  = col_spacing * n_inv
        cabin_height = tmpl_y_max - min_y_val

        view_h = max(cabin_height * 1.05, row_w / (420.0 / 297.0) * 1.05)

        layout.add_viewport(
            center=(210, 148.5),
            size=(420, 297),
            view_center_point=(row_cx, row_cy),
            view_height=view_h,
        )
        print(f"  Created layout '{lname}' with {n_inv} inverters (View Height: {view_h:.0f})")

    # ── 6. Save Output ────────────────────────────────────────────────────────
    print(f"Saving generated DXF to: {OUTPUT_PATH}")
    doc.saveas(OUTPUT_PATH)
    print("Done! SLD generation completed successfully.")

# ─────────────────────────────────────────────────────────────────────────────
#  HISTORY & PRESET STORAGE
# ─────────────────────────────────────────────────────────────────────────────

HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'smart_history.json')

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_history(data):
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception as ex:
        print(f"Failed to save history: {ex}", file=sys.stderr)

# ─────────────────────────────────────────────────────────────────────────────
#  GUI IMPLEMENTATION  (CustomTkinter – Material Light)
# ─────────────────────────────────────────────────────────────────────────────

try:
    import customtkinter as ctk
    HAS_CTK = True
except ImportError:
    HAS_CTK = False

# ── Material colour palette ──────────────────────────────────────────────────
_BG        = "#F8F9FA"   # window background
_CARD      = "#FFFFFF"   # card / section background
_BLUE      = "#1A73E8"   # primary accent (Google Blue)
_BLUE_H    = "#1558B0"   # hover blue
_TEXT      = "#202124"   # primary text
_MUTED     = "#5F6368"   # secondary / muted text
_BORDER    = "#DADCE0"   # subtle card border
_SUCCESS   = "#34A853"
_ERROR     = "#EA4335"
_STAR      = "#EA4335"   # required-field asterisk colour

# ── Helper: labelled row with red asterisk for required fields ───────────────
def _row(parent, label, required=False, **kwargs):
    """Return (row_frame, label_widget). Appends * in red if required."""
    row = ctk.CTkFrame(parent, fg_color="transparent")
    row.pack(fill="x", pady=4)
    lbl_text = label + ("  *" if required else "")
    lbl = ctk.CTkLabel(row, text=lbl_text, text_color=_TEXT if not required else _TEXT,
                       anchor="w", width=200, font=("Segoe UI", 13))
    if required:
        # Draw asterisk in red via a separate tiny label
        lbl = ctk.CTkLabel(row, text=label, text_color=_TEXT,
                           anchor="w", width=190, font=("Segoe UI", 13))
        lbl.pack(side="left")
        ctk.CTkLabel(row, text=" *", text_color=_STAR,
                     font=("Segoe UI", 13, "bold"), width=12).pack(side="left")
    else:
        lbl.pack(side="left")
    return row

def _card(parent, title):
    """Return a white-background card frame with a section title."""
    outer = ctk.CTkFrame(parent, fg_color=_CARD, corner_radius=10,
                         border_width=1, border_color=_BORDER)
    outer.pack(fill="x", padx=20, pady=10)
    ctk.CTkLabel(outer, text=title, font=("Segoe UI", 14, "bold"),
                 text_color=_BLUE, anchor="w").pack(anchor="w", padx=20, pady=(14, 2))
    sep = ctk.CTkFrame(outer, height=1, fg_color=_BORDER)
    sep.pack(fill="x", padx=20, pady=(0, 10))
    body = ctk.CTkFrame(outer, fg_color="transparent")
    body.pack(fill="x", padx=20, pady=(0, 16))
    return body


class SmartSLDGui:
    def __init__(self):
        if not HAS_CTK:
            # Fallback: try plain tkinter
            if not HAS_TK:
                print("[ERROR] Neither customtkinter nor tkinter is available.", file=sys.stderr)
                sys.exit(1)
            print("[WARNING] customtkinter not found, falling back to basic tkinter GUI.")
            self._run_fallback_tk()
            return

        ctk.set_appearance_mode("Light")
        ctk.set_default_color_theme("blue")

        self.root = ctk.CTk()
        self.root.title("A176LAB - Generatore Smart SLD DC")
        self.root.geometry("940x780")
        self.root.configure(fg_color=_BG)
        self.root.resizable(True, True)

        self.history = load_history()
        # Merge any user-saved custom presets back into the live module-level dicts
        for name, (dc, ac) in self.history.get('custom_inverters', {}).items():
            _INVERTER_POWERS.setdefault(name, (dc, ac))
        for name in self.history.get('custom_panels', []):
            if name not in _PANELS_PRESETS:
                _PANELS_PRESETS.append(name)
        self._setup_vars()
        self._build_ui()

        self.root.bind("<F5>", lambda _e: self._start_generation())

    # ── Variables ─────────────────────────────────────────────────────────────
    def _setup_vars(self):
        _DEFAULT_TEMPLATE = os.environ.get('SLD_TEMPLATE_PATH', r'\\S01\5 temp\A176Lab Tools\SLD Generator\Lista Cavi - Cavi LV-DC 28.05.2026.dxf')
        self.var_excel    = ctk.StringVar(value=self.history.get('xlsx_path', ''))
        self.var_template = ctk.StringVar(value=self.history.get('template_dxf', _DEFAULT_TEMPLATE))
        self.var_out      = ctk.StringVar(value=self.history.get('output_path', ''))

        self.var_panel_model = ctk.StringVar(value=self.history.get('panel_model', 'JA Solar JAM72D42-625/LB'))
        self.var_panel_map   = ctk.StringVar(value=self.history.get('panel_model_map', ''))
        self.var_panels      = ctk.StringVar(value=self.history.get('panels_per_string', '28'))
        self.var_strings_mppt = ctk.StringVar(value=self.history.get('strings_per_mppt', '2'))
        self.var_mppts_switch = ctk.StringVar(value=self.history.get('mppts_per_switch', '4'))
        self.var_inv_model   = ctk.StringVar(value=self.history.get('inverter_model', 'Sungrow SG350HX'))
        self.var_dc_power    = ctk.StringVar(value=self.history.get('dc_power_kwp', '350'))
        self.var_ac_power    = ctk.StringVar(value=self.history.get('ac_power_kwac', '320'))
        self.var_ac_power_30c = ctk.StringVar(value=self.history.get('ac_power_30c', ''))
        self.var_max_ac_current = ctk.StringVar(value=self.history.get('max_ac_current', ''))
        self.var_max_pv_current = ctk.StringVar(value=self.history.get('max_pv_current_per_mppt', ''))
        self.var_max_dc_sc = ctk.StringVar(value=self.history.get('max_dc_sc_current_per_mppt', ''))
        self.var_mppt_vrange = ctk.StringVar(value=self.history.get('mppt_voltage_range', ''))
        self.var_max_vdc   = ctk.StringVar(value=self.history.get('max_vdc', ''))
        self.var_temp        = ctk.StringVar(value=self.history.get('temp_rating', '40'))

        self.var_show_cable          = ctk.BooleanVar(value=self.history.get('show_cable_info', False))
        self.var_hide_string_details = ctk.BooleanVar(value=self.history.get('hide_string_details', True))
        self.var_show_annot_circle   = ctk.BooleanVar(value=self.history.get('show_annot_circle', False))

        self.var_col_spacing  = ctk.StringVar(value=self.history.get('col_spacing', str(COL_SPACING_DEFAULT)))
        self.var_row_spacing  = ctk.StringVar(value=self.history.get('row_spacing', str(ROW_SPACING_DEFAULT)))
        self.var_circle_radius= ctk.StringVar(value=self.history.get('circle_radius', '24.59'))
        self.var_text_size    = ctk.StringVar(value=self.history.get('text_size', '60.44'))
        self.var_heavy_section= ctk.StringVar(value=self.history.get('heavy_section', '1x10'))
        self.var_heavy_linetype=ctk.StringVar(value=self.history.get('heavy_linetype', 'TRATTEGGIATA'))
        self.var_heavy_color  = ctk.StringVar(value=self.history.get('heavy_color', '40'))
        self.var_heavy_layer  = ctk.StringVar(value=self.history.get('heavy_layer', 'TRATTEGGIATA'))

        # MPPT layout grid
        self.var_num_mppts_layout = ctk.StringVar(value=str(self.history.get('num_mppts_layout', '6')))
        self.var_mppt_layout      = ctk.StringVar(value=self.history.get('mppt_layout', ''))
        self._mppt_slot_btns:  dict = {}   # {(m, p): CTkButton}
        self._mppt_slot_state: dict = {}   # {(m, p): 'A'|'R'}
        self._mppt_grid_frame       = None

        # Dynamic per-Wp panel model fields
        self._wp_model_vars: dict = {}   # {wp_int: ctk.StringVar}
        self._wp_model_frame      = None  # set in _build_step2

        self.var_excel.trace_add('write', self._sync_output_path)
        self.var_excel.trace_add('write', self._on_excel_changed)

    def _sync_output_path(self, *_):
        xlsx = self.var_excel.get()
        if xlsx:
            self.var_out.set(os.path.splitext(xlsx)[0] + '_SLD_Generated.dxf')

    # ── Dynamic Wp model fields ───────────────────────────────────────────────
    def _on_excel_changed(self, *_):
        xlsx = self.var_excel.get()
        if xlsx and os.path.isfile(xlsx):
            threading.Thread(target=self._parse_excel_wp, args=(xlsx,), daemon=True).start()

    def _parse_excel_wp(self, xlsx_path):
        try:
            import openpyxl as _xl2
            wb = _xl2.load_workbook(xlsx_path, data_only=True, read_only=True)
            ws = None
            for name in wb.sheetnames:
                if name.strip().lower() == '2e802-3':
                    ws = wb[name]; break
            if ws is None:
                ws = wb.active
            header_row, col_wp = 30, 21
            for r_idx in (30, 29, 28, 31, 32, 27, 26):
                row_vals = [str(ws.cell(row=r_idx, column=c).value or '').strip().lower()
                            for c in range(1, 35)]
                if any('string name' in v for v in row_vals):
                    header_row = r_idx
                    _strong = False
                    for c in range(1, 35):
                        val = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
                        if any(kw in val for kw in ('module type', 'module power', 'watt peak')):
                            col_wp = c; _strong = True
                        elif not _strong and ('potenza' in val or re.search(r'(?<![a-z])wp(?![a-z])', val)):
                            col_wp = c
                    break
            wp_set: set = set()
            for row in ws.iter_rows(min_row=header_row + 1, min_col=col_wp, max_col=col_wp):
                for cell in row:
                    v = cell.value
                    if v is None: continue
                    try:
                        w = int(float(str(v).replace(',', '.')))
                        if 50 < w < 2000: wp_set.add(w)
                    except (ValueError, TypeError):
                        pass
            wb.close()
            wp_values = sorted(wp_set)
        except Exception as e:
            print(f"[WARNING] Could not parse Excel for Wp values: {e}")
            wp_values = []
        self.root.after(0, self._update_wp_model_ui, wp_values)

    def _update_wp_model_ui(self, wp_values):
        if not self._wp_model_frame:
            return
        for w in self._wp_model_frame.winfo_children():
            w.destroy()
        if not wp_values:
            ctk.CTkLabel(self._wp_model_frame,
                         text="Carica il file Excel per rilevare i tipi di pannello.",
                         text_color=_MUTED, font=("Segoe UI", 10),
                         anchor="w").pack(padx=8, pady=6, anchor="w")
            self._wp_model_vars = {}
            return
        # Parse saved panel_model_map (e.g. "720=Jinko JKM720N-...; 655=...") into a lookup dict
        saved_map = {}
        for part in re.split(r';\s*', self.var_panel_map.get().strip()):
            if '=' in part:
                k, _, v = part.partition('=')
                try:
                    saved_map[int(k.strip())] = v.strip()
                except ValueError:
                    pass
        default_model = self.var_panel_model.get().strip()
        self._wp_model_vars = {}
        for wp in wp_values:
            if wp in saved_map:
                prefill = saved_map[wp]
            elif len(wp_values) == 1:
                prefill = default_model
            else:
                prefill = ''
            var = ctk.StringVar(value=prefill)
            self._wp_model_vars[wp] = var
            row = ctk.CTkFrame(self._wp_model_frame, fg_color="transparent")
            row.pack(fill="x", padx=6, pady=2)
            ctk.CTkLabel(row, text=f"{wp} Wp", font=("Segoe UI", 10, "bold"),
                         text_color=_TEXT, width=55, anchor="e").pack(side="left", padx=(0, 6))
            ctk.CTkEntry(row, textvariable=var, height=28, corner_radius=6,
                         fg_color="white", border_color=_BORDER, font=("Segoe UI", 10),
                         placeholder_text="Nome modello").pack(side="left", fill="x", expand=True)

    def _build_panel_map_string(self):
        parts = [f"{wp}={var.get().strip()}"
                 for wp, var in sorted(self._wp_model_vars.items())
                 if var.get().strip()]
        if not parts:
            return self.var_panel_map.get().strip()
        return "; ".join(parts)

    def _on_inverter_select(self, choice):
        if choice in _INVERTER_POWERS:
            dc, ac = _INVERTER_POWERS[choice]
            self.var_dc_power.set(dc)
            self.var_ac_power.set(ac)

    # ── UI builder ────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Fixed Header ──────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self.root, fg_color=_BLUE, corner_radius=0, height=70)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        # Help button — must be packed BEFORE left-anchored labels so side="right" works
        ctk.CTkButton(
            hdr, text="?", width=36, height=36, corner_radius=18,
            fg_color="white", text_color=_BLUE, hover_color="#D2E3FC",
            font=("Segoe UI", 16, "bold"), border_width=0,
            command=self._show_help
        ).pack(side="right", padx=18, pady=17)
        ctk.CTkLabel(hdr, text="A176LAB - Generatore Smart SLD DC",
                     font=("Segoe UI", 18, "bold"), text_color="white").pack(anchor="w", padx=24, pady=(12, 0))
        ctk.CTkLabel(hdr, text="Flusso di generazione automatica geometrie AutoCAD",
                     font=("Segoe UI", 11), text_color="#C8D8F8").pack(anchor="w", padx=24)

        # ── Scrollable body ───────────────────────────────────────────────────
        self.scroll = ctk.CTkScrollableFrame(self.root, fg_color=_BG, corner_radius=0)
        self.scroll.pack(fill="both", expand=True)

        self._build_step1()
        self._build_step2()
        self._build_step3()
        self._build_step4()

        # Auto-parse saved Excel on startup to restore Wp model fields
        if self.var_excel.get() and os.path.isfile(self.var_excel.get()):
            self.root.after(200, self._on_excel_changed)

        # ── Fixed Footer ──────────────────────────────────────────────────────
        footer = ctk.CTkFrame(self.root, fg_color=_CARD, corner_radius=0,
                               border_width=1, border_color=_BORDER, height=190)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        # Progress bar
        self.progress = ctk.CTkProgressBar(footer, mode="indeterminate",
                                            fg_color=_BORDER, progress_color=_BLUE, height=4)
        self.progress.pack(fill="x", padx=20, pady=(10, 0))
        self.progress.set(0)

        # Status + clear log
        status_row = ctk.CTkFrame(footer, fg_color="transparent")
        status_row.pack(fill="x", padx=20, pady=(4, 2))
        self.lbl_status = ctk.CTkLabel(status_row, text="Pronto.", font=("Segoe UI", 11),
                                        text_color=_MUTED, anchor="w")
        self.lbl_status.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(status_row, text="Pulisci Log", width=90, height=26,
                       font=("Segoe UI", 11), fg_color="#E8EAED", text_color=_TEXT,
                       hover_color="#DADCE0", corner_radius=6,
                       command=self._clear_log).pack(side="right")

        # Log box
        self.log_box = ctk.CTkTextbox(footer, height=80, font=("Consolas", 10),
                                       fg_color="#F1F3F4", text_color=_TEXT,
                                       border_width=1, border_color=_BORDER,
                                       corner_radius=6, state="disabled", wrap="word")
        self.log_box.pack(fill="x", padx=20, pady=(0, 6))

        # Generate button
        self.btn_generate = ctk.CTkButton(
            footer, text="▶  Esegui Generazione  (F5)",
            font=("Segoe UI", 14, "bold"),
            fg_color=_BLUE, hover_color=_BLUE_H,
            text_color="white", corner_radius=22, height=42,
            command=self._start_generation)
        self.btn_generate.pack(padx=20, pady=(2, 10), fill="x")

    # ── Help dialog ───────────────────────────────────────────────────────────
    def _show_help(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Guida - Smart SLD Generator")
        dlg.geometry("700x600")
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.configure(fg_color=_BG)
        dlg.lift()

        # Title bar
        title_bar = ctk.CTkFrame(dlg, fg_color=_BLUE, corner_radius=0, height=52)
        title_bar.pack(fill="x")
        title_bar.pack_propagate(False)
        ctk.CTkLabel(title_bar, text="  ?  Come funziona il Generatore Smart SLD",
                     font=("Segoe UI", 15, "bold"), text_color="white",
                     anchor="w").pack(side="left", padx=16, pady=14)

        scroll = ctk.CTkScrollableFrame(dlg, fg_color=_BG, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        sections = [
            ("PANORAMICA", (
                "Il generatore crea automaticamente schemi unifilari DC (DXF) partendo da un\n"
                "foglio Excel del progetto e da un file DXF template di riferimento.\n"
                "Ogni inverter viene replicato e adattato in base ai dati reali del progetto,\n"
                "producendo uno schema pronto per AutoCAD con layout Paper Space A3."
            )),
            ("STEP 1 - FILE DI INPUT", (
                "Lista Cavi Excel\n"
                "  Il file Excel del progetto. Deve contenere due fogli:\n"
                "  - '2E802-3': riga intestazione (riga 30) con colonne Inverter, String Name,\n"
                "               MPPT, Section, Module Type (Wp).\n"
                "  - 'Inverter To String': colonne String Name, L+ (m), L- (m), Table No.\n\n"
                "Template DXF\n"
                "  Schema DXF di riferimento contenente la geometria dell'inverter 1.1.\n"
                "  Il generatore individua automaticamente le posizioni dei testi e degli\n"
                "  elementi grafici e li usa come base per tutti gli inverter generati.\n\n"
                "Output DXF\n"
                "  Percorso del file DXF generato. Si compila automaticamente come\n"
                "  <nome_excel>_SLD_Generated.dxf nella stessa cartella dell'Excel."
            )),
            ("STEP 2 - APPARECCHIATURE", (
                "Modello Inverter\n"
                "  Seleziona dall'elenco o aggiungine uno nuovo con [+].\n"
                "  La potenza DC/AC si compila automaticamente al momento della selezione.\n"
                "  Usa [...] per modificare o eliminare i modelli dall'elenco.\n\n"
                "Potenza DC (kWp) / AC (kWac)\n"
                "  Potenze nominali usate nell'intestazione INVERTER X.X del disegno.\n\n"
                "Temperatura (C)\n"
                "  Temperatura di riferimento per la potenza AC (es. 40 C).\n\n"
                "Modello Pannello\n"
                "  Nome del modulo fotovoltaico. Compare nell'etichetta di ogni stringa.\n"
                "  Usa [+] per aggiungere un nuovo modello, [...] per gestirli.\n\n"
                "Modelli per Wp (mappa)\n"
                "  Per inverter con piu tipi di pannello: associa ogni potenza modulo\n"
                "  (colonna 'Module type') a un modello, es. '720=Modello A; 725=Modello B'.\n"
                "  Ogni stringa usa il modello del suo Wp; le altre usano il Modello Pannello.\n"
                "  (Visibile solo con 'Nascondi Dettagli Stringa' disattivato.)\n\n"
                "Pannelli per Stringa\n"
                "  Numero di pannelli in serie per stringa (es. 28).\n\n"
                "Stringhe per MPPT\n"
                "  Numero di righe stringa disegnate per ogni MPPT (default 2). Se un MPPT\n"
                "  nell'Excel ha piu stringhe di questo valore, le righe vengono aumentate\n"
                "  automaticamente per non perdere stringhe (con avviso nel log).\n\n"
                "MPPT per Sezionatore DC\n"
                "  Quanti MPPT raggruppare sotto ogni etichetta DC SWITCH (default 4).\n"
                "  Il numero di sezionatori = MPPT totali / questo valore (arrotondato su)."
            )),
            ("STEP 3 - STILI AVANZATI", (
                "Includi Lunghezze Cavi\n"
                "  Se attivo, aggiunge L+ e L- in metri all'etichetta di ogni stringa.\n\n"
                "Nascondi Dettagli Stringa\n"
                "  Se attivo, mostra solo il nome stringa senza pannelli e modello.\n\n"
                "Step Colonne / Righe\n"
                "  Spaziatura orizzontale e verticale tra gli inverter nel DXF.\n\n"
                "Raggio cerchio\n"
                "  Raggio del terminale circolare di connessione stringa nel disegno.\n\n"
                "Dim. testo\n"
                "  Altezza del testo per le etichette stringa (unita DXF).\n\n"
                "Sezione heavy\n"
                "  La sezione cavo (es. 1x10) da visualizzare con linetype tratteggiata.\n"
                "  Il colore e il layer corrispondenti si impostano nei campi adiacenti."
            )),
            ("GENERAZIONE E OUTPUT", (
                "Premi F5 oppure il pulsante blu 'Esegui Generazione' in fondo alla finestra.\n"
                "Il log mostra l'avanzamento in tempo reale.\n\n"
                "Il file DXF generato contiene:\n"
                "  - Model Space con tutti gli inverter disposti per Cabin/Trasformatore.\n"
                "  - Un layout Paper Space A3 (420x297mm) per ciascun Cabin.\n"
                "  - Etichette stringa, intestazioni inverter e layer di stile aggiornati."
            )),
            ("SUPPORTO", "Per assistenza tecnica: info@a176lab.it"),
        ]

        for heading, body_text in sections:
            sec = ctk.CTkFrame(scroll, fg_color=_CARD, corner_radius=8,
                               border_width=1, border_color=_BORDER)
            sec.pack(fill="x", padx=16, pady=6)
            ctk.CTkLabel(sec, text=heading, font=("Segoe UI", 12, "bold"),
                         text_color=_BLUE, anchor="w").pack(anchor="w", padx=14, pady=(10, 2))
            ctk.CTkLabel(sec, text=body_text, font=("Segoe UI", 11),
                         text_color=_TEXT, justify="left", anchor="w",
                         wraplength=620).pack(anchor="w", padx=14, pady=(0, 10))

        ctk.CTkButton(dlg, text="Chiudi", fg_color=_BLUE, hover_color=_BLUE_H,
                       text_color="white", corner_radius=16, height=36, width=110,
                       command=dlg.destroy).pack(pady=14)

    # ── Step 1: Input files ───────────────────────────────────────────────────
    def _build_step1(self):
        body = _card(self.scroll,
                     "Step 1 - Dati di Input Obbligatori")
        ctk.CTkLabel(body, text="Seleziona i file fondamentali per avviare la generazione.",
                     font=("Segoe UI", 11), text_color=_MUTED, anchor="w").pack(anchor="w", pady=(0, 8))

        for label, var, browse_cmd, required in [
            ("Lista Cavi Excel",   self.var_excel,    self._browse_excel,    True),
            ("Template DXF base", self.var_template,  self._browse_template, True),
            ("Percorso Output DXF",self.var_out,      self._browse_output,   True),
        ]:
            row = _row(body, label, required=required)
            ctk.CTkEntry(row, textvariable=var, height=34, corner_radius=6,
                          fg_color="white", border_color=_BORDER,
                          font=("Segoe UI", 11)).pack(side="left", fill="x", expand=True, padx=(6, 4))
            ctk.CTkButton(row, text="Sfoglia...", width=90, height=34,
                           font=("Segoe UI", 11), corner_radius=6,
                           fg_color="#E8EAED", text_color=_TEXT, hover_color="#DADCE0",
                           command=browse_cmd).pack(side="left")

    # ── Step 2: Equipment specs ───────────────────────────────────────────────
    def _build_step2(self):
        body = _card(self.scroll, "Step 2 - Specifiche Apparecchiature")
        ctk.CTkLabel(body,
                     text="Parametri dell'impianto (lascia i default se non specificato).",
                     font=("Segoe UI", 11), text_color=_MUTED, anchor="w").pack(anchor="w", pady=(0, 8))

        grid = ctk.CTkFrame(body, fg_color="transparent")
        grid.pack(fill="x")
        col_l = ctk.CTkFrame(grid, fg_color="transparent")
        col_l.pack(side="left", fill="both", expand=True, padx=(0, 10))
        col_r = ctk.CTkFrame(grid, fg_color="transparent")
        col_r.pack(side="left", fill="both", expand=True, padx=(10, 0))

        def _lbl_entry(parent, label, var, width=None):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", pady=4)
            ctk.CTkLabel(f, text=label, text_color=_TEXT, anchor="w",
                          font=("Segoe UI", 12), width=170).pack(side="left")
            e = ctk.CTkEntry(f, textvariable=var, height=32, corner_radius=6,
                              fg_color="white", border_color=_BORDER,
                              font=("Segoe UI", 11), width=width or 120)
            e.pack(side="left", fill="x", expand=(width is None))
            return f

        # Left column
        # Inverter model
        f_inv = ctk.CTkFrame(col_l, fg_color="transparent")
        f_inv.pack(fill="x", pady=4)
        ctk.CTkLabel(f_inv, text="Modello Inverter:", text_color=_TEXT, anchor="w",
                      font=("Segoe UI", 12), width=170).pack(side="left")
        self.cb_inv = ctk.CTkComboBox(f_inv, variable=self.var_inv_model,
                                       values=sorted(_INVERTER_POWERS.keys()),
                                       command=self._on_inverter_select,
                                       height=32, corner_radius=6,
                                       fg_color="white", border_color=_BORDER,
                                       font=("Segoe UI", 11))
        self.cb_inv.pack(side="left", fill="x", expand=True, padx=(0, 4))
        ctk.CTkButton(f_inv, text="+", width=32, height=32, corner_radius=6,
                       fg_color="#E8EAED", text_color=_BLUE, hover_color="#C8D4EC",
                       font=("Segoe UI", 13, "bold"),
                       command=self._add_new_inverter).pack(side="left", padx=(0, 2))
        ctk.CTkButton(f_inv, text="...", width=32, height=32, corner_radius=6,
                       fg_color="#E8EAED", text_color=_MUTED, hover_color="#DADCE0",
                       font=("Segoe UI", 13),
                       command=self._manage_inverters).pack(side="left")

        _lbl_entry(col_l, "Potenza DC (kWp):", self.var_dc_power, 100)
        _lbl_entry(col_l, "Potenza AC @40°C (kVA):", self.var_ac_power, 100)
        _lbl_entry(col_l, "Potenza AC @30°C (kVA):", self.var_ac_power_30c, 100)
        _lbl_entry(col_l, "Corrente AC max:", self.var_max_ac_current, 120)
        _lbl_entry(col_l, "Temperatura (°C):", self.var_temp, 100)
        _lbl_entry(col_l, "I max PV per MPPT:", self.var_max_pv_current, 120)
        _lbl_entry(col_l, "I cc max per MPPT:", self.var_max_dc_sc, 120)
        _lbl_entry(col_l, "Range tensione MPPT:", self.var_mppt_vrange, 140)
        _lbl_entry(col_l, "Vdc max:", self.var_max_vdc, 120)

        # Right column
        # Panel model
        f_pan = ctk.CTkFrame(col_r, fg_color="transparent")
        f_pan.pack(fill="x", pady=4)
        ctk.CTkLabel(f_pan, text="Modello Pannello:", text_color=_TEXT, anchor="w",
                      font=("Segoe UI", 12), width=170).pack(side="left")
        self.cb_panel = ctk.CTkComboBox(f_pan, variable=self.var_panel_model,
                                         values=_PANELS_PRESETS,
                                         height=32, corner_radius=6,
                                         fg_color="white", border_color=_BORDER,
                                         font=("Segoe UI", 11))
        self.cb_panel.pack(side="left", fill="x", expand=True, padx=(0, 4))
        ctk.CTkButton(f_pan, text="+", width=32, height=32, corner_radius=6,
                       fg_color="#E8EAED", text_color=_BLUE, hover_color="#C8D4EC",
                       font=("Segoe UI", 13, "bold"),
                       command=self._add_new_panel).pack(side="left", padx=(0, 2))
        ctk.CTkButton(f_pan, text="...", width=32, height=32, corner_radius=6,
                       fg_color="#E8EAED", text_color=_MUTED, hover_color="#DADCE0",
                       font=("Segoe UI", 13),
                       command=self._manage_panels).pack(side="left")

        # Panels per string – free-text entry (editable)
        f_pps = ctk.CTkFrame(col_r, fg_color="transparent")
        f_pps.pack(fill="x", pady=4)
        ctk.CTkLabel(f_pps, text="Pannelli per Stringa:", text_color=_TEXT, anchor="w",
                      font=("Segoe UI", 12), width=170).pack(side="left")
        ctk.CTkEntry(f_pps, textvariable=self.var_panels, width=80,
                      height=32, corner_radius=6,
                      fg_color="white", border_color=_BORDER,
                      font=("Segoe UI", 11)).pack(side="left")

        # Panel model map – dynamic per-Wp fields, auto-populated when Excel is loaded
        f_pmap_hdr = ctk.CTkFrame(col_r, fg_color="transparent")
        f_pmap_hdr.pack(fill="x", pady=(6, 0))
        ctk.CTkLabel(f_pmap_hdr, text="Modelli per Wp:", text_color=_TEXT, anchor="w",
                      font=("Segoe UI", 12), width=170).pack(side="left")
        ctk.CTkLabel(f_pmap_hdr, text="(dal file Excel)",
                      text_color=_MUTED, font=("Segoe UI", 10)).pack(side="left")
        self._wp_model_frame = ctk.CTkFrame(col_r, fg_color="#F3F4F6", corner_radius=6,
                                             border_width=1, border_color=_BORDER)
        self._wp_model_frame.pack(fill="x", pady=(2, 4))
        ctk.CTkLabel(self._wp_model_frame,
                     text="Carica il file Excel per rilevare i tipi di pannello.",
                     text_color=_MUTED, font=("Segoe UI", 10),
                     anchor="w").pack(padx=8, pady=6, anchor="w")

        # Strings per MPPT – number of string rows drawn per MPPT channel
        f_spm = ctk.CTkFrame(col_r, fg_color="transparent")
        f_spm.pack(fill="x", pady=4)
        ctk.CTkLabel(f_spm, text="Stringhe per MPPT:", text_color=_TEXT, anchor="w",
                      font=("Segoe UI", 12), width=170).pack(side="left")
        ctk.CTkEntry(f_spm, textvariable=self.var_strings_mppt, width=80,
                      height=32, corner_radius=6,
                      fg_color="white", border_color=_BORDER,
                      font=("Segoe UI", 11)).pack(side="left")

        # MPPTs per DC switch – how many MPPTs grouped under each DC-switch label
        f_mps = ctk.CTkFrame(col_r, fg_color="transparent")
        f_mps.pack(fill="x", pady=4)
        ctk.CTkLabel(f_mps, text="MPPT per Sezionatore DC:", text_color=_TEXT, anchor="w",
                      font=("Segoe UI", 12), width=170).pack(side="left")
        ctk.CTkEntry(f_mps, textvariable=self.var_mppts_switch, width=80,
                      height=32, corner_radius=6,
                      fg_color="white", border_color=_BORDER,
                      font=("Segoe UI", 11)).pack(side="left")

    # ── Step 3: Styles & geometry ─────────────────────────────────────────────
    def _build_step3(self):
        body = _card(self.scroll, "Step 3 - Stili e Geometrie (Avanzate)")
        ctk.CTkLabel(body, text="Impostazioni CAD avanzate. Modifica solo se necessario.",
                     font=("Segoe UI", 11), text_color=_MUTED, anchor="w").pack(anchor="w", pady=(0, 8))

        # Toggles row 1
        tog = ctk.CTkFrame(body, fg_color="transparent")
        tog.pack(fill="x", pady=4)
        ctk.CTkLabel(tog, text="Includi Lunghezze Cavi:", text_color=_TEXT,
                      font=("Segoe UI", 12), width=200).pack(side="left")
        ctk.CTkSwitch(tog, variable=self.var_show_cable, text="",
                       onvalue=True, offvalue=False,
                       progress_color=_BLUE).pack(side="left", padx=(0, 30))
        ctk.CTkLabel(tog, text="Nascondi Dettagli Stringa:", text_color=_TEXT,
                      font=("Segoe UI", 12)).pack(side="left")
        ctk.CTkSwitch(tog, variable=self.var_hide_string_details, text="",
                       onvalue=True, offvalue=False,
                       progress_color=_BLUE).pack(side="left")

        # Toggles row 2
        tog2 = ctk.CTkFrame(body, fg_color="transparent")
        tog2.pack(fill="x", pady=4)
        ctk.CTkLabel(tog2, text="Mostra cerchio annotazione:", text_color=_TEXT,
                      font=("Segoe UI", 12), width=200).pack(side="left")
        ctk.CTkSwitch(tog2, variable=self.var_show_annot_circle, text="",
                       onvalue=True, offvalue=False,
                       progress_color=_BLUE).pack(side="left", padx=(0, 30))
        ctk.CTkLabel(tog2,
                      text="(cerchio rosso nel diagramma pannelli — disattiva se non necessario)",
                      text_color=_MUTED, font=("Segoe UI", 11)).pack(side="left")

        # Grid spacing
        g = ctk.CTkFrame(body, fg_color="transparent")
        g.pack(fill="x", pady=4)
        col_l = ctk.CTkFrame(g, fg_color="transparent")
        col_l.pack(side="left", fill="both", expand=True, padx=(0, 10))
        col_r = ctk.CTkFrame(g, fg_color="transparent")
        col_r.pack(side="left", fill="both", expand=True, padx=(10, 0))

        def _pair(parent, label, var):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", pady=3)
            ctk.CTkLabel(f, text=label, text_color=_TEXT, anchor="w",
                          font=("Segoe UI", 12), width=170).pack(side="left")
            ctk.CTkEntry(f, textvariable=var, height=30, corner_radius=6,
                          fg_color="white", border_color=_BORDER,
                          font=("Segoe UI", 11), width=110).pack(side="left")

        _pair(col_l, "Step Colonne:", self.var_col_spacing)
        _pair(col_l, "Step Righe:",   self.var_row_spacing)
        _pair(col_l, "Raggio cerchio:", self.var_circle_radius)
        _pair(col_l, "Dim. testo:",   self.var_text_size)

        _pair(col_r, "Sezione heavy (es. 1x10):", self.var_heavy_section)
        _pair(col_r, "Linetype:",     self.var_heavy_linetype)
        _pair(col_r, "Colore (indice):", self.var_heavy_color)
        _pair(col_r, "Layer heavy:", self.var_heavy_layer)

    # ── Step 4: MPPT layout ───────────────────────────────────────────────────
    def _build_step4(self):
        body = _card(self.scroll, "Step 4 - Layout MPPT (Opzionale)")
        ctk.CTkLabel(body,
                     text="Clicca su uno slot per alternare ATTIVO / RISERVA. "
                          "Lascia tutti ATTIVI per rilevamento automatico dal template.",
                     font=("Segoe UI", 11), text_color=_MUTED, anchor="w").pack(anchor="w", pady=(0, 8))

        # Controls row
        ctrl = ctk.CTkFrame(body, fg_color="transparent")
        ctrl.pack(fill="x", pady=(0, 6))

        def _ctrl_entry(lbl, var, w=60):
            ctk.CTkLabel(ctrl, text=lbl, text_color=_TEXT,
                         font=("Segoe UI", 12)).pack(side="left", padx=(0, 4))
            e = ctk.CTkEntry(ctrl, textvariable=var, width=w, height=30, corner_radius=6,
                             fg_color="white", border_color=_BORDER, font=("Segoe UI", 11))
            e.pack(side="left", padx=(0, 16))

        _ctrl_entry("N. MPPTs:", self.var_num_mppts_layout, 60)
        _ctrl_entry("Slot/MPPT:", self.var_strings_mppt, 60)
        _ctrl_entry("MPPTs/Switch:", self.var_mppts_switch, 60)

        ctk.CTkButton(ctrl, text="Aggiorna Griglia", height=30, corner_radius=6,
                       fg_color=_BLUE, hover_color=_BLUE_H, text_color="white",
                       font=("Segoe UI", 11), width=130,
                       command=self._rebuild_mppt_grid).pack(side="left", padx=(0, 6))
        ctk.CTkButton(ctrl, text="Reset tutto ATTIVO", height=30, corner_radius=6,
                       fg_color="#E8EAED", hover_color="#DADCE0", text_color=_TEXT,
                       font=("Segoe UI", 11), width=140,
                       command=self._reset_mppt_grid).pack(side="left")

        # Scrollable grid frame
        self._mppt_grid_frame = ctk.CTkScrollableFrame(body, fg_color="#F1F3F4",
                                                        corner_radius=8, height=280)
        self._mppt_grid_frame.pack(fill="x", pady=(4, 0))

        # Build initial grid from current layout var
        self._rebuild_mppt_grid()

    def _rebuild_mppt_grid(self):
        import json
        for w in self._mppt_grid_frame.winfo_children():
            w.destroy()
        self._mppt_slot_btns.clear()
        self._mppt_slot_state.clear()

        try:
            n_mppts = max(1, int(self.var_num_mppts_layout.get() or 6))
            n_slots  = max(1, int(self.var_strings_mppt.get() or 5))
        except ValueError:
            return

        # Load existing layout
        existing = {}
        try:
            raw = self.var_mppt_layout.get().strip()
            if raw:
                existing = {int(k): list(v) for k, v in json.loads(raw).items()}
        except Exception:
            pass

        # Column headers
        hdr = ctk.CTkFrame(self._mppt_grid_frame, fg_color="transparent")
        hdr.pack(fill="x", padx=4, pady=(4, 0))
        ctk.CTkLabel(hdr, text="", width=72).pack(side="left")
        for p in range(1, n_slots + 1):
            ctk.CTkLabel(hdr, text=f"Slot {p}", width=82,
                         font=("Segoe UI", 11, "bold"), text_color=_MUTED).pack(side="left", padx=2)

        # MPPT rows
        for m in range(1, n_mppts + 1):
            row_f = ctk.CTkFrame(self._mppt_grid_frame, fg_color="transparent")
            row_f.pack(fill="x", padx=4, pady=2)
            ctk.CTkLabel(row_f, text=f"MPP{m}", width=72,
                         font=("Segoe UI", 12, "bold"), text_color=_TEXT).pack(side="left")
            ex_row = existing.get(m, [])
            for p in range(1, n_slots + 1):
                state = 'R' if (p - 1 < len(ex_row) and ex_row[p - 1] == 'R') else 'A'
                self._mppt_slot_state[(m, p)] = state
                btn = ctk.CTkButton(
                    row_f,
                    text="RISERVA" if state == 'R' else "ATTIVO",
                    width=80, height=28, corner_radius=6,
                    fg_color=_ERROR if state == 'R' else _BLUE,
                    hover_color="#C0392B" if state == 'R' else _BLUE_H,
                    text_color="white", font=("Segoe UI", 10),
                    command=lambda _m=m, _p=p: self._toggle_mppt_slot(_m, _p))
                btn.pack(side="left", padx=2)
                self._mppt_slot_btns[(m, p)] = btn

        self._serialize_mppt_layout()

    def _toggle_mppt_slot(self, m, p):
        cur = self._mppt_slot_state.get((m, p), 'A')
        new = 'R' if cur == 'A' else 'A'
        self._mppt_slot_state[(m, p)] = new
        btn = self._mppt_slot_btns.get((m, p))
        if btn:
            btn.configure(
                text="RISERVA" if new == 'R' else "ATTIVO",
                fg_color=_ERROR if new == 'R' else _BLUE,
                hover_color="#C0392B" if new == 'R' else _BLUE_H)
        self._serialize_mppt_layout()

    def _reset_mppt_grid(self):
        for (m, p), btn in self._mppt_slot_btns.items():
            self._mppt_slot_state[(m, p)] = 'A'
            btn.configure(text="ATTIVO", fg_color=_BLUE, hover_color=_BLUE_H)
        self._serialize_mppt_layout()

    def _serialize_mppt_layout(self):
        import json
        if not self._mppt_slot_state or not any(v == 'R' for v in self._mppt_slot_state.values()):
            self.var_mppt_layout.set('')
            return
        try:
            n_mppts = max(1, int(self.var_num_mppts_layout.get() or 6))
            n_slots  = max(1, int(self.var_strings_mppt.get() or 5))
        except ValueError:
            return
        layout = {}
        for m in range(1, n_mppts + 1):
            act = 1
            row = []
            for p in range(1, n_slots + 1):
                if self._mppt_slot_state.get((m, p), 'A') == 'R':
                    row.append('R')
                else:
                    row.append(act)
                    act += 1
            layout[str(m)] = row
        self.var_mppt_layout.set(json.dumps(layout))

    # ── Add-new preset dialogs ────────────────────────────────────────────────
    def _add_new_panel(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Aggiungi Modello Pannello")
        dlg.geometry("400x170")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(fg_color=_BG)

        ctk.CTkLabel(dlg, text="Modello Pannello:", font=("Segoe UI", 12),
                      text_color=_TEXT).grid(row=0, column=0, sticky="w", padx=16, pady=12)
        e_model = ctk.CTkEntry(dlg, width=220, height=32, corner_radius=6,
                                fg_color="white", border_color=_BORDER)
        e_model.grid(row=0, column=1, padx=10, pady=12, sticky="ew")
        e_model.focus_set()

        ctk.CTkLabel(dlg, text="(Wp letto dall'Excel)",
                      font=("Segoe UI", 10, "italic"), text_color=_MUTED
                      ).grid(row=1, column=0, columnspan=2, sticky="w", padx=16)

        def _confirm():
            model = e_model.get().strip()
            if not model:
                return
            if model not in _PANELS_PRESETS:
                _PANELS_PRESETS.append(model)
            self.cb_panel.configure(values=_PANELS_PRESETS)
            self.var_panel_model.set(model)
            self._save_custom_presets()
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.grid(row=2, column=0, columnspan=2, pady=14)
        ctk.CTkButton(btn_row, text="Aggiungi", fg_color=_BLUE, hover_color=_BLUE_H,
                       text_color="white", corner_radius=16, width=100,
                       command=_confirm).pack(side="left", padx=8)
        ctk.CTkButton(btn_row, text="Annulla", fg_color="#E8EAED", text_color=_TEXT,
                       hover_color="#DADCE0", corner_radius=16, width=80,
                       command=dlg.destroy).pack(side="left", padx=8)
        dlg.bind("<Return>", lambda _: _confirm())

    def _add_new_inverter(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Aggiungi Modello Inverter")
        dlg.geometry("430x230")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(fg_color=_BG)

        labels = ["Modello Inverter:", "Potenza DC (kWp):", "Potenza AC (kWac):"]
        entries = []
        for i, lbl in enumerate(labels):
            ctk.CTkLabel(dlg, text=lbl, font=("Segoe UI", 12),
                          text_color=_TEXT).grid(row=i, column=0, sticky="w", padx=16, pady=8)
            e = ctk.CTkEntry(dlg, width=210, height=32, corner_radius=6,
                              fg_color="white", border_color=_BORDER)
            e.grid(row=i, column=1, padx=10, pady=8, sticky="ew")
            entries.append(e)
        entries[0].focus_set()

        def _confirm():
            model = entries[0].get().strip()
            dc    = entries[1].get().strip()
            ac    = entries[2].get().strip()
            if not model:
                return
            _INVERTER_POWERS[model] = (dc, ac)
            self.cb_inv.configure(values=sorted(_INVERTER_POWERS.keys()))
            self.var_inv_model.set(model)
            if dc: self.var_dc_power.set(dc)
            if ac: self.var_ac_power.set(ac)
            self._save_custom_presets()
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.grid(row=3, column=0, columnspan=2, pady=14)
        ctk.CTkButton(btn_row, text="Aggiungi", fg_color=_BLUE, hover_color=_BLUE_H,
                       text_color="white", corner_radius=16, width=100,
                       command=_confirm).pack(side="left", padx=8)
        ctk.CTkButton(btn_row, text="Annulla", fg_color="#E8EAED", text_color=_TEXT,
                       hover_color="#DADCE0", corner_radius=16, width=80,
                       command=dlg.destroy).pack(side="left", padx=8)
        dlg.bind("<Return>", lambda _: _confirm())


    # ── Manage existing presets ───────────────────────────────────────────────

    def _save_custom_presets(self):
        """Persist only the user-added (non built-in) presets to history."""
        _BUILTIN_INV = {
            'Sungrow SG100HX','Sungrow SG125HX','Sungrow SG250HX','Sungrow SG350HX','Sungrow SG500HX',
            'Huawei SUN2000-100KTL-M3','Huawei SUN2000-215KTL-H3','Huawei SUN2000-275KTL-H1',
            'Huawei SUN2000-330KTL','Huawei SUN2000-450KTL-H1',
            'ABB PVS-120-TL','ABB PVS-250-TL','ABB PVS-350-TL','ABB PVS-500-TL',
            'SMA Sunny Tripower Core1 25','SMA Sunny Tripower Core2 150',
            'SMA Sunny Tripower 25000TL','SMA Sunny Tripower 60000TL','SMA Sunny Tripower 100000TL',
            'Fronius Symo GEN24 25.0 Plus','Fronius Symo GEN24 50.0 Plus',
            'Fronius Symo GEN24 60.0 Plus','Fronius Symo GEN24 100.0 Plus',
            'KACO blueplanet 100.0 TL3','KACO blueplanet 125.0 TL3','KACO blueplanet 250.0 TL3',
            'Growatt 50000MT','Growatt 60000MT','Growatt 100000MT',
        }
        _BUILTIN_PAN = {
            'JA Solar JAM72D42-625/LB','JA Solar JAM72S20-460/MR',
            'Longi Solar LR5-72HBD-580M','Longi Solar LR5-72HBD-545M',
            'Canadian Solar HiKu7 CS7N-655MB','Jinko Solar JKM660M-78HL4-V',
            'Trina Solar TSM-670NEG21C.20','REC Alpha Pure-R 430AA',
        }
        custom_inv = {k: list(v) for k, v in _INVERTER_POWERS.items() if k not in _BUILTIN_INV}
        custom_pan = [p for p in _PANELS_PRESETS if p not in _BUILTIN_PAN]
        h = load_history()
        h['custom_inverters'] = custom_inv
        h['custom_panels'] = custom_pan
        save_history(h)

    def _manage_panels(self):
        """Open a dialog to edit or delete panel presets."""
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Gestisci Modelli Pannello")
        dlg.geometry("520x420")
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.configure(fg_color=_BG)

        ctk.CTkLabel(dlg, text="Modelli Pannello", font=("Segoe UI", 14, "bold"),
                      text_color=_BLUE).pack(anchor="w", padx=16, pady=(14, 4))

        scroll = ctk.CTkScrollableFrame(dlg, fg_color=_CARD, corner_radius=8)
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        def _refresh():
            for w in scroll.winfo_children():
                w.destroy()
            for i, name in enumerate(_PANELS_PRESETS):
                row = ctk.CTkFrame(scroll, fg_color="transparent")
                row.pack(fill="x", pady=2)
                ctk.CTkLabel(row, text=name, anchor="w", font=("Segoe UI", 12),
                              text_color=_TEXT).pack(side="left", fill="x", expand=True, padx=6)

                def _edit(n=name, idx=i):
                    ed = ctk.CTkToplevel(dlg)
                    ed.title("Modifica Pannello")
                    ed.geometry("380x120")
                    ed.resizable(False, False)
                    ed.grab_set()
                    ed.configure(fg_color=_BG)
                    ctk.CTkLabel(ed, text="Nome:", font=("Segoe UI", 12)).grid(row=0, column=0, sticky="w", padx=12, pady=12)
                    e = ctk.CTkEntry(ed, width=230, height=30, corner_radius=6,
                                     fg_color="white", border_color=_BORDER)
                    e.insert(0, n)
                    e.grid(row=0, column=1, padx=8, pady=12)
                    def _ok():
                        new = e.get().strip()
                        if new and new != n:
                            _PANELS_PRESETS[idx] = new
                        ed.destroy()
                        _refresh()
                        self.cb_panel.configure(values=_PANELS_PRESETS)
                        self._save_custom_presets()
                    ctk.CTkButton(ed, text="OK", fg_color=_BLUE, text_color="white",
                                   corner_radius=14, width=80, command=_ok).grid(row=1, column=0, columnspan=2, pady=8)
                    ed.bind("<Return>", lambda _: _ok())

                def _delete(n=name):
                    if n in _PANELS_PRESETS:
                        _PANELS_PRESETS.remove(n)
                    _refresh()
                    self.cb_panel.configure(values=_PANELS_PRESETS)
                    self._save_custom_presets()

                ctk.CTkButton(row, text="...", width=30, height=28, corner_radius=6,
                               fg_color="#E8EAED", text_color=_MUTED, hover_color="#DADCE0",
                               command=_edit).pack(side="left", padx=2)
                ctk.CTkButton(row, text="X", width=30, height=28, corner_radius=6,
                               fg_color="#FCE8E6", text_color=_ERROR, hover_color="#F5C6C2",
                               command=_delete).pack(side="left", padx=(0, 4))

        _refresh()
        ctk.CTkButton(dlg, text="Chiudi", fg_color="#E8EAED", text_color=_TEXT,
                       hover_color="#DADCE0", corner_radius=14, width=90,
                       command=dlg.destroy).pack(pady=(0, 14))

    def _manage_inverters(self):
        """Open a dialog to edit or delete inverter presets."""
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("Gestisci Modelli Inverter")
        dlg.geometry("600x440")
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.configure(fg_color=_BG)

        ctk.CTkLabel(dlg, text="Modelli Inverter", font=("Segoe UI", 14, "bold"),
                      text_color=_BLUE).pack(anchor="w", padx=16, pady=(14, 4))

        scroll = ctk.CTkScrollableFrame(dlg, fg_color=_CARD, corner_radius=8)
        scroll.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        def _refresh():
            for w in scroll.winfo_children():
                w.destroy()
            for name in sorted(_INVERTER_POWERS.keys()):
                dc, ac = _INVERTER_POWERS[name]
                row = ctk.CTkFrame(scroll, fg_color="transparent")
                row.pack(fill="x", pady=2)
                ctk.CTkLabel(row, text=f"{name}  ({dc} kWp / {ac} kWac)",
                              anchor="w", font=("Segoe UI", 11),
                              text_color=_TEXT).pack(side="left", fill="x", expand=True, padx=6)

                def _edit(n=name):
                    cur_dc, cur_ac = _INVERTER_POWERS[n]
                    ed = ctk.CTkToplevel(dlg)
                    ed.title("Modifica Inverter")
                    ed.geometry("400x200")
                    ed.resizable(False, False)
                    ed.grab_set()
                    ed.configure(fg_color=_BG)
                    entries = []
                    for r, (lbl, val) in enumerate([("Nome:", n), ("DC (kWp):", cur_dc), ("AC (kWac):", cur_ac)]):
                        ctk.CTkLabel(ed, text=lbl, font=("Segoe UI", 12)).grid(row=r, column=0, sticky="w", padx=12, pady=7)
                        e = ctk.CTkEntry(ed, width=230, height=30, corner_radius=6,
                                         fg_color="white", border_color=_BORDER)
                        e.insert(0, val)
                        e.grid(row=r, column=1, padx=8, pady=7)
                        entries.append(e)
                    def _ok():
                        new_name = entries[0].get().strip()
                        new_dc   = entries[1].get().strip()
                        new_ac   = entries[2].get().strip()
                        if new_name:
                            if new_name != n and n in _INVERTER_POWERS:
                                del _INVERTER_POWERS[n]
                            _INVERTER_POWERS[new_name] = (new_dc, new_ac)
                        ed.destroy()
                        _refresh()
                        self.cb_inv.configure(values=sorted(_INVERTER_POWERS.keys()))
                        self._save_custom_presets()
                    ctk.CTkButton(ed, text="OK", fg_color=_BLUE, text_color="white",
                                   corner_radius=14, width=80, command=_ok).grid(row=3, column=0, columnspan=2, pady=10)
                    ed.bind("<Return>", lambda _: _ok())

                def _delete(n=name):
                    if n in _INVERTER_POWERS:
                        del _INVERTER_POWERS[n]
                    _refresh()
                    self.cb_inv.configure(values=sorted(_INVERTER_POWERS.keys()))
                    self._save_custom_presets()

                ctk.CTkButton(row, text="...", width=30, height=28, corner_radius=6,
                               fg_color="#E8EAED", text_color=_MUTED, hover_color="#DADCE0",
                               command=_edit).pack(side="left", padx=2)
                ctk.CTkButton(row, text="X", width=30, height=28, corner_radius=6,
                               fg_color="#FCE8E6", text_color=_ERROR, hover_color="#F5C6C2",
                               command=_delete).pack(side="left", padx=(0, 4))

        _refresh()
        ctk.CTkButton(dlg, text="Chiudi", fg_color="#E8EAED", text_color=_TEXT,
                       hover_color="#DADCE0", corner_radius=14, width=90,
                       command=dlg.destroy).pack(pady=(0, 14))

    # ── File browse helpers ───────────────────────────────────────────────────
    def _browse_excel(self):

        if HAS_TK:
            from tkinter import filedialog
            p = filedialog.askopenfilename(
                filetypes=[("File Excel", "*.xlsx *.xls"), ("Tutti i file", "*.*")])
            if p: self.var_excel.set(p)

    def _browse_template(self):
        if HAS_TK:
            from tkinter import filedialog
            p = filedialog.askopenfilename(
                filetypes=[("File DXF", "*.dxf"), ("Tutti i file", "*.*")])
            if p: self.var_template.set(p)

    def _browse_output(self):
        if HAS_TK:
            from tkinter import filedialog
            p = filedialog.asksaveasfilename(
                defaultextension=".dxf",
                filetypes=[("File DXF", "*.dxf"), ("Tutti i file", "*.*")])
            if p: self.var_out.set(p)

    # ── Log helpers ───────────────────────────────────────────────────────────
    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("0.0", "end")
        self.log_box.configure(state="disabled")

    def log_message(self, msg):
        def _append():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.root.after(0, _append)

    # ── Generation ────────────────────────────────────────────────────────────
    def _start_generation(self):
        cfg = {
            'template_dxf':       self.var_template.get().strip(),
            'xlsx_path':          self.var_excel.get().strip(),
            'output_path':        self.var_out.get().strip(),
            'panel_model':        self.var_panel_model.get().strip(),
            'panel_model_map':    self._build_panel_map_string(),
            'panels_per_string':  self.var_panels.get().strip(),
            'strings_per_mppt':   self.var_strings_mppt.get().strip(),
            'mppts_per_switch':   self.var_mppts_switch.get().strip(),
            'num_mppts_layout':   self.var_num_mppts_layout.get().strip(),
            'mppt_layout':        self.var_mppt_layout.get().strip(),
            'inverter_model':     self.var_inv_model.get().strip(),
            'dc_power_kwp':       self.var_dc_power.get().strip(),
            'ac_power_kwac':      self.var_ac_power.get().strip(),
            'ac_power_30c':       self.var_ac_power_30c.get().strip(),
            'max_ac_current':     self.var_max_ac_current.get().strip(),
            'max_pv_current_per_mppt':    self.var_max_pv_current.get().strip(),
            'max_dc_sc_current_per_mppt': self.var_max_dc_sc.get().strip(),
            'mppt_voltage_range': self.var_mppt_vrange.get().strip(),
            'max_vdc':            self.var_max_vdc.get().strip(),
            'temp_rating':        self.var_temp.get().strip(),
            'transformer_power':  '',
            'show_cable_info':    self.var_show_cable.get(),
            'hide_string_details':self.var_hide_string_details.get(),
            'show_annot_circle':  self.var_show_annot_circle.get(),
            'col_spacing':        self.var_col_spacing.get().strip(),
            'row_spacing':        self.var_row_spacing.get().strip(),
            'circle_radius':      self.var_circle_radius.get().strip(),
            'text_size':          self.var_text_size.get().strip(),
            'heavy_section':      self.var_heavy_section.get().strip(),
            'heavy_linetype':     self.var_heavy_linetype.get().strip(),
            'heavy_color':        self.var_heavy_color.get().strip(),
            'heavy_layer':        self.var_heavy_layer.get().strip(),
        }

        errors = []
        if not os.path.isfile(cfg['template_dxf']):
            errors.append(f"File template DXF non trovato:\n  {cfg['template_dxf']}")
        if not os.path.isfile(cfg['xlsx_path']):
            errors.append(f"File Excel non trovato:\n  {cfg['xlsx_path']}")
        if not cfg['output_path']:
            errors.append("Il percorso del file DXF di output è obbligatorio.")
        for key in ('col_spacing','row_spacing','circle_radius','text_size','heavy_color',
                    'panels_per_string','strings_per_mppt','mppts_per_switch'):
            val = cfg[key]
            if val:
                try:
                    float(val)
                except ValueError:
                    errors.append(f"Il campo '{key}' deve essere numerico (ricevuto '{val}').")

        if errors:
            if HAS_TK:
                from tkinter import messagebox as _mb
                _mb.showerror("Errore di Configurazione", "\n\n".join(errors), parent=self.root)
            return

        save_history(cfg)

        self.btn_generate.configure(state="disabled")
        self.progress.configure(mode="indeterminate")
        self.progress.start()
        self.lbl_status.configure(text="Generazione in corso…", text_color=_BLUE)
        self._clear_log()

        def worker():
            try:
                generate(cfg, log_cb=self.log_message)
                self.root.after(0, self._on_success)
            except Exception as ex:
                import traceback
                self.log_message(f"\n[FATAL ERROR] {ex}\n{traceback.format_exc()}")
                self.root.after(0, self._on_error, str(ex))

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self):
        self.progress.stop()
        self.progress.set(1)
        self.btn_generate.configure(state="normal")
        self.lbl_status.configure(text="✓  Generazione completata.", text_color=_SUCCESS)
        if HAS_TK:
            from tkinter import messagebox as _mb
            _mb.showinfo("Successo",
                         f"Schema SLD generato con successo!\n\nSalvato in:\n{self.var_out.get()}",
                         parent=self.root)

    def _on_error(self, err_msg):
        self.progress.stop()
        self.progress.set(0)
        self.btn_generate.configure(state="normal")
        self.lbl_status.configure(text="✗  Errore di generazione.", text_color=_ERROR)
        if HAS_TK:
            from tkinter import messagebox as _mb
            _mb.showerror("Errore di Generazione",
                          f"Impossibile generare lo schema SLD:\n\n{err_msg}",
                          parent=self.root)

    def _run_fallback_tk(self):
        """Minimal tkinter fallback if CTk is unavailable."""
        import tkinter as _tk
        root = _tk.Tk()
        root.title("A176LAB - SLD Generator (fallback)")
        _tk.Label(root, text="customtkinter non disponibile.\nInstalla con: pip install customtkinter",
                  pady=40, padx=60).pack()
        root.mainloop()

    def mainloop(self):
        self.root.mainloop()



# ─────────────────────────────────────────────────────────────────────────────
#  CLI SUPPORT & MAIN ROUTINE
# ─────────────────────────────────────────────────────────────────────────────

def build_parser():
    p = argparse.ArgumentParser(
        description="Smart CLI Generator for DC Single Line Diagrams (DXF)"
    )
    p.add_argument('--template', help="Path to the template DXF file")
    p.add_argument('--excel', help="Path to the Excel cable schedule")
    p.add_argument('--out', help="Path to save the generated DXF file (defaults to <excel_name>_SLD_Generated.dxf)")
    p.add_argument('--panels', default='26', help="Panels per string (default: 28)")
    p.add_argument('--strings-per-mppt', default='2', help="String rows drawn per MPPT channel (default: 2)")
    p.add_argument('--mppts-per-switch', default='4', help="MPPTs grouped under each DC-switch label (default: 4)")
    p.add_argument('--panel-model', default='', help="Solar panel model name (default: '')")
    p.add_argument('--panel-model-map', default='', help="Per-watt model override for mixed-panel inverters, e.g. '720=Model A; 725=Model B'")
    p.add_argument('--inv-model', default='Sungrow SG350HX', help="Inverter model name (default: 'Sungrow SG350HX')")
    p.add_argument('--dc-power', default='350', help="Inverter DC power rating in KWp (default: '350')")
    p.add_argument('--ac-power', default='320', help="Inverter AC power rating in KWac (default: '320')")
    p.add_argument('--temp', default='40', help="Inverter ambient temperature rating in °C (default: '40')")
    p.add_argument('--tx-power', default='', help="Transformer power text (default: '')")
    p.add_argument('--show-cable', action='store_true', help="Include cable length/section details on labels")
    p.add_argument('--hide-string-details', action='store_true', help="Do not append panels and module type details to string labels")
    p.add_argument('--col-spacing', default=str(COL_SPACING_DEFAULT), help="Horizontal stamp spacing column step")
    p.add_argument('--row-spacing', default=str(ROW_SPACING_DEFAULT), help="Vertical stamp spacing row step")
    p.add_argument('--circle-radius', default='24.59', help="Module circle terminal radius")
    p.add_argument('--text-size', default='60.44', help="String label text size")
    p.add_argument('--heavy-section', default='1x10', help="String cable section to style as heavy (default: '1x10')")
    p.add_argument('--heavy-linetype', default='TRATTEGGIATA', help="AutoCAD linetype to apply to heavy cable sections")
    p.add_argument('--heavy-color', default='40', help="AutoCAD color index for heavy sections (default: 40)")
    p.add_argument('--heavy-layer', default='TRATTEGGIATA', help="Layer where heavy cable runs will be placed")
    return p


def main():
    parser = build_parser()
    args, unknown = parser.parse_known_args()

    # If no parameters (or specifically excel and template) are specified, start the GUI
    if not args.excel or not args.template:
        if not HAS_TK:
            print("[ERROR] Arguments --excel and --template are required in headless CLI mode.", file=sys.stderr)
            print("To run in GUI mode, make sure tkinter is installed and run with no arguments.", file=sys.stderr)
            sys.exit(1)
        print("Launching Graphical User Interface (GUI)...")
        app = SmartSLDGui()
        app.mainloop()
        return

    # Run in CLI mode
    if not args.out:
        args.out = os.path.splitext(args.excel)[0] + '_SLD_Generated.dxf'

    cfg = {
        'template_dxf':      args.template,
        'xlsx_path':         args.excel,
        'output_path':       args.out,
        'panel_model':       args.panel_model,
        'panel_model_map':   args.panel_model_map,
        'panels_per_string': args.panels,
        'strings_per_mppt':  args.strings_per_mppt,
        'mppts_per_switch':  args.mppts_per_switch,
        'inverter_model':    args.inv_model,
        'dc_power_kwp':      args.dc_power,
        'ac_power_kwac':     args.ac_power,
        'temp_rating':       args.temp,
        'transformer_power': args.tx_power,
        'show_cable_info':   args.show_cable,
        'hide_string_details': args.hide_string_details,
        'col_spacing':       args.col_spacing,
        'row_spacing':       args.row_spacing,
        'circle_radius':     args.circle_radius,
        'text_size':         args.text_size,
        'heavy_section':     args.heavy_section,
        'heavy_linetype':    args.heavy_linetype,
        'heavy_color':       args.heavy_color,
        'heavy_layer':       args.heavy_layer,
    }

    # Verify input paths
    if not os.path.isfile(cfg['template_dxf']):
        print(f"[ERROR] Template file does not exist: {cfg['template_dxf']}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(cfg['xlsx_path']):
        print(f"[ERROR] Excel file does not exist: {cfg['xlsx_path']}", file=sys.stderr)
        sys.exit(1)

    # Validate numeric fields
    num_fields = (
        ('panels_per_string', 'Panels per String'),
        ('strings_per_mppt',  'Strings per MPPT'),
        ('mppts_per_switch',  'MPPTs per DC Switch'),
        ('dc_power_kwp',      'DC Power'),
        ('ac_power_kwac',     'AC Power'),
        ('temp_rating',       'Temperature Rating'),
        ('col_spacing',       'Horizontal Col Step'),
        ('row_spacing',       'Vertical Row Step'),
        ('circle_radius',     'Module Circle Radius'),
        ('text_size',         'String Label Text Size'),
        ('heavy_color',       'Heavy Cable Run Color Index'),
    )
    for key, label in num_fields:
        val = cfg.get(key, '')
        if val:
            try:
                float(val)
            except ValueError:
                print(f"[ERROR] {label}: must be a number (got '{val}').", file=sys.stderr)
                sys.exit(1)

    print("=" * 72)
    print("Starting Headless Smart DC SLD Generation...")
    print("=" * 72)
    try:
        generate(cfg)
    except Exception as ex:
        import traceback
        print(f"\n[FATAL ERROR] Generation failed: {ex}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
