# -*- coding: utf-8 -*-
"""
sld_gui_v2.py  -  Upgraded GUI front-end and CAD engine for the DC Single Line Diagram generator.
Run:  python sld_gui_v2.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import threading
import os
import re
import sys
import json
from collections import defaultdict
from datetime import datetime

_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logoA176LAB.jpg')


def _load_logo_tk(size=(90, 90)):
    """Return a tkinter PhotoImage of the logo, or None if PIL is unavailable."""
    try:
        from PIL import Image, ImageTk
        img = Image.open(_LOGO_PATH).convert('RGBA')
        img.thumbnail(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  GENERATION — CONSTANTS & DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────

# Y-band in the template DXF that contains Inverter 1.1
_TMPL_Y_MIN = 159_400
_TMPL_Y_MAX = 168_000

# Original column/row step measured from the template
_COL_STEP = 11_740
_ROW_STEP = 10_200

# Spacing between stamped inverter copies (slightly larger than template dims)
_COL_SPACING_DEFAULT = int(_COL_STEP * 1.22)
_ROW_SPACING_DEFAULT = int(_ROW_STEP * 1.18)

# Y-proximity (drawing units) for matching port labels to string-label slots.
_PORT_Y_TOL = 400

# Template bottom-right placeholder rectangle filter.
_PLACEHOLDER_X_MIN = 23_500
_PLACEHOLDER_MIN_W = 500

# MTEXT text-box width override so long string labels never wrap
_STRING_LABEL_MIN_WIDTH = 4_500

PORT_RE   = re.compile(r'^\d+-\d+$')
STRING_RE = re.compile(r'String \d+\.\d+\.\d+')

# Inverter model → (DC KWp, AC KWac) for autofill
_INVERTER_POWERS = {
    'Sungrow SG100HX': ('100', '100'),
    'Sungrow SG125HX': ('125', '125'),
    'Sungrow SG250HX': ('250', '250'),
    'Sungrow SG350HX': ('350', '350'),
    'Sungrow SG500HX': ('500', '500'),
    'Huawei SUN2000-100KTL-M3': ('100', '100'),
    'Huawei SUN2000-215KTL-H3': ('215', '160'),
    'Huawei SUN2000-275KTL-H1': ('275', '220'),
    'Huawei SUN2000-330KTL': ('330', '275'),
    'Huawei SUN2000-450KTL-H1': ('450', '400'),
    'ABB PVS-120-TL': ('120', '120'),
    'ABB PVS-250-TL': ('250', '220'),
    'ABB PVS-350-TL': ('350', '280'),
    'ABB PVS-500-TL': ('500', '400'),
    'SMA Sunny Tripower Core1 25': ('25', '25'),
    'SMA Sunny Tripower Core2 150': ('150', '150'),
    'SMA Sunny Tripower 25000TL': ('25', '25'),
    'SMA Sunny Tripower 60000TL': ('60', '60'),
    'SMA Sunny Tripower 100000TL': ('100', '100'),
    'Fronius Symo GEN24 25.0 Plus': ('25', '25'),
    'Fronius Symo GEN24 50.0 Plus': ('50', '50'),
    'Fronius Symo GEN24 60.0 Plus': ('60', '60'),
    'Fronius Symo GEN24 100.0 Plus': ('100', '100'),
    'KACO blueplanet 100.0 TL3': ('100', '100'),
    'KACO blueplanet 125.0 TL3': ('125', '125'),
    'KACO blueplanet 250.0 TL3': ('250', '250'),
    'Growatt 50000MT': ('50', '50'),
    'Growatt 60000MT': ('60', '60'),
    'Growatt 100000MT': ('100', '100'),
}

# ─────────────────────────────────────────────────────────────────────────────
#  GENERATION — HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _strip_mtext_fmt(txt):
    """Strip AutoCAD inline MTEXT formatting codes to get plain text."""
    txt = re.sub(r'\\[A-Za-z][^;]*;', '', txt)
    return re.sub(r'\{[^}]*\}', '', txt).strip()


def _ent_y(e):
    """Return a representative Y coordinate for an entity, or None."""
    try:
        t = e.dxftype()
        if t in ('TEXT', 'MTEXT', 'INSERT'):  return e.dxf.insert.y
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return e.dxf.center.y
        if t == 'LINE':                        return e.dxf.start.y
        if t == 'LWPOLYLINE':
            pts = list(e.get_points())
            return pts[0][1] if pts else None
    except Exception:
        pass
    return None


def _ent_y_dict(d):
    """Return a representative Y coordinate for a dictionary entity representation."""
    t = d['type']
    if t in ('TEXT', 'MTEXT', 'INSERT'):  return d.get('y') or d.get('iy')
    if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return d.get('cy')
    if t == 'LINE':                        return d.get('sy')
    if t == 'LWPOLYLINE':
        pts = d.get('pts', [])
        return pts[0][1] if pts else None
    if t == 'POLYLINE':
        pts = d.get('pts3d', [])
        return pts[0][1] if pts else None
    return None


def _is_cable_line(d, sl_y):
    """True if d is the horizontal connection line at the Y-level of sl_y."""
    t = d['type']
    if t == 'LINE':
        length = abs(d['sx'] - d['ex'])
        is_horiz = abs(d['sy'] - d['ey']) < 5
        is_near_y = abs(d['sy'] - sl_y) < 80
        return is_horiz and is_near_y and length > 200
    elif t == 'LWPOLYLINE':
        xs = [p[0] for p in d['pts']]
        ys = [p[1] for p in d['pts']]
        if not xs or not ys:
            return False
        length = max(xs) - min(xs)
        is_horiz = (max(ys) - min(ys)) < 5
        is_near_y = abs(ys[0] - sl_y) < 80
        return is_horiz and is_near_y and length > 200
    elif t == 'POLYLINE':
        xs = [p[0] for p in d['pts3d']]
        ys = [p[1] for p in d['pts3d']]
        if not xs or not ys:
            return False
        length = max(xs) - min(xs)
        is_horiz = (max(ys) - min(ys)) < 5
        is_near_y = abs(ys[0] - sl_y) < 80
        return is_horiz and is_near_y and length > 200
    return False


def _common_attrs(e):
    d = {}
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if e.dxf.hasattr(a):
                d[a] = getattr(e.dxf, a)
        except Exception:
            pass
    return d


def _extract_entity(e, log):
    """Return a plain dict with all drawing data, or None for unsupported types."""
    t = e.dxftype()
    d = {'type': t}
    d.update(_common_attrs(e))
    try:
        if t == 'LWPOLYLINE':
            d['pts']    = [list(p) for p in e.get_points()]
            d['closed'] = e.closed
            if e.dxf.hasattr('const_width'):
                d['const_width'] = e.dxf.const_width
        elif t == 'MTEXT':
            pos = e.dxf.insert
            d.update({'x': pos.x, 'y': pos.y, 'text': e.text})
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'style'):
                try:
                    if e.dxf.hasattr(a):
                        d[a] = getattr(e.dxf, a)
                except Exception:
                    pass
        elif t == 'ARC':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'radius': e.dxf.radius,
                      'start_angle': e.dxf.start_angle,
                      'end_angle':   e.dxf.end_angle})
        elif t == 'CIRCLE':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z, 'radius': e.dxf.radius})
        elif t == 'LINE':
            s, en = e.dxf.start, e.dxf.end
            d.update({'sx': s.x, 'sy': s.y, 'sz': s.z,
                      'ex': en.x, 'ey': en.y, 'ez': en.z})
        elif t == 'ELLIPSE':
            c, ma = e.dxf.center, e.dxf.major_axis
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'major_axis': [ma.x, ma.y, ma.z],
                      'ratio': e.dxf.ratio,
                      'start_param': e.dxf.start_param,
                      'end_param':   e.dxf.end_param})
        elif t == 'INSERT':
            ins = e.dxf.insert
            d.update({'name': e.dxf.name,
                      'ix': ins.x, 'iy': ins.y,
                      'iz': ins.z if hasattr(ins, 'z') else 0.0})
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if e.dxf.hasattr(a):
                        d[a] = getattr(e.dxf, a)
                except Exception:
                    pass
        elif t == 'POLYLINE':
            d['pts3d'] = [[v.dxf.location.x, v.dxf.location.y, v.dxf.location.z]
                          for v in e.vertices]
        else:
            return None
    except Exception as ex:
        log(f"  [warn] extract {t}: {ex}")
        return None
    return d


def _entity_min_x(d):
    """Leftmost X coordinate for an extracted entity dict."""
    t = d['type']
    try:
        if t == 'LWPOLYLINE':                  return min(p[0] for p in d['pts'])
        if t == 'MTEXT':                       return d['x']
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return d['cx'] - d.get('radius', 0)
        if t == 'LINE':                        return min(d['sx'], d['ex'])
        if t == 'INSERT':                      return d['ix']
        if t == 'POLYLINE':                    return min(p[0] for p in d['pts3d'])
    except Exception:
        pass
    return 0


def _is_placeholder_rect(d):
    """True if d is the template's bottom-right placeholder rectangle artifact."""
    if d.get('type') != 'LWPOLYLINE' or not d.get('closed'):
        return False
    pts = d.get('pts', [])
    if len(pts) != 4:
        return False
    xs = [p[0] for p in pts]
    return min(xs) > _PLACEHOLDER_X_MIN and (max(xs) - min(xs)) > _PLACEHOLDER_MIN_W


def _classify_mtext(txt):
    c = _strip_mtext_fmt(txt).replace('\n', ' ')
    if re.search(r'INVERTER 1\.1', txt, re.I) and 'P=' in txt:
        return 'title'
    if re.search(r'Cabin Tx\.\d+.*Inverter', txt):
        return 'cabin_label'
    if re.match(r'^CABIN \d+$', c):
        return 'cabin_header'
    if re.search(r'\bPV modules?\b', txt, re.I):
        return 'panel_count'
    if STRING_RE.search(txt) or c == 'reserve':
        return 'string_label'
    return 'fixed'


def _mtext_attribs(d, dx, dy):
    """Build a dxfattribs dict for placing an MTEXT entity at an offset."""
    att = {'insert': (d['x'] + dx, d['y'] + dy)}
    for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
              'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
        if a in d:
            att[a] = d[a]
    return att


def _apply_common(ne, d):
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if a in d:
                setattr(ne.dxf, a, d[a])
        except Exception:
            pass


def _place_entity(layout, d, dx, dy, log):
    """Stamp one entity from its dict, shifted by (dx, dy)."""
    t = d['type']
    try:
        if t == 'LWPOLYLINE':
            pts = [(p[0] + dx, p[1] + dy) + tuple(p[2:]) for p in d['pts']]
            ne  = layout.add_lwpolyline(pts)
            ne.closed = d.get('closed', False)
            if 'const_width' in d:
                ne.dxf.const_width = d['const_width']
            _apply_common(ne, d)
        elif t == 'MTEXT':
            ne = layout.add_mtext(d['text'], dxfattribs=_mtext_attribs(d, dx, dy))
            _apply_common(ne, d)
        elif t == 'ARC':
            ne = layout.add_arc(
                (d['cx'] + dx, d['cy'] + dy, d['cz']),
                d['radius'], d['start_angle'], d['end_angle'])
            _apply_common(ne, d)
        elif t == 'CIRCLE':
            ne = layout.add_circle(
                (d['cx'] + dx, d['cy'] + dy, d['cz']), d['radius'])
            _apply_common(ne, d)
        elif t == 'LINE':
            ne = layout.add_line(
                (d['sx'] + dx, d['sy'] + dy, d['sz']),
                (d['ex'] + dx, d['ey'] + dy, d['ez']))
            _apply_common(ne, d)
        elif t == 'ELLIPSE':
            ne = layout.add_ellipse(
                center=(d['cx'] + dx, d['cy'] + dy, d['cz']),
                major_axis=d['major_axis'], ratio=d['ratio'],
                start_param=d['start_param'], end_param=d['end_param'])
            _apply_common(ne, d)
        elif t == 'INSERT':
            ne = layout.add_blockref(
                d['name'], (d['ix'] + dx, d['iy'] + dy, d['iz']))
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d:
                        setattr(ne.dxf, a, d[a])
                except Exception:
                    pass
            _apply_common(ne, d)
        elif t == 'POLYLINE':
            pts = [(p[0] + dx, p[1] + dy, p[2]) for p in d['pts3d']]
            if pts:
                ne = layout.add_polyline3d(pts)
                _apply_common(ne, d)
    except Exception as ex:
        log(f"  [warn] place {t}: {ex}")


def _place_entity_stretched(layout, d, dx, dy, split_y, extra_h, log):
    """Like _place_entity but shifts any coordinate below split_y further down by extra_h.

    This stretches the inverter box and bottom annotations to cover extrapolated
    MPPT rows, while leaving all existing MPPT geometry above split_y intact.
    """
    if extra_h <= 0:
        _place_entity(layout, d, dx, dy, log)
        return

    def _sy(y):
        return y - extra_h if y < split_y else y

    t = d['type']
    try:
        if t == 'LWPOLYLINE':
            pts = [(p[0] + dx, _sy(p[1]) + dy) + tuple(p[2:]) for p in d['pts']]
            ne  = layout.add_lwpolyline(pts)
            ne.closed = d.get('closed', False)
            if 'const_width' in d:
                ne.dxf.const_width = d['const_width']
            _apply_common(ne, d)
        elif t == 'MTEXT':
            ne = layout.add_mtext(d['text'],
                                  dxfattribs=_mtext_attribs(dict(d, y=_sy(d['y'])), dx, dy))
            _apply_common(ne, d)
        elif t == 'ARC':
            ne = layout.add_arc(
                (d['cx'] + dx, _sy(d['cy']) + dy, d['cz']),
                d['radius'], d['start_angle'], d['end_angle'])
            _apply_common(ne, d)
        elif t == 'CIRCLE':
            ne = layout.add_circle(
                (d['cx'] + dx, _sy(d['cy']) + dy, d['cz']), d['radius'])
            _apply_common(ne, d)
        elif t == 'LINE':
            ne = layout.add_line(
                (d['sx'] + dx, _sy(d['sy']) + dy, d['sz']),
                (d['ex'] + dx, _sy(d['ey']) + dy, d['ez']))
            _apply_common(ne, d)
        elif t == 'ELLIPSE':
            ne = layout.add_ellipse(
                center=(d['cx'] + dx, _sy(d['cy']) + dy, d['cz']),
                major_axis=d['major_axis'], ratio=d['ratio'],
                start_param=d['start_param'], end_param=d['end_param'])
            _apply_common(ne, d)
        elif t == 'INSERT':
            ne = layout.add_blockref(
                d['name'], (d['ix'] + dx, _sy(d['iy']) + dy, d['iz']))
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d:
                        setattr(ne.dxf, a, d[a])
                except Exception:
                    pass
            _apply_common(ne, d)
        elif t == 'POLYLINE':
            pts = [(p[0] + dx, _sy(p[1]) + dy, p[2]) for p in d['pts3d']]
            if pts:
                ne = layout.add_polyline3d(pts)
                _apply_common(ne, d)
    except Exception as ex:
        log(f"  [warn] place_stretched {t}: {ex}")


def _place_mtext(layout, d, dx, dy, text, log):
    """Stamp an MTEXT at (dx, dy) offset with a custom text string."""
    try:
        ne = layout.add_mtext(text, dxfattribs=_mtext_attribs(d, dx, dy))
        _apply_common(ne, d)
    except Exception as ex:
        log(f"  [warn] place_mtext: {ex}")


# ─────────────────────────────────────────────────────────────────────────────
#  GENERATION CORE
# ─────────────────────────────────────────────────────────────────────────────

def _generate(cfg, log):
    """Run the full SLD generation with a config dict. log(str) sends text to UI."""
    try:
        import ezdxf as _ez
    except ImportError:
        raise RuntimeError("ezdxf not installed.  Run:  pip install ezdxf")
    try:
        import openpyxl as _xl
    except ImportError:
        raise RuntimeError("openpyxl not installed.  Run:  pip install openpyxl")

    TEMPLATE_DXF = cfg['template_dxf']
    XLSX_PATH    = cfg['xlsx_path']
    OUTPUT_PATH  = cfg['output_path']

    panel_model       = cfg.get('panel_model', '').strip()
    panels_per_string = int(float(cfg.get('panels_per_string') or 0))
    inverter_model    = cfg.get('inverter_model', '').strip()
    dc_power_kwp      = float(cfg.get('dc_power_kwp') or 0)
    ac_power_kwac     = float(cfg.get('ac_power_kwac') or 0)
    temp_rating       = float(cfg.get('temp_rating') or 40)
    transformer_power = cfg.get('transformer_power', '').strip()

    # Workspace parameters
    col_spacing       = float(cfg.get('col_spacing') or _COL_SPACING_DEFAULT)
    row_spacing       = float(cfg.get('row_spacing') or _ROW_SPACING_DEFAULT)
    circle_radius_cfg = float(cfg.get('circle_radius') or 24.59)
    text_size_cfg     = float(cfg.get('text_size') or 60.44)
    heavy_section     = cfg.get('heavy_section', '1x10').strip()
    heavy_linetype    = cfg.get('heavy_linetype', 'TRATTEGGIATA').strip()
    heavy_color       = int(float(cfg.get('heavy_color') or 40))
    heavy_layer       = cfg.get('heavy_layer', 'TRATTEGGIATA').strip()

    # ── 1. Read Excel sheets ──────────────────────────────────────────────────
    log("Reading Excel workbook ...")
    wb = _xl.load_workbook(XLSX_PATH, data_only=True)
    
    # Locate sheet 'Inverter To String' (case-insensitive)
    its_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == 'inverter to string':
            its_sheet_name = name
            break
            
    its_data = {}
    if its_sheet_name:
        log(f"Loading string routing data from sheet '{its_sheet_name}'...")
        ws_its = wb[its_sheet_name]
        # Col 1: To (string ID), Col 2: Cable length +, Col 3: Cable length -, Col 5: Table #
        for r in range(2, ws_its.max_row + 1):
            s_id = ws_its.cell(r, 1).value
            l_p = ws_its.cell(r, 2).value
            l_m = ws_its.cell(r, 3).value
            tbl = ws_its.cell(r, 5).value
            if s_id:
                its_data[str(s_id).strip()] = {
                    'l_plus': float(l_p) if l_p is not None else 0.0,
                    'l_minus': float(l_m) if l_m is not None else 0.0,
                    'table_num': str(tbl).strip() if tbl else 'Default'
                }
        log(f"  Loaded {len(its_data)} routing records.")
    else:
        log("[warn] Sheet 'Inverter To String' not found. Routing lengths will default to 0.0.")

    # Locate sheet '2E802-3' (case-insensitive)
    master_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == '2e802-3':
            master_sheet_name = name
            break
            
    if not master_sheet_name:
        raise RuntimeError(
            f"Master sheet '2E802-3' not found in the workbook.\n"
            f"Available sheets: {', '.join(wb.sheetnames)}"
        )
        
    ws = wb[master_sheet_name]
    excel = {}
    cur = None

    # Dynamic column mapping based on header detection (usually in row 30)
    col_inverter = 1
    col_str_name = 3
    col_mppt = 4
    col_pos = 12
    col_wp = 20
    col_sec = 25

    header_row = 30
    found_headers = False
    for r_idx in (30, 29, 28, 31, 32, 27, 26):
        row_vals = [str(ws.cell(row=r_idx, column=c).value or '').strip().lower() for c in range(1, 35)]
        if any('string name' in val for val in row_vals):
            header_row = r_idx
            found_headers = True
            break

    if found_headers:
        log(f"Found headers row at: {header_row}")
        for c in range(1, 35):
            val = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
            if val == 'inverter':
                col_inverter = c
            elif val == 'string name':
                col_str_name = c
            elif val == 'mppt':
                col_mppt = c
            elif 'posizione stringa' in val:
                col_pos = c
            elif val == 'module type':
                col_wp = c
            elif 'section' in val:
                col_sec = c
    else:
        log("[warn] Headers row containing 'string name' not found. Using default column mappings.")

    log(f"Dynamic column indices detected -> Inverter: {col_inverter}, StringName: {col_str_name}, MPPT: {col_mppt}, Position: {col_pos}, Wp: {col_wp}, Section: {col_sec}")

    log(f"Parsing master cable schedule from sheet '{master_sheet_name}'...")
    for r in range(header_row + 1, ws.max_row + 1):
        inv = ws.cell(row=r, column=col_inverter).value
        sname = ws.cell(row=r, column=col_str_name).value
        mppt = ws.cell(row=r, column=col_mppt).value

        if inv is not None and '.' in str(inv):
            try:
                parts = str(inv).split('.')
                t = int(parts[0])
                v = int(parts[1])
                cur = (t, v)
                excel.setdefault(cur, defaultdict(list))
            except Exception:
                pass

        if sname and cur and mppt:
            sname_str = str(sname).strip()
            route = its_data.get(sname_str, {'l_plus': 0.0, 'l_minus': 0.0, 'table_num': 'Default'})

            # Read Section (mm²)
            section = ws.cell(row=r, column=col_sec).value
            section_str = str(section).strip() if section else '1x6'

            # Read Tracker position
            tracker_pos = ws.cell(row=r, column=col_pos).value
            if not tracker_pos:
                tracker_pos = route['table_num']
            if not tracker_pos or str(tracker_pos).strip() == 'No piling information':
                tracker_pos = 'Default'

            # Wp module rating
            wp = 0
            module_wp_raw = ws.cell(row=r, column=col_wp).value
            try:
                wp = int(float(module_wp_raw))
            except Exception:
                pass

            excel[cur][int(float(mppt))].append({
                'name': sname_str,
                'wp': wp,
                'l_plus': route['l_plus'],
                'l_minus': route['l_minus'],
                'section': section_str,
                'tracker_pos': str(tracker_pos).strip()
            })

    inv_list = sorted(excel.keys())
    transformers = {}
    for (T, I) in inv_list:
        transformers.setdefault(T, []).append(I)
    for T in transformers:
        transformers[T] = sorted(transformers[T])
    transformer_list = sorted(transformers.keys())

    log(f"Inverters parsed : {len(inv_list)}")
    log(f"Transformers     : {len(transformer_list)}")
    for T in transformer_list:
        ii = transformers[T]
        mc = [len(excel.get((T, i), {})) for i in ii]
        sc = [sum(len(v) for v in excel.get((T, i), {}).values()) for i in ii]
        log(f"  Tx{T}: {len(ii)} inverters  "
            f"MPPTs {min(mc)}-{max(mc)}  Strings {min(sc)}-{max(sc)}")

    # ── 2. Load template DXF and extract template-band entities ──────────────
    log("Loading template DXF ...")
    doc = _ez.readfile(TEMPLATE_DXF)
    msp = doc.modelspace()

    # Ensure custom linetype is defined in doc
    if heavy_linetype not in doc.linetypes:
        try:
            doc.linetypes.new(heavy_linetype, dxfattribs={
                'description': f'Custom dash linetype {heavy_linetype}',
                'pattern': [20.0, -10.0]
            })
            log(f"  Defined missing linetype '{heavy_linetype}' in output DXF")
        except Exception as ex:
            log(f"  [warn] Could not define linetype '{heavy_linetype}': {ex}")

    # Ensure custom layer is defined
    if heavy_layer not in doc.layers:
        try:
            doc.layers.new(heavy_layer, dxfattribs={'color': heavy_color, 'linetype': heavy_linetype})
            log(f"  Created missing layer '{heavy_layer}' in output DXF")
        except Exception as ex:
            log(f"  [warn] Could not define layer '{heavy_layer}': {ex}")

    # Auto-detect the template Y-band by anchoring on the "INVERTER 1.1" title.
    anchor_y = None
    for _e in msp:
        if _e.dxftype() == 'MTEXT':
            if re.search(r'INVERTER\s+1\.1\b', _e.text, re.I) and 'P=' in _e.text:
                anchor_y = _e.dxf.insert.y
                break
    if anchor_y is not None:
        tmpl_y_min = anchor_y - 10000
        tmpl_y_max = anchor_y + 5000
        log(f"Template anchor Y={anchor_y:.0f}  band {tmpl_y_min:.0f}-{tmpl_y_max:.0f}")
    else:
        tmpl_y_min, tmpl_y_max = _TMPL_Y_MIN, _TMPL_Y_MAX
        log(f"[warn] 'INVERTER 1.1' not found; using default Y-band {_TMPL_Y_MIN}-{_TMPL_Y_MAX}")

    raw_ents  = [e for e in msp
                 if (y := _ent_y(e)) is not None
                 and tmpl_y_min <= y <= tmpl_y_max]
    log(f"Template band: {len(raw_ents)} entities found")
    all_dicts = [d for e in raw_ents if (d := _extract_entity(e, log)) is not None]

    # Left edge of template in model space
    xs = []
    for d in all_dicts:
        t = d['type']
        try:
            if t == 'LWPOLYLINE': xs += [p[0] for p in d['pts']]
            elif t == 'MTEXT':    xs.append(d['x'])
            elif t in ('ARC', 'CIRCLE', 'ELLIPSE'): xs.append(d['cx'])
            elif t == 'LINE':     xs += [d['sx'], d['ex']]
            elif t == 'INSERT':   xs.append(d['ix'])
            elif t == 'POLYLINE': xs += [p[0] for p in d['pts3d']]
        except Exception:
            pass
    xmin = min(xs) if xs else 0
    xcut = xmin + _COL_STEP

    tmpl = [d for d in all_dicts
            if _entity_min_x(d) <= xcut and not _is_placeholder_rect(d)]
    log(f"Template: {len(tmpl)}/{len(all_dicts)} entities in column slice")

    # ── 3. Classify MTEXT entities ────────────────────────────────────────────
    tmpl_texts = []
    for d in tmpl:
        if d['type'] == 'MTEXT':
            d['cls'] = _classify_mtext(d['text'])
            tmpl_texts.append(d)

    # ── 4. MPPT-port → string label position map ──────────────────────────────
    port_lbl = [(m['x'], m['y'], _strip_mtext_fmt(m['text']))
                for m in tmpl_texts if PORT_RE.match(_strip_mtext_fmt(m['text']))]
    str_lbl  = [m for m in tmpl_texts if m['cls'] == 'string_label']

    mppt_map = {}
    for px, py, ptxt in port_lbl:
        best, bd = None, _PORT_Y_TOL
        for sl in str_lbl:
            if sl['x'] > px:
                dd = abs(sl['y'] - py)
                if dd < bd:
                    bd, best = dd, sl
        if best:
            m2 = re.match(r'^(\d+)-(\d+)$', ptxt)
            if m2:
                mppt_map[(int(m2.group(1)), int(m2.group(2)))] = best

    # Any string_label not within _PORT_Y_TOL of any port → fallback for slot 1-1
    panel_sl = next(
        (sl for sl in str_lbl
         if not any(abs(sl['y'] - py) < _PORT_Y_TOL for _, py, _ in port_lbl)),
        None)
    if panel_sl and (1, 1) not in mppt_map:
        mppt_map[(1, 1)] = panel_sl

    log(f"MPPT/port slots in template: {len(mppt_map)}")

    # Extrapolate missing MPPT rows when template has fewer rows than Excel.
    tmpl_mpputs = sorted(set(m for m, _ in mppt_map))
    excel_mpputs = sorted(set(m for invd in excel.values() for m in invd.keys()))
    missing = [m for m in excel_mpputs if m not in tmpl_mpputs]

    _stretch_split_y = float('inf')
    _stretch_extra_h = 0.0

    if missing:
        log(f"[warn] Template missing MPPT slot(s): {missing} - extrapolating positions.")
        port1_ys = [(m, mppt_map[(m, 1)]['y']) for m in tmpl_mpputs if (m, 1) in mppt_map]
        port1_ys.sort()

        if port1_ys:
            if len(port1_ys) >= 2:
                steps = [port1_ys[i][1] - port1_ys[i + 1][1] for i in range(len(port1_ys) - 1)]
                if len(steps) > 1:
                    rest = sorted(steps[1:])
                    median_step = rest[len(rest) // 2]
                    use_steps = steps[1:] if steps[0] > 2 * median_step else steps
                else:
                    use_steps = steps
                avg_step = sum(use_steps) / len(use_steps) if use_steps else 210.0
            else:
                avg_step = 210.0

            last_m, last_y1 = port1_ys[-1]
            if (last_m, 2) in mppt_map:
                port_inner_offset = mppt_map[(last_m, 2)]['y'] - last_y1
            else:
                port_inner_offset = -105

            _stretch_split_y = last_y1 + port_inner_offset - avg_step * 0.5
            _stretch_extra_h = len(missing) * avg_step
            log(f"  Stretching inverter box: split_y={_stretch_split_y:.0f}, extra_h={_stretch_extra_h:.0f}")

            ref_p1 = mppt_map.get((last_m, 1), str_lbl[-1] if str_lbl else None)
            ref_p2 = mppt_map.get((last_m, 2), ref_p1)

            if ref_p1:
                for miss_m in missing:
                    delta = miss_m - last_m
                    new_y1 = last_y1 - delta * avg_step
                    new_y2 = new_y1 + port_inner_offset
                    mppt_map[(miss_m, 1)] = dict(ref_p1, y=new_y1)
                    if ref_p2:
                        mppt_map[(miss_m, 2)] = dict(ref_p2, y=new_y2)
                log(f"Extrapolated MPPT rows {missing}")

    # ── 5. Precompute inverter positions ──────────────────────────────────────
    inv_col = {(T, I): idx for T, invs in transformers.items() for idx, I in enumerate(invs)}
    tx_row  = {T: idx for idx, T in enumerate(transformer_list)}
    _eff_row_spacing = int(row_spacing + _stretch_extra_h)

    def inv_offset(T, I):
        return (inv_col[(T, I)] * col_spacing, -tx_row[T] * _eff_row_spacing)

    # ── 6. Label builders ─────────────────────────────────────────────────────
    def make_string_label(T, I, mppt, port):
        lst = excel.get((T, I), {}).get(mppt, [])
        if port - 1 < len(lst):
            sdata = lst[port - 1]
            name = sdata['name']
            wp = sdata['wp']
            l_plus = sdata['l_plus']
            l_minus = sdata['l_minus']
            sec = sdata['section']
            
            label = f"String {name}"
            if panels_per_string > 0:
                suffix = f" {wp}Wp" if wp > 0 else ""
                if panel_model:
                    label += f" - {panels_per_string}x {panel_model}{suffix}"
                else:
                    label += f" - {panels_per_string}P{suffix}"
            label += f" (L+={l_plus:.3f}m, L-={l_minus:.3f}m, {sec})"
            return label
        return "reserve"

    def make_inv_title(T, I):
        dc = dc_power_kwp
        if dc <= 0 and panels_per_string > 0:
            inv_data = excel.get((T, I), {})
            total_kwp = sum(
                panels_per_string * sd['wp'] / 1000.0
                for strings in inv_data.values()
                for sd in strings
                if sd['wp'] > 0
            )
            if total_kwp > 0:
                dc = total_kwp
        parts = [f"INVERTER {T}.{I}"]
        if inverter_model:
            parts.append(inverter_model)
        if dc > 0:
            parts.append(f"P= {dc:.0f} KWp")
        if ac_power_kwac > 0:
            parts.append(f"P= {ac_power_kwac:.0f} KWac @{temp_rating:.0f}°C")
        return " - ".join(parts)

    def make_cabin_hdr(T):
        return f"CABIN {T}\n{transformer_power}" if transformer_power else f"CABIN {T}"

    # ── 7. Clear modelspace and existing paper-space layouts ──────────────────
    log("Preparing output DXF ...")
    msp.delete_all_entities()
    paper_layouts = [l.name for l in doc.layouts if not l.is_modelspace]
    for name in paper_layouts:
        try:
            doc.layouts.delete(name)
        except Exception:
            pass

    # ── 8. Stamp all inverter sections ────────────────────────────────────────
    td  = next((m for m in tmpl_texts if m['cls'] == 'title'),        None)
    chd = next((m for m in tmpl_texts if m['cls'] == 'cabin_header'), None)
    cld = next((m for m in tmpl_texts if m['cls'] == 'cabin_label'),  None)

    log(f"Generating {len(inv_list)} inverter sections ...")
    for idx, (T, I) in enumerate(inv_list):
        dx, dy = inv_offset(T, I)
        
        # Stamp background geometry & static annotations
        for tmpl_idx, d in enumerate(tmpl):
            if d['type'] == 'MTEXT':
                cls = d.get('cls', 'fixed')
                if cls == 'fixed':
                    _place_entity_stretched(msp, d, dx, dy,
                                            _stretch_split_y, _stretch_extra_h, log)
                elif cls == 'panel_count':
                    if panels_per_string > 0:
                        updated = re.sub(
                            r'\d+(\s*PV modules?)',
                            lambda m: f'{panels_per_string}{m.group(1)}',
                            d['text'], flags=re.I)
                        d_s = dict(d, y=(d['y'] - _stretch_extra_h if d['y'] < _stretch_split_y else d['y']))
                        _place_mtext(msp, d_s, dx, dy, updated, log)
                    else:
                        _place_entity_stretched(msp, d, dx, dy,
                                                _stretch_split_y, _stretch_extra_h, log)
                elif 'mmq' in d['text']:
                    # Update cable section specs at the bottom dynamically based on this inverter's strings
                    sections_used = set()
                    for mppt_s in excel.get((T, I), {}).values():
                        for sdata in mppt_s:
                            sections_used.add(sdata['section'])
                    
                    if len(sections_used) == 1:
                        sec = list(sections_used)[0]
                        new_txt = f"2/ {sec} mmq - Cu - H1Z2Z2k"
                    else:
                        new_txt = "2/ (1x6/10)mmq - Cu - H1Z2Z2k"
                        
                    d_s = dict(d, y=(d['y'] - _stretch_extra_h if d['y'] < _stretch_split_y else d['y']))
                    _place_mtext(msp, d_s, dx, dy, new_txt, log)
            else:
                # Check if this geometry is a connection line for a slot (mppt, port)
                slot = None
                if d['type'] in ('LINE', 'LWPOLYLINE', 'POLYLINE'):
                    for (mppt, port), sl in mppt_map.items():
                        if _is_cable_line(d, sl['y']):
                            slot = (mppt, port)
                            break
                            
                slot_circle = None
                if d['type'] == 'CIRCLE':
                    for (mppt, port), sl in mppt_map.items():
                        if abs(d['cy'] - sl['y']) < 80:
                            slot_circle = (mppt, port)
                            break

                if slot:
                    mppt, port = slot
                    lst = excel.get((T, I), {}).get(mppt, [])
                    is_active = (port - 1 < len(lst))
                    
                    d_s = dict(d)
                    if is_active:
                        sdata = lst[port - 1]
                        if sdata['section'] == heavy_section:
                            d_s['linetype'] = heavy_linetype
                            d_s['color']    = heavy_color
                            d_s['layer']    = heavy_layer
                            d_s['ltscale']  = 0.5
                        else:
                            # Reset standard cables to layer 0 and solid
                            d_s['linetype'] = 'Continuous'
                            d_s['color']    = 40
                            d_s['layer']    = '0'
                    else:
                        # Unused / reserve connection line is drawn in dim color
                        d_s['linetype'] = 'Continuous'
                        d_s['color']    = 8  # grey
                        d_s['layer']    = '0'
                        
                    _place_entity_stretched(msp, d_s, dx, dy,
                                            _stretch_split_y, _stretch_extra_h, log)
                elif slot_circle:
                    mppt, port = slot_circle
                    lst = excel.get((T, I), {}).get(mppt, [])
                    is_active = (port - 1 < len(lst))
                    
                    d_s = dict(d)
                    if is_active:
                        d_s['radius'] = circle_radius_cfg
                        orig_color = d.get('color', 7)
                        d_s['color'] = 7 if orig_color in (8, 253) else orig_color
                    else:
                        d_s['color']  = 8  # gray out reserve circle terminal
                        
                    _place_entity_stretched(msp, d_s, dx, dy,
                                            _stretch_split_y, _stretch_extra_h, log)
                else:
                    # Stretches/places default template geometry
                    _place_entity_stretched(msp, d, dx, dy,
                                            _stretch_split_y, _stretch_extra_h, log)
        
        # Stamp inverter/cabin headers
        if td:
            _place_mtext(msp, td,  dx, dy, make_inv_title(T, I), log)
        if chd:
            _place_mtext(msp, chd, dx, dy, make_cabin_hdr(T), log)
        if cld:
            _place_mtext(msp, cld, dx, dy, f"\\pxqc;Cabin Tx.{T}\\PInverter {T}.{I}", log)
            
        # Stamp parameterised string labels
        for (mppt, port), sl in mppt_map.items():
            sl_wide = dict(sl, width=max(sl.get('width', 0), _STRING_LABEL_MIN_WIDTH))
            sl_wide['char_height'] = text_size_cfg
            
            label = make_string_label(T, I, mppt, port)
            if label == "reserve":
                sl_wide['color'] = 8  # grey
            else:
                orig_color = sl.get('color', 7)
                sl_wide['color'] = 7 if orig_color in (8, 253) else orig_color
                
            _place_mtext(msp, sl_wide, dx, dy, label, log)
            
        if (idx + 1) % 10 == 0 or (idx + 1) == len(inv_list):
            log(f"  {idx + 1}/{len(inv_list)} inverters done")

    # ── 9. Paper-space layouts (A3 landscape) ─────────────────────────────────
    eff_tmpl_y_min  = tmpl_y_min - _stretch_extra_h
    tmpl_height     = tmpl_y_max - eff_tmpl_y_min
    tmpl_y_center   = (eff_tmpl_y_min + tmpl_y_max) / 2
    tmpl_x_center   = xmin + _COL_STEP / 2

    log(f"Creating {len(transformer_list)} paper-space layout(s) ...")
    for T in transformer_list:
        lname = f"Tx{T}"
        try:
            layout = doc.layouts.new(lname)
        except Exception:
            layout = doc.layouts.get(lname)

        n_inv  = len(transformers[T])
        row_cx = tmpl_x_center + (n_inv - 1) * col_spacing / 2
        row_cy = tmpl_y_center - tx_row[T] * _eff_row_spacing
        row_w  = col_spacing * n_inv

        view_h = max(tmpl_height * 1.05, row_w / (420.0 / 297.0) * 1.05)

        layout.add_viewport(
            center=(210, 148.5),
            size=(420, 297),
            view_center_point=(row_cx, row_cy),
            view_height=view_h,
        )
        log(f"  Created layout '{lname}' with {n_inv} inverter(s)")

    # ── 10. Save ──────────────────────────────────────────────────────────────
    log(f"Saving -> {OUTPUT_PATH}")
    doc.saveas(OUTPUT_PATH)
    log("Done! SLD generated successfully.")


# ─────────────────────────────────────────────────────────────────────────────
#  HISTORY STORE
# ─────────────────────────────────────────────────────────────────────────────

_HISTORY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'history_v2.json')

_DEFAULTS = {
    'panel_model': [
        'JA Solar JAM72D42-625/LB',
        'JA Solar JAM72S20-460/MR',
        'Longi Solar LR5-72HBD-580M',
        'Longi Solar LR5-72HBD-545M',
        'Canadian Solar HiKu7 CS7N-655MB',
        'Jinko Solar JKM660M-78HL4-V',
        'Trina Solar TSM-670NEG21C.20',
        'REC Alpha Pure-R 430AA',
    ],
    'panels_per_string': ['16', '18', '20', '22', '24', '26', '28', '30'],
    'inverter_model': [
        'Sungrow SG350HX',
        'Sungrow SG250HX',
        'Sungrow SG125HX',
        'Huawei SUN2000-330KTL',
        'Huawei SUN2000-275KTL-H1',
        'Huawei SUN2000-215KTL-H3',
        'ABB PVS-250-TL',
        'SMA Sunny Tripower Core2 150',
        'Fronius Symo GEN24 25.0 Plus',
    ],
    'dc_power_kwp':  ['250', '275', '300', '320', '330', '350', '375', '400', '450', '500'],
    'ac_power_kwac': ['200', '225', '250', '275', '300', '320', '330', '350'],
}


class _HistoryStore:
    def __init__(self, path):
        self._path = path
        self._data: dict = {}
        self._load()

    def _load(self):
        try:
            if os.path.isfile(self._path):
                with open(self._path, encoding='utf-8') as f:
                    self._data = json.load(f)
        except Exception:
            self._data = {}

    def _save(self):
        try:
            with open(self._path, 'w', encoding='utf-8') as f:
                json.dump(self._data, f, indent=2, ensure_ascii=False)
        except Exception as ex:
            print(f"[warn] Could not save preset history: {ex}", file=sys.stderr)

    def values(self, key):
        entries = sorted(self._data.get(key, []),
                         key=lambda e: e.get('ts', ''), reverse=True)
        history_vals = [e['value'] for e in entries]
        result = list(history_vals)
        for v in _DEFAULTS.get(key, []):
            if v not in result:
                result.append(v)
        return result

    def record(self, key, value):
        value = value.strip()
        if not value:
            return
        ts = datetime.now().isoformat(timespec='seconds')
        entries = self._data.setdefault(key, [])
        for e in entries:
            if e['value'] == value:
                e['ts'] = ts
                break
        else:
            entries.append({'value': value, 'ts': ts})
        self._save()

    def all_entries(self, key):
        return list(self._data.get(key, []))

    def add_manual(self, key, value, note=''):
        value = value.strip()
        if not value:
            return
        entries = self._data.setdefault(key, [])
        for e in entries:
            if e['value'] == value:
                if note.strip():
                    e['note'] = note.strip()
                break
        else:
            entry = {'value': value}
            if note.strip():
                entry['note'] = note.strip()
            entries.append(entry)
        self._save()

    def delete_entry(self, key, value):
        entries = self._data.get(key, [])
        self._data[key] = [e for e in entries if e['value'] != value]
        self._save()


_history = _HistoryStore(_HISTORY_PATH)

_FIELD_LABELS: dict[str, str] = {
    'panel_model':       'Panel Model',
    'panels_per_string': 'Panels per String',
    'inverter_model':    'Inverter Model',
    'dc_power_kwp':      'DC Power (KWp)',
    'ac_power_kwac':     'AC Power (KWac)',
}

# ─────────────────────────────────────────────────────────────────────────────
#  GUI WIDGETS
# ─────────────────────────────────────────────────────────────────────────────

_PAD = {'padx': 6, 'pady': 3}


class _HistoryCombo(ttk.Frame):
    def __init__(self, parent, label, history_key, default='', unit='', width=28,
                 refresh_callback=None, on_select=None, **kw):
        super().__init__(parent, **kw)
        self._key        = history_key
        self._refresh_cb = refresh_callback
        self._select_cb  = on_select
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=default)
        self._combo = ttk.Combobox(self, textvariable=self.var, width=width,
                                   values=_history.values(history_key))
        self._combo.pack(side='left', **_PAD)
        if unit:
            ttk.Label(self, text=unit, foreground='gray').pack(side='left')
        ttk.Button(self, text='...', width=3,
                   command=self._open_manager).pack(side='left', padx=(6, 0))
        self._combo.bind('<<ComboboxSelected>>', self._on_combo_select)

    def _on_combo_select(self, event=None):
        if self._select_cb:
            self._select_cb(self.get())

    def _open_manager(self):
        root = self.winfo_toplevel()
        _PresetManagerDialog(root, initial_key=self._key,
                             on_close=self._refresh_cb)

    def get(self):
        return self.var.get().strip()

    def refresh(self):
        self._combo['values'] = _history.values(self._key)

    def record(self):
        v = self.get()
        if v:
            _history.record(self._key, v)
            self._combo['values'] = _history.values(self._key)


class _FileRow(ttk.Frame):
    def __init__(self, parent, label, default='', save=False,
                 filetypes=None, **kw):
        super().__init__(parent, **kw)
        self._save      = save
        self._filetypes = filetypes or [('All files', '*.*')]
        ttk.Label(self, text=label, width=18, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=default)
        ttk.Entry(self, textvariable=self.var, width=52).pack(
            side='left', padx=(0, 4), fill='x', expand=True)
        ttk.Button(self, text='Browse…', width=9,
                   command=self._browse).pack(side='left')

    def _browse(self):
        cur = self.var.get()
        init_dir = os.path.dirname(cur) if cur else os.path.expanduser('~')
        if self._save:
            p = filedialog.asksaveasfilename(
                initialdir=init_dir,
                defaultextension='.dxf',
                filetypes=self._filetypes)
        else:
            p = filedialog.askopenfilename(
                initialdir=init_dir,
                filetypes=self._filetypes)
        if p:
            self.var.set(p)

    def get(self):
        return self.var.get().strip()


class _FieldRow(ttk.Frame):
    def __init__(self, parent, label, default='', unit='', width=20, **kw):
        super().__init__(parent, **kw)
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=str(default))
        ttk.Entry(self, textvariable=self.var, width=width).pack(side='left', **_PAD)
        if unit:
            ttk.Label(self, text=unit, foreground='gray').pack(side='left')

    def get(self):
        return self.var.get().strip()


class _PresetManagerDialog(tk.Toplevel):
    def __init__(self, parent, initial_key=None, on_close=None):
        super().__init__(parent)
        self.title("Manage Presets")
        self.geometry("720x520")
        self.minsize(580, 380)
        self._on_close = on_close
        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._close)
        self._build(initial_key)

    def _build(self, initial_key):
        top = ttk.Frame(self, padding=(12, 10, 12, 4))
        top.pack(fill='x')
        ttk.Label(top, text="Field:").pack(side='left', padx=(0, 6))
        self._field_var = tk.StringVar()
        self._keys = list(_FIELD_LABELS.keys())
        labels = [_FIELD_LABELS[k] for k in self._keys]
        self._field_cb = ttk.Combobox(top, textvariable=self._field_var,
                                      values=labels, state='readonly', width=32)
        self._field_cb.pack(side='left')
        self._field_cb.bind('<<ComboboxSelected>>', lambda _: self._refresh_tree())

        mid = ttk.Frame(self, padding=(12, 4))
        mid.pack(fill='both', expand=True)
        cols = ('value', 'note', 'last_used', 'kind')
        self._tree = ttk.Treeview(mid, columns=cols, show='headings', selectmode='browse')
        self._tree.heading('value',     text='Value')
        self._tree.heading('note',      text='Note / Details')
        self._tree.heading('last_used', text='Last Used')
        self._tree.heading('kind',      text='Type')
        self._tree.column('value',     width=210, minwidth=100)
        self._tree.column('note',      width=210, minwidth=100)
        self._tree.column('last_used', width=140, minwidth=90)
        self._tree.column('kind',      width=70,  minwidth=55)
        vsb = ttk.Scrollbar(mid, orient='vertical', command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='left', fill='y')
        self._tree.bind('<<TreeviewSelect>>', self._on_select)

        form = ttk.LabelFrame(self, text='Add / Update Entry', padding=(10, 6))
        form.pack(fill='x', padx=12, pady=(4, 0))
        r1 = ttk.Frame(form)
        r1.pack(fill='x', pady=2)
        ttk.Label(r1, text='Value:', width=8, anchor='e').pack(side='left')
        self._val_var = tk.StringVar()
        ttk.Entry(r1, textvariable=self._val_var, width=40).pack(side='left', padx=4)
        r2 = ttk.Frame(form)
        r2.pack(fill='x', pady=2)
        ttk.Label(r2, text='Note:', width=8, anchor='e').pack(side='left')
        self._note_var = tk.StringVar()
        ttk.Entry(r2, textvariable=self._note_var, width=56).pack(side='left', padx=4)
        btns = ttk.Frame(form)
        btns.pack(fill='x', pady=(6, 2))
        ttk.Button(btns, text='Add / Update',
                   command=self._add_entry).pack(side='left', padx=(0, 8))
        self._del_btn = ttk.Button(btns, text='Delete Selected',
                                   command=self._delete_entry, state='disabled')
        self._del_btn.pack(side='left')

        foot = ttk.Frame(self, padding=(12, 6))
        foot.pack(fill='x')
        ttk.Button(foot, text='Close', command=self._close).pack(side='right')

        if initial_key and initial_key in self._keys:
            self._field_cb.current(self._keys.index(initial_key))
        else:
            self._field_cb.current(0)
        self._refresh_tree()

    def _current_key(self):
        label = self._field_var.get()
        for k, v in _FIELD_LABELS.items():
            if v == label:
                return k
        return None

    def _refresh_tree(self):
        self._tree.delete(*self._tree.get_children())
        key = self._current_key()
        if not key:
            return
        user_entries = _history.all_entries(key)
        user_values  = {e['value'] for e in user_entries}
        for e in sorted(user_entries, key=lambda x: x.get('ts', ''), reverse=True):
            self._tree.insert('', 'end', iid=f'u|{e["value"]}',
                              values=(e['value'], e.get('note', ''),
                                      e.get('ts', ''), 'User'),
                              tags=('user',))
        for v in _DEFAULTS.get(key, []):
            if v not in user_values:
                self._tree.insert('', 'end', iid=f'd|{v}',
                                  values=(v, '', '', 'Default'),
                                  tags=('default',))
        self._tree.tag_configure('default', foreground='gray')
        self._del_btn.configure(state='disabled')

    def _on_select(self, _):
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        row = self._tree.item(iid, 'values')
        self._val_var.set(row[0])
        self._note_var.set(row[1])
        self._del_btn.configure(state='normal' if iid.startswith('u|') else 'disabled')

    def _add_entry(self):
        key   = self._current_key()
        value = self._val_var.get().strip()
        if not key or not value:
            messagebox.showwarning("Input Required", "Please enter a value.", parent=self)
            return
        _history.add_manual(key, value, self._note_var.get())
        self._val_var.set('')
        self._note_var.set('')
        self._refresh_tree()

    def _delete_entry(self):
        sel = self._tree.selection()
        if not sel or not sel[0].startswith('u|'):
            return
        value = self._tree.item(sel[0], 'values')[0]
        if messagebox.askyesno("Confirm Delete",
                                f"Delete preset:\n\n  {value}\n\nThis cannot be undone.",
                                parent=self):
            _history.delete_entry(self._current_key(), value)
            self._refresh_tree()

    def _close(self):
        self.grab_release()
        self.destroy()
        if self._on_close:
            self._on_close()


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN APP WINDOW
# ─────────────────────────────────────────────────────────────────────────────

class SLDAppV2(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("A176LAB - DC Single Line Diagram Generator v2")
        self.geometry("880x820")
        self.minsize(720, 680)
        self.resizable(True, True)

        style = ttk.Style(self)
        for theme in ('vista', 'winnative', 'clam', 'default'):
            try:
                style.theme_use(theme)
                break
            except Exception:
                pass

        self._build_ui()
        self._set_icon()
        self.bind('<F5>', lambda _: self._on_generate())

    def _build_ui(self):
        self._add_header()
        ttk.Separator(self, orient='horizontal').pack(fill='x', padx=8, pady=(0, 4))

        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        self._build_files_tab(nb)
        self._build_equip_tab(nb)
        self._build_workspace_tab(nb)
        self._build_run_tab(nb)

    def _add_header(self):
        header = ttk.Frame(self, padding=(10, 8, 10, 6))
        header.pack(fill='x')
        logo = _load_logo_tk((72, 72))
        if logo:
            self._logo_ref = logo
            ttk.Label(header, image=logo).pack(side='left', padx=(0, 14))
        txt = ttk.Frame(header)
        txt.pack(side='left', fill='y', pady=2)
        ttk.Label(txt, text="DC Single Line Diagram Generator v2",
                  font=('Segoe UI', 14, 'bold')).pack(anchor='w')
        ttk.Label(txt, text="A176 LAB  –  Think different project",
                  foreground='gray', font=('Segoe UI', 9)).pack(anchor='w')

    def _set_icon(self):
        """Set window and taskbar icon from A176LAB logo."""
        try:
            from PIL import Image, ImageTk
            img = Image.open(_LOGO_PATH).convert('RGBA')
            img.thumbnail((256, 256), Image.LANCZOS)
            self._icon_ref = ImageTk.PhotoImage(img)
            self.iconphoto(True, self._icon_ref)
        except Exception as ex:
            print(f"[warn] could not set window icon: {ex}", file=sys.stderr)

    # ── Files Tab ─────────────────────────────────────────────────────────────
    def _build_files_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Files  ')

        ttk.Label(tab, text="File Paths Configuration",
                  font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))

        # Default paths setup
        default_tmpl = r'C:\Users\user\Desktop\SLD Diagram\YANEL\26S001_2E103 - DC Single Line Diagram.dxf'
        if not os.path.exists(default_tmpl):
            default_tmpl = ''

        self.fe_tmpl = _FileRow(
            tab, "Template DXF:",
            default=default_tmpl,
            filetypes=[('DXF files', '*.dxf'), ('All files', '*.*')])
        self.fe_tmpl.pack(fill='x', pady=2)

        self.fe_xlsx = _FileRow(
            tab, "Excel Cable List:",
            filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')])
        self.fe_xlsx.pack(fill='x', pady=2)
        self.fe_xlsx.var.trace_add('write', self._sync_output_path)

        self.fe_out = _FileRow(
            tab, "Output DXF:", save=True,
            filetypes=[('DXF files', '*.dxf'), ('All files', '*.*')])
        self.fe_out.pack(fill='x', pady=2)

        ttk.Separator(tab).pack(fill='x', pady=12)

        hint = (
            "Template DXF      - Source DXF template containing Inverter 1.1 layout slice.\n"
            "Excel Cable List  - Project schedule with sheets '2E802-3' and 'Inverter To String'.\n"
            "Output DXF        - Saved diagram path (auto-fills to target folder as Excel file).\n\n"
            "F5 starts diagram generation automatically."
        )
        ttk.Label(tab, text=hint, foreground='gray',
                  wraplength=720, justify='left').pack(anchor='w')

    # ── Equipment Tab ─────────────────────────────────────────────────────────
    def _build_equip_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Equipment  ')

        def _on_inverter_model_select(model):
            if model in _INVERTER_POWERS:
                dc, ac = _INVERTER_POWERS[model]
                self.f_dc_power.var.set(dc)
                self.f_ac_power.var.set(ac)

        # Solar Panel section
        self._section(tab, "Solar Panel")
        self.f_panel_model = _HistoryCombo(
            tab, "Panel Model:", 'panel_model',
            refresh_callback=self._refresh_all_combos)
        self.f_panel_model.pack(fill='x')

        ttk.Label(tab, text="  Module ratings are parsed dynamically from Excel sheet '2E802-3'.",
                  foreground='gray', font=('Segoe UI', 8)).pack(anchor='w', padx=6)

        self.f_panels_str = _HistoryCombo(
            tab, "Panels per String:", 'panels_per_string',
            default='20', unit='panels', width=14,
            refresh_callback=self._refresh_all_combos)
        self.f_panels_str.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=10)

        # Inverter section
        self._section(tab, "Inverter Specs")
        self.f_inv_model = _HistoryCombo(
            tab, "Inverter Model:", 'inverter_model',
            refresh_callback=self._refresh_all_combos,
            on_select=_on_inverter_model_select)
        self.f_inv_model.pack(fill='x')

        self.f_dc_power = _HistoryCombo(
            tab, "DC Power per Inverter:", 'dc_power_kwp',
            default='350', unit='KWp  (0 = calculate from module ratings)', width=14,
            refresh_callback=self._refresh_all_combos)
        self.f_dc_power.pack(fill='x')

        self.f_ac_power = _HistoryCombo(
            tab, "AC Power:", 'ac_power_kwac',
            default='320', unit='KWac', width=14,
            refresh_callback=self._refresh_all_combos)
        self.f_ac_power.pack(fill='x')

        self.f_temp = _FieldRow(tab, "Temperature Rating:", '40', unit='°C')
        self.f_temp.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=10)

        # Transformer section
        self._section(tab, "Transformer specs")
        self.f_tx_power = _FieldRow(tab, "Transformer Power:", '',
                                     unit='e.g. 2.5 MVA  (appended to cabin header)', width=20)
        self.f_tx_power.pack(fill='x')

    # ── Workspace Parameters Tab ──────────────────────────────────────────────
    def _build_workspace_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Workspace Parameters  ')

        self._section(tab, "Array Grid Layout Steps")
        self.f_col_spacing = _FieldRow(tab, "Horizontal Grid Col Step:", str(_COL_SPACING_DEFAULT),
                                       unit="units (AutoCAD distance coordinates)")
        self.f_col_spacing.pack(fill='x')

        self.f_row_spacing = _FieldRow(tab, "Vertical Grid Row Step:", str(_ROW_SPACING_DEFAULT),
                                       unit="units (AutoCAD distance coordinates)")
        self.f_row_spacing.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=8)

        self._section(tab, "Visual Node Elements")
        self.f_circle_radius = _FieldRow(tab, "Module Circle Radius:", "24.59",
                                         unit="units (AutoCAD drawing units)")
        self.f_circle_radius.pack(fill='x')

        self.f_text_size = _FieldRow(tab, "String Label Text Height:", "60.44",
                                     unit="units (AutoCAD text size)")
        self.f_text_size.pack(fill='x')

        ttk.Separator(tab).pack(fill='x', pady=8)

        self._section(tab, "Heavy Cable Run Custom Styling (Layer / Linetype Override)")
        
        self.f_heavy_section = _FieldRow(tab, "Target Heavy Section:", "1x10",
                                         unit="mm² (cables matching this string are custom-styled)")
        self.f_heavy_section.pack(fill='x')

        self.f_heavy_linetype = _FieldRow(tab, "Heavy Run Linetype:", "TRATTEGGIATA",
                                          unit="Linetype name to apply")
        self.f_heavy_linetype.pack(fill='x')

        self.f_heavy_color = _FieldRow(tab, "Heavy Run Color (ACI):", "40",
                                       unit="AutoCAD Color Index (e.g. 1=Red, 40=Orange/Brown)")
        self.f_heavy_color.pack(fill='x')

        self.f_heavy_layer = _FieldRow(tab, "Heavy Run Layer Name:", "TRATTEGGIATA",
                                       unit="Destination layer for heavy runs")
        self.f_heavy_layer.pack(fill='x')

    # ── Generate / Log Tab ────────────────────────────────────────────────────
    def _build_run_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Generate  ')

        ctrl = ttk.Frame(tab)
        ctrl.pack(fill='x', pady=(0, 6))

        self.gen_btn = ttk.Button(ctrl, text='▶  Generate SLD  (F5)',
                                   command=self._on_generate)
        self.gen_btn.pack(side='left', ipadx=10, ipady=4)

        ttk.Button(ctrl, text='Clear Log',
                   command=self._clear_log).pack(side='left', padx=8)

        self.progress = ttk.Progressbar(ctrl, mode='indeterminate', length=160)
        self.progress.pack(side='left', padx=8)

        self.status_var = tk.StringVar(value='Ready.')
        ttk.Label(tab, textvariable=self.status_var,
                  foreground='gray').pack(anchor='w', pady=(0, 4))

        ttk.Label(tab, text='Execution Progress Log',
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(4, 2))

        self.log_box = scrolledtext.ScrolledText(
            tab, state='disabled',
            font=('Consolas', 9),
            bg='#1e1e1e', fg='#d4d4d4',
            insertbackground='white',
            relief='flat', borderwidth=1)
        self.log_box.pack(fill='both', expand=True)

    def _section(self, parent, title):
        ttk.Label(parent, text=title,
                  font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(4, 2))

    # ── Helpers & Actions ─────────────────────────────────────────────────────
    def _log(self, msg):
        def _do():
            self.log_box.configure(state='normal')
            self.log_box.insert('end', msg + '\n')
            self.log_box.see('end')
            self.log_box.configure(state='disabled')
        self.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state='normal')
        self.log_box.delete('1.0', 'end')
        self.log_box.configure(state='disabled')

    def _collect(self):
        return {
            'template_dxf':      self.fe_tmpl.get(),
            'xlsx_path':         self.fe_xlsx.get(),
            'output_path':       self.fe_out.get(),
            'panel_model':       self.f_panel_model.get(),
            'panels_per_string': self.f_panels_str.get(),
            'inverter_model':    self.f_inv_model.get(),
            'dc_power_kwp':      self.f_dc_power.get(),
            'ac_power_kwac':     self.f_ac_power.get(),
            'temp_rating':       self.f_temp.get(),
            'transformer_power': self.f_tx_power.get(),
            
            # Workspace Parameters
            'col_spacing':       self.f_col_spacing.get(),
            'row_spacing':       self.f_row_spacing.get(),
            'circle_radius':     self.f_circle_radius.get(),
            'text_size':         self.f_text_size.get(),
            'heavy_section':     self.f_heavy_section.get(),
            'heavy_linetype':    self.f_heavy_linetype.get(),
            'heavy_color':       self.f_heavy_color.get(),
            'heavy_layer':       self.f_heavy_layer.get(),
        }

    def _validate(self, cfg):
        errs = []
        if not cfg.get('template_dxf'):
            errs.append("Template DXF path is required.")
        elif not os.path.isfile(cfg['template_dxf']):
            errs.append(f"Template DXF not found:\n  {cfg['template_dxf']}")
        if not cfg['xlsx_path']:
            errs.append("Excel file path is required.")
        elif not os.path.isfile(cfg['xlsx_path']):
            errs.append(f"Excel file not found:\n  {cfg['xlsx_path']}")
        if not cfg['output_path']:
            errs.append("Output DXF path is required.")
        else:
            out_dir = os.path.dirname(cfg['output_path'])
            if out_dir and not os.path.isdir(out_dir):
                errs.append(f"Output directory does not exist:\n  {out_dir}")
                
        # Numeric conversions check
        num_fields = (
            ('panels_per_string', 'Panels per String'),
            ('dc_power_kwp',      'DC Power'),
            ('ac_power_kwac',     'AC Power'),
            ('temp_rating',       'Temperature Rating'),
            ('col_spacing',       'Horizontal Col Step'),
            ('row_spacing',       'Vertical Row Step'),
            ('circle_radius',     'Module Circle Radius'),
            ('text_size',         'String Label Text Size'),
            ('heavy_color',       'Heavy Cable Run Color Index'),
        )
        for key, label in num_fields:
            val = cfg.get(key, '')
            if val:
                try:
                    float(val)
                except ValueError:
                    errs.append(f"{label}: must be a number (got '{val}').")
        return errs

    def _sync_output_path(self, *_):
        xlsx = self.fe_xlsx.get()
        if xlsx:
            self.fe_out.var.set(os.path.splitext(xlsx)[0] + '_SLD_Generated.dxf')

    def _refresh_all_combos(self):
        for combo in (self.f_panel_model, self.f_panels_str,
                      self.f_inv_model, self.f_dc_power, self.f_ac_power):
            combo.refresh()

    def _on_generate(self):
        cfg  = self._collect()
        errs = self._validate(cfg)
        if errs:
            messagebox.showerror("Input Error", "\n\n".join(errs))
            return

        self.gen_btn.configure(state='disabled')
        self.progress.start(12)
        self.status_var.set("Generating ...")
        self._log("=" * 64)
        self._log("Starting upgraded SLD diagram generation ...")

        def worker():
            try:
                _generate(cfg, self._log)
                self.after(0, self._on_success, cfg['output_path'])
            except Exception as ex:
                self.after(0, self._on_error, str(ex))

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self, path):
        self.progress.stop()
        self.gen_btn.configure(state='normal')
        self.status_var.set("Done — file saved.")
        for combo in (self.f_panel_model, self.f_panels_str,
                      self.f_inv_model, self.f_dc_power, self.f_ac_power):
            combo.record()
        
        # Open directory link confirm box
        msg = f"SLD generated successfully!\n\nOutput saved at:\n{path}"
        messagebox.showinfo("Success", msg)

    def _on_error(self, msg):
        self.progress.stop()
        self.gen_btn.configure(state='normal')
        self.status_var.set("Error — see log.")
        self._log(f"\n[ERROR] {msg}")
        messagebox.showerror("Generation Error", msg)


if __name__ == '__main__':
    app = SLDAppV2()
    app.mainloop()
