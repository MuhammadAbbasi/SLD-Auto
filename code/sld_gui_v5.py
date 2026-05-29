# -*- coding: utf-8 -*-
"""
sld_gui_v5.py  -  Autonomous DC Single Line Diagram Generator  v5
=================================================================
Geometry calibrated from "26S001_2E103 - DC Single Line Diagram.dxf"
(Yanel reference project).

Key design facts (measured from reference DXF):
  - Column step (xmin to xmin)     : 13 163 units
  - Row step (band_top to band_top): ~16 060 units (scaled per MPPT count)
  - xmin of column 0               : 15 277
  - band_top of row 0              : 167 671
  - Port labels at x-offset        : 5 531 from xmin
  - Circle centre x-offset         : 5 668 from xmin  (r = 24.6)
  - String labels x-offset         : 6 905 from xmin
  - Left connection line           : 5 194 --> 5 643 (circle left edge)
  - Right connection line          : 5 693 --> 6 892 (circle right edge --> label)
  - P00 fuse block insert x-offset : 5 432 from xmin
  - MPPT header between the 2 ports of each MPPT pair
  - No explicit outer box rectangle (open layout)

GUI:  python sld_gui_v5.py
CLI:  python sld_gui_v5.py --excel in.xlsx --output out.dxf [options]
"""

import argparse
import json
import os
import re
import sys
import threading
from collections import defaultdict
from datetime import datetime

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, scrolledtext, messagebox
    _HAS_TK = True
except ImportError:
    _HAS_TK = False

# ------------------------------------------------------------------ paths ---
_DIR       = os.path.dirname(os.path.abspath(__file__))
_LOGO_PATH = os.path.join(_DIR, 'logoA176LAB.jpg')
_HIST_PATH = os.path.join(_DIR, 'history_v2.json')

# ------------------------------------------ drawing geometry (DXF units) ---
# Calibrated from "Priory Farm - Lista Cavi - Cavi LV-DC.dxf" template

ORIGIN_X     = 15_277
ORIGIN_Y     = 167_671
COL_STEP     = 14_322   # column xmin-to-xmin
ROW_GAP      = 2_000    # extra gap below cable-info to next band_top

# X offsets from inv_x (column xmin)
PX_TITLE     = 553      # INVERTER title (above box)
PX_MODEL     = 1_279    # model name (above box)
PX_BOX_L     = 1_232    # AC box left wall
PX_AC_INNER  = 1_432    # 3~ symbol and AC section text
PX_BOX_R     = 3_404    # AC box right wall = vertical AC/DC divider
PX_CC        = 3_450    # CC side specs (just inside DC section)
PX_SW_LBL    = 5_683    # DC SWITCH text label
PX_SW_BUS    = 5_914    # DC SWITCH vertical bus X
PX_MPP       = 6_028    # MPP label
PX_CABLE_L   = 6_144    # port left cable start (from bus)
PX_P00       = 6_382    # P00 fuse insert X
PX_CIRC      = 6_618    # circle centre X
CIRCLE_R     = 24.6
PX_CABLE_R   = 7_843    # right cable end X
PX_STR_LBL   = 7_857    # string label X
PX_STR_W     = 5_800    # string label MTEXT width

# Y offsets from band_top (real_y = band_top - offset)
YO_TITLE     = 2_291    # INVERTER title (above box)
YO_MODEL     = 3_331    # model name text (above box top)
BOX_TOP_OFF  = 3_551    # box top edge
YO_PORT1     = 3_655    # first port Y (MPP1, string 1)
YO_3PHASE    = 3_677    # "3~" symbol (inside AC section)
YO_AC_SPECS  = 3_985    # AC specs text
YO_CABIN_LBL = 5_101    # cabin label inside AC section
YO_CC_SPECS  = 8_105    # CC specs (inside DC section)

# Port row spacing
PORT_INNER     = 211    # between consecutive strings of the same MPPT
PORT_BETWEEN   = 140    # between adjacent MPPTs within a switch group
DC_SW_EXTRA    = 358    # extra Y gap at each DC switch boundary
DC_SW_INTERVAL = 3      # MPPTs per switch group

CABLE_TAIL_GAP = 200    # gap from last port to cable-info line

# Text heights
TH_TITLE   = 220
TH_MODEL   = 170
TH_SPECS   = 50
TH_PANEL   = 60
TH_3PHASE  = 180
TH_MPP     = 50
TH_PORT    = 29
TH_STR     = 60
TH_CABIN   = 1_000
TH_SWITCH  = 40
TH_CABLE   = 47

# ACI colours
C_DEFAULT  = 7
C_RED      = 1
C_ORANGE   = 40
C_GREY     = 8
C_CYAN     = 4

# Layer names
L_0      = "0"
L_TESTI  = "-Testi"

# ----------------------------------------- inverter model power look-up ---
# ------------------------------------------ Excel auto-detection tables ---
_SHEET_CABLE = {
    '2e802-3', '2e802', 'cable schedule', 'cables', 'string list',
    'strings', 'dc cables', 'cavi dc', 'lista cavi', 'schedule',
    'inverter strings', 'mppt strings', 'dc string',
}
_SHEET_ROUTE = {
    'inverter to string', 'routing', 'cavi lv-dc', 'inverter_to_string',
    'its', 'routes', 'lv-dc',
}
_HDR_INV  = {'inverter', 'inv', 'invertitore', 'convertitore', 'inv id'}
_HDR_STR  = {'string name', 'string', 'stringa', 'nome stringa', 'cable name', 'nome cavo'}
_HDR_MPPT = {'mppt', 'mpp', 'mppt no', 'mppt number', 'mppt channel'}
_HDR_SEC  = {'section', 'sezione', 'sez', 'mm2', 'mmq', 'cross section', 'cable section'}
_HDR_WP   = {'module type', 'wp', 'watt peak', 'panel', 'modulo', 'module'}
_HDR_POS  = {'posizione stringa', 'position', 'tracker', 'table no', 'struttura'}
_HDR_LP   = {'l+', 'l plus', 'length +', 'lung+', 'l_plus'}
_HDR_LM   = {'l-', 'l minus', 'length -', 'lung-', 'l_minus'}

# ------------------------------------------------ DXF primitive helpers ---

def _a(**kw):
    """Return a dxfattribs dict, dropping None values."""
    return {k: v for k, v in kw.items() if v is not None}


def _hline(msp, x1, x2, y, layer=L_0, color=C_DEFAULT, lt=None, lw=None):
    msp.add_line((x1, y, 0), (x2, y, 0),
                 dxfattribs=_a(layer=layer, color=color, linetype=lt, lineweight=lw))


def _vline(msp, x, y1, y2, layer=L_0, color=C_DEFAULT):
    msp.add_line((x, y1, 0), (x, y2, 0),
                 dxfattribs=_a(layer=layer, color=color))


def _circle(msp, cx, cy, r, layer=L_0, color=C_DEFAULT):
    msp.add_circle((cx, cy, 0), r,
                   dxfattribs=_a(layer=layer, color=color))


def _mtext(msp, x, y, text, height=60, width=5000,
           layer=L_0, color=C_DEFAULT, attach=4, style='Standard'):
    """
    attach 4 = middle-left   (insert = left edge, vertically centred)
    attach 5 = middle-centre
    attach 2 = top-centre
    """
    msp.add_mtext(text, dxfattribs=_a(
        insert=(x, y, 0),
        char_height=height,
        width=width,
        attachment_point=attach,
        layer=layer,
        color=color,
        style=style,
    ))


def _blockref(msp, name, x, y, layer=L_0):
    msp.add_blockref(name, (x, y, 0), dxfattribs=_a(layer=layer))


# -------------------------------------------- geometry helpers -----------

def _band_top(row_idx, row_step):
    return ORIGIN_Y - row_idx * row_step


def _col_x(col_idx):
    return ORIGIN_X + col_idx * COL_STEP


def _port_y(band_top, m, p, n_spm, sw_interval=DC_SW_INTERVAL):
    """
    Y coordinate for port (m, p) where m=MPPT index (1-based), p=string index (1-based).
    Uses PORT_INNER within MPPT, PORT_BETWEEN between MPPTs, DC_SW_EXTRA at switch gaps.
    """
    if m == 1 and p == 1:
        return band_top - YO_PORT1
    # How many complete MPPTs have passed before this one
    completed_mpts = m - 1
    # How many DC switch boundaries have been crossed
    n_sw_gaps = completed_mpts // sw_interval
    # Cumulative offset
    offset = (completed_mpts * (PORT_INNER * (n_spm - 1) + PORT_BETWEEN)
              + (p - 1) * PORT_INNER
              + n_sw_gaps * DC_SW_EXTRA)
    return band_top - YO_PORT1 - offset


def _mpp_label_y(band_top, m, n_spm, sw_interval=DC_SW_INTERVAL):
    """Y for the MPP label — midpoint between port (m,1) and port (m,2), or just above if n_spm=1."""
    p1_y = _port_y(band_top, m, 1, n_spm, sw_interval)
    if n_spm >= 2:
        p2_y = _port_y(band_top, m, 2, n_spm, sw_interval)
        return (p1_y + p2_y) / 2
    return p1_y - 20


def _cable_y(band_top, n_mppts, n_spm, sw_interval=DC_SW_INTERVAL):
    """Y of the cable-info footer line."""
    last_y = _port_y(band_top, n_mppts, n_spm, n_spm, sw_interval)
    return last_y - CABLE_TAIL_GAP


def compute_row_step(n_mppts, n_spm, sw_interval=DC_SW_INTERVAL):
    """Total vertical step from band_top to next band_top."""
    last_port_offset = (
        (n_mppts - 1) * (PORT_INNER * (n_spm - 1) + PORT_BETWEEN)
        + (n_spm - 1) * PORT_INNER
        + ((n_mppts - 1) // sw_interval) * DC_SW_EXTRA
    )
    return YO_PORT1 + last_port_offset + CABLE_TAIL_GAP + ROW_GAP


# ----------------------------------------------- Excel parsing -----------

def _sheet_by_name(wb, known_set):
    for name in wb.sheetnames:
        if name.strip().lower() in known_set:
            return wb[name]
    return None


def _find_cable_sheet(wb, log):
    ws = _sheet_by_name(wb, _SHEET_CABLE)
    if ws:
        log(f"  Cable-schedule sheet: '{ws.title}'")
        return ws
    best_ws, best_hits = None, 0
    for name in wb.sheetnames:
        ws = wb[name]
        hits = sum(
            1 for r in range(1, min(400, ws.max_row + 1))
            for c in range(1, min(15, ws.max_column + 1))
            if re.match(r'^\d+\.\d+$', str(ws.cell(r, c).value or '').strip())
        )
        if hits > best_hits:
            best_hits, best_ws = hits, ws
    if best_ws and best_hits >= 3:
        log(f"  Auto-detected cable sheet: '{best_ws.title}' ({best_hits} hits)")
        return best_ws
    ws = wb[wb.sheetnames[0]]
    log(f"  [warn] Using first sheet '{ws.title}'")
    return ws


def _find_headers(ws, log):
    best_row, best_map, best_score = None, {}, 0
    for r in range(1, 81):
        found = {}
        for c in range(1, min(55, ws.max_column + 1)):
            val = str(ws.cell(r, c).value or '').strip().lower()
            if not val:
                continue
            if any(k == val or val.startswith(k) for k in _HDR_INV):  found.setdefault('inv',  c)
            if any(k in val for k in _HDR_STR):   found.setdefault('str',  c)
            if any(k in val for k in _HDR_MPPT):  found.setdefault('mppt', c)
            if any(k in val for k in _HDR_SEC):   found.setdefault('sec',  c)
            if any(k in val for k in _HDR_WP):    found.setdefault('wp',   c)
            if any(k in val for k in _HDR_POS):   found.setdefault('pos',  c)
            if any(k in val for k in _HDR_LP):    found.setdefault('lp',   c)
            if any(k in val for k in _HDR_LM):    found.setdefault('lm',   c)
        score = sum(1 for k in ('inv', 'str', 'mppt') if k in found)
        if score > best_score:
            best_score, best_row, best_map = score, r, found
    if best_score >= 2:
        log(f"  Header row {best_row}: {best_map}")
        return best_row, best_map
    log("  [warn] No header row found; using legacy column positions")
    return 30, {'inv': 1, 'str': 3, 'mppt': 4, 'sec': 25, 'wp': 20}


def _parse_routing(wb, log):
    ws = _sheet_by_name(wb, _SHEET_ROUTE)
    if not ws:
        return {}
    data = {}
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        lp  = ws.cell(r, 2).value
        lm  = ws.cell(r, 3).value
        tbl = ws.cell(r, 5).value
        if sid:
            data[str(sid).strip()] = {
                'l_plus':    float(lp)  if lp  else 0.0,
                'l_minus':   float(lm)  if lm  else 0.0,
                'table_num': str(tbl).strip() if tbl else '',
            }
    log(f"  Routing sheet '{ws.title}': {len(data)} records")
    return data


def _parse_cable_sheet(ws, header_row, col_map, routing, log):
    c_inv  = col_map.get('inv',  1)
    c_str  = col_map.get('str',  3)
    c_mppt = col_map.get('mppt', 4)
    c_sec  = col_map.get('sec')
    c_wp   = col_map.get('wp')
    c_pos  = col_map.get('pos')
    excel  = {}
    cur    = None
    for r in range(header_row + 1, ws.max_row + 1):
        inv   = ws.cell(r, c_inv).value
        sname = ws.cell(r, c_str).value
        mppt  = ws.cell(r, c_mppt).value
        if inv is not None and '.' in str(inv):
            try:
                parts = str(inv).split('.')
                T, I  = int(parts[0]), int(parts[1])
                cur   = (T, I)
                excel.setdefault(cur, defaultdict(list))
            except Exception:
                pass
        if sname and cur and mppt:
            sname_s = str(sname).strip()
            route   = routing.get(sname_s, {})
            section = '1x6'
            if c_sec:
                s = str(ws.cell(r, c_sec).value or '').strip()
                if s and s.lower() not in ('none', 'n/a', ''):
                    section = s
            wp = 0
            if c_wp:
                try:
                    wp = int(float(ws.cell(r, c_wp).value or 0))
                except Exception:
                    pass
            tracker = ''
            if c_pos:
                t = str(ws.cell(r, c_pos).value or '').strip()
                if t.lower() not in ('none', 'no piling information', 'n/a', ''):
                    tracker = t
                else:
                    tracker = route.get('table_num', '')
            try:
                excel[cur][int(float(mppt))].append({
                    'name':        sname_s,
                    'wp':          wp,
                    'l_plus':      float(route.get('l_plus', 0)),
                    'l_minus':     float(route.get('l_minus', 0)),
                    'section':     section,
                    'tracker_pos': tracker,
                })
            except Exception:
                pass
    return excel


def parse_excel(xlsx_path, log):
    """Autonomous Excel parse. Returns (excel, inv_list, transformers, transformer_list)."""
    try:
        import openpyxl as _xl
    except ImportError:
        raise RuntimeError("openpyxl not installed.  Run:  pip install openpyxl")
    log("Reading Excel workbook ...")
    wb = _xl.load_workbook(xlsx_path, data_only=True)
    log(f"  Sheets: {wb.sheetnames}")
    routing          = _parse_routing(wb, log)
    ws               = _find_cable_sheet(wb, log)
    hdr_row, col_map = _find_headers(ws, log)
    excel            = _parse_cable_sheet(ws, hdr_row, col_map, routing, log)
    inv_list = sorted(excel.keys())
    transformers = {}
    for T, I in inv_list:
        transformers.setdefault(T, []).append(I)
    for T in transformers:
        transformers[T] = sorted(transformers[T])
    transformer_list = sorted(transformers.keys())
    log(f"  Inverters: {len(inv_list)}   Transformers: {len(transformer_list)}")
    for T in transformer_list:
        ii = transformers[T]
        mc = [len(excel.get((T, i), {}))                           for i in ii]
        sc = [sum(len(v) for v in excel.get((T, i), {}).values()) for i in ii]
        log(f"    Tx{T}: {len(ii)} inv  MPPTs {min(mc)}-{max(mc)}  Strings {min(sc)}-{max(sc)}")
    return excel, inv_list, transformers, transformer_list


# ----------------------------------------------- DXF generation engine ---

def _define_p00_block(doc, log):
    """
    Define the P00 fuse/disconnector block (two diagonal lines).
    Matches the P00 block from the reference DXF exactly.
    """
    if 'P00' in doc.blocks:
        return
    try:
        blk = doc.blocks.new('P00')
        blk.add_lwpolyline([(-18.7, 25.7), (6.8, -25.3)],
                           dxfattribs={'layer': L_0})
        blk.add_lwpolyline([(19.6, -25.3), (-5.9, 25.7)],
                           dxfattribs={'layer': L_0})
    except Exception as ex:
        log(f"  [warn] P00 block: {ex}")


def _setup_doc(doc, log):
    """Register layers and text styles in a new blank DXF."""
    layer_defs = [
        (L_0,      C_DEFAULT, 'Continuous'),
        (L_TESTI,  C_CYAN,    'Continuous'),
    ]
    for lname, col, lt in layer_defs:
        if lname not in doc.layers:
            try:
                doc.layers.new(lname, dxfattribs={'color': col, 'linetype': lt})
            except Exception:
                pass
    for name in ('Standard', 'SIMPLEX'):
        try:
            if name in doc.styles:
                doc.styles.get(name).dxf.font = 'arial.ttf'
            else:
                doc.styles.new(name, dxfattribs={'font': 'arial.ttf'})
        except Exception:
            pass


def _draw_inv_box(msp, inv_x, band_top, n_mppts, n_spm):
    """
    Draw the inverter box: AC section rectangle + diagonal inverter symbol.
    Returns (box_top, box_bot).
    """
    box_top = band_top - BOX_TOP_OFF
    cable_y = _cable_y(band_top, n_mppts, n_spm)
    box_bot = cable_y - 150
    box_l   = inv_x + PX_BOX_L
    box_r   = inv_x + PX_BOX_R

    # Outer AC section rectangle
    pts = [(box_l, box_top), (box_r, box_top), (box_r, box_bot), (box_l, box_bot)]
    msp.add_lwpolyline(pts, close=True, dxfattribs={'layer': L_0, 'color': C_DEFAULT})

    # Diagonal "/" inverter symbol inside AC box
    mg = 180
    msp.add_line((box_l + mg, box_bot + mg), (box_r - mg, box_top - mg),
                 dxfattribs={'layer': L_0, 'color': C_DEFAULT})

    return box_top, box_bot


def _draw_dc_switch_bus(msp, inv_x, band_top, n_mppts, n_spm, sw_num, sw_interval=DC_SW_INTERVAL):
    """
    Draw the DC SWITCH bus structure for switch group sw_num (1-based).
    Groups the MPPTs in range [(sw_num-1)*sw_interval+1 .. sw_num*sw_interval].
    """
    m_first = (sw_num - 1) * sw_interval + 1
    m_last  = min(sw_num * sw_interval, n_mppts)

    top_y  = _port_y(band_top, m_first, 1,    n_spm, sw_interval)
    bot_y  = _port_y(band_top, m_last,  n_spm, n_spm, sw_interval)
    mid_y  = (top_y + bot_y) / 2

    bx = inv_x + PX_BOX_R
    vx = inv_x + PX_SW_BUS
    lx = inv_x + PX_CABLE_L

    # Horizontal stub: AC box wall → vertical bus
    msp.add_line((bx, mid_y), (vx, mid_y),
                 dxfattribs={'layer': L_0, 'color': C_DEFAULT})

    # Vertical bus spanning all ports in this switch group
    msp.add_line((vx, top_y), (vx, bot_y),
                 dxfattribs={'layer': L_0, 'color': C_DEFAULT})

    # Short horizontal stub from bus to each port's left cable start
    for m in range(m_first, m_last + 1):
        for p in range(1, n_spm + 1):
            py = _port_y(band_top, m, p, n_spm, sw_interval)
            msp.add_line((vx, py), (lx, py),
                         dxfattribs={'layer': L_0, 'color': C_DEFAULT})

    # DC SWITCH label
    _mtext(msp, inv_x + PX_SW_LBL, mid_y,
           f"DC SWITCH {sw_num}", height=TH_SWITCH,
           width=PX_SW_BUS - PX_SW_LBL - 50,
           color=C_CYAN, layer=L_TESTI, attach=4)


def _draw_port_row(msp, inv_x, band_top, m, p, sdata, cfg, n_spm,
                   sw_interval=DC_SW_INTERVAL, use_dc_sw=True):
    """
    Draw one port row.
    Active: left stub + P00 fuse + circle + right cable + port label + string label.
    Reserve: short grey stub + port label only.
    Left cable starts at PX_CABLE_L (DC switch bus) when DC switch enabled,
    or at PX_BOX_R (AC box wall) when disabled.
    """
    py = _port_y(band_top, m, p, n_spm, sw_interval)

    is_active = sdata is not None

    # Reserve port: minimal stub
    if not is_active:
        stub_start = inv_x + (PX_CABLE_L if use_dc_sw else PX_BOX_R)
        _hline(msp, stub_start, stub_start + 360, py, color=C_GREY)
        _mtext(msp, inv_x + PX_MPP + (PX_CIRC - PX_MPP) // 2, py + TH_PORT * 0.5,
               f"{m}-{p}", height=TH_PORT, width=300, color=C_GREY, attach=4)
        return

    heavy_sec  = cfg.get('heavy_section', '1x10')
    heavy_col  = int(float(cfg.get('heavy_color') or C_ORANGE))
    line_color = heavy_col if sdata.get('section') == heavy_sec else C_ORANGE

    # Left cable: from cable start to circle left edge
    lx_start = inv_x + (PX_CABLE_L if use_dc_sw else PX_BOX_R)
    _hline(msp, lx_start, inv_x + PX_CIRC - CIRCLE_R, py, color=line_color)

    # P00 fuse block (only draw if it's within the cable range)
    _blockref(msp, 'P00', inv_x + PX_P00, py)

    # Junction circle
    _circle(msp, inv_x + PX_CIRC, py, CIRCLE_R, color=line_color)

    # Right cable: from circle right edge to string label area
    _hline(msp, inv_x + PX_CIRC + CIRCLE_R, inv_x + PX_CABLE_R, py, color=line_color)

    # Port label above the line
    _mtext(msp, inv_x + PX_MPP + 20, py + TH_PORT * 0.5,
           f"{m}-{p}", height=TH_PORT, width=PX_CIRC - PX_MPP - 30,
           color=line_color, attach=4)

    # String label
    pps       = int(float(cfg.get('panels_per_string') or 0))
    panel_mdl = cfg.get('panel_model', '').strip()
    label     = f"String {sdata['name']}"
    if pps > 0 and cfg.get('inline_panel_info', False):
        wp  = sdata.get('wp', 0)
        sfx = f" {wp}Wp" if wp > 0 else ""
        label += f" - {pps}x {panel_mdl}{sfx}" if panel_mdl else f" - {pps}P{sfx}"
    if cfg.get('show_cable_info', False):
        label += (f" (L+={sdata.get('l_plus', 0):.1f}m"
                  f" L-={sdata.get('l_minus', 0):.1f}m"
                  f" {sdata.get('section', '')})")

    _mtext(msp, inv_x + PX_STR_LBL, py,
           label, height=TH_STR, width=PX_STR_W, color=line_color, attach=4)


def _draw_inverter(msp, inv_x, band_top, T, I, inv_data,
                   global_max_mppt, global_max_spm, cfg, log):
    """Draw one complete inverter column."""
    use_dc_sw = bool(cfg.get('dc_switch', True))
    sw_iv     = DC_SW_INTERVAL

    box_top, box_bot = _draw_inv_box(msp, inv_x, band_top, global_max_mppt, global_max_spm)

    inv_model = cfg.get('inverter_model', '').strip()
    pps       = int(float(cfg.get('panels_per_string') or 0))
    panel_mdl = cfg.get('panel_model', '').strip()

    # ---------------------------------------------------------- title (above box)
    _mtext(msp, inv_x + PX_TITLE, band_top - YO_TITLE,
           f"INVERTER {T}.{I}", height=TH_TITLE,
           width=COL_STEP - PX_TITLE - 100, color=C_DEFAULT, attach=4)

    # ---------------------------------------------------------- model (above box)
    if inv_model:
        _mtext(msp, inv_x + PX_MODEL, band_top - YO_MODEL,
               inv_model.upper(), height=TH_MODEL,
               width=PX_BOX_R - PX_MODEL - 80, color=C_DEFAULT, attach=4)

    # ---------------------------------------------------------- AC section text
    ac_w = PX_BOX_R - PX_AC_INNER - 60

    # "3~" three-phase symbol
    _mtext(msp, inv_x + PX_AC_INNER, band_top - YO_3PHASE,
           '3~', height=TH_3PHASE, width=ac_w, color=C_DEFAULT, attach=4)

    # Cabin label
    _mtext(msp, inv_x + PX_AC_INNER, band_top - YO_CABIN_LBL,
           f'Cabin Tx.{T}\\PInverter {T}.{I}',
           height=TH_SPECS + 10, width=ac_w, color=C_GREY, attach=4)

    # ---------------------------------------------------------- CC specs (DC section)
    n_inputs = global_max_mppt * global_max_spm
    cc_txt = (f"CC side\\P{n_inputs} input - {global_max_mppt} MPPT\\P"
              f"Max Vdc : 1'500V")
    _mtext(msp, inv_x + PX_CC, band_top - YO_CC_SPECS,
           cc_txt, height=TH_SPECS, width=PX_SW_LBL - PX_CC - 60,
           color=C_DEFAULT, attach=4)

    # ---------------------------------------------------------- panel info
    all_wp = sorted({sd['wp'] for strs in inv_data.values()
                     for sd in strs if sd.get('wp', 0) > 0})
    wp_txt = f"{all_wp[-1]}Wp" if all_wp else ""
    if pps > 0:
        panel_info = f"{pps} PV modules in series"
        if panel_mdl:
            panel_info += f" - {panel_mdl}"
            if wp_txt:
                panel_info += f"  {wp_txt}"
        elif wp_txt:
            panel_info += f" - {wp_txt}"
        top_port_y = band_top - YO_PORT1
        _mtext(msp, inv_x + PX_STR_LBL, top_port_y + TH_PANEL * 0.8,
               panel_info, height=TH_PANEL, width=PX_STR_W, color=C_DEFAULT, attach=4)

    # ---------------------------------------------------------- DC SWITCH buses
    if use_dc_sw:
        n_switches = (global_max_mppt + sw_iv - 1) // sw_iv
        for sw in range(1, n_switches + 1):
            _draw_dc_switch_bus(msp, inv_x, band_top,
                                global_max_mppt, global_max_spm, sw, sw_iv)
    else:
        # Direct connection: horizontal bar from AC box wall to cable start area
        for m in range(1, global_max_mppt + 1):
            for p in range(1, global_max_spm + 1):
                py = _port_y(band_top, m, p, global_max_spm, sw_iv)
                msp.add_line((inv_x + PX_BOX_R, py), (inv_x + PX_CABLE_L, py),
                             dxfattribs={'layer': L_0, 'color': C_DEFAULT})

    # ---------------------------------------------------------- port rows
    for m in range(1, global_max_mppt + 1):
        for p in range(1, global_max_spm + 1):
            strings   = inv_data.get(m, [])
            is_active = (p - 1) < len(strings)
            sdata     = strings[p - 1] if is_active else None
            _draw_port_row(msp, inv_x, band_top, m, p, sdata, cfg,
                           global_max_spm, sw_iv, use_dc_sw)

        # MPP label between first and second port of this MPPT
        mpp_y = _mpp_label_y(band_top, m, global_max_spm, sw_iv)
        _mtext(msp, inv_x + PX_MPP, mpp_y,
               f"MPP{m}", height=TH_MPP,
               width=PX_CABLE_L - PX_MPP - 30,
               color=C_CYAN, layer=L_TESTI, attach=4)

    # ---------------------------------------------------------- cable info footer
    cable_y = _cable_y(band_top, global_max_mppt, global_max_spm, sw_iv)
    sections = {sd['section']
                for strs in inv_data.values()
                for sd in strs if sd.get('section')}
    cable_txt = (f"2/ {next(iter(sections))} mmq - Cu - H1Z2Z2k" if len(sections) == 1
                 else "2/ (1x6/10)mmq - Cu - H1Z2Z2k" if sections
                 else "2/ 1x6 mmq - Cu - H1Z2Z2k")
    _mtext(msp, inv_x + PX_STR_LBL, cable_y,
           cable_txt, height=TH_CABLE, width=PX_STR_W,
           color=C_DEFAULT, layer=L_TESTI, attach=4)


def _draw_cabin_header(msp, center_x, band_top, T, n_inv, col_step=COL_STEP):
    """Draw the CABIN N header centered over all inverter columns for transformer T."""
    _mtext(msp, center_x, band_top,
           f"CABIN {T}", height=TH_CABIN, width=n_inv * col_step,
           color=C_DEFAULT, layer=L_0, attach=2)


def _generate(cfg, log):
    """Core generation: parse Excel -> draw from scratch -> save DXF."""
    try:
        import ezdxf as _ez
    except ImportError:
        raise RuntimeError("ezdxf not installed.  Run:  pip install ezdxf")

    xlsx   = cfg['xlsx_path']
    output = cfg['output_path']

    max_mppts_raw = str(cfg.get('max_mppts', 'Auto')).strip()

    # -- 1. Parse Excel -------------------------------------------------------
    excel, inv_list, transformers, transformer_list = parse_excel(xlsx, log)
    if not inv_list:
        raise RuntimeError("No inverter data found in the Excel file.")

    # -- 2. Global layout dimensions ------------------------------------------
    if max_mppts_raw.lower() == 'auto':
        global_max_mppt = max(max(d.keys()) for d in excel.values() if d)
        log(f"Auto-detected max MPPT: {global_max_mppt}")
    else:
        global_max_mppt = max(1, int(float(max_mppts_raw)))
        log(f"User-specified max MPPT: {global_max_mppt}")

    global_max_spm = max(
        (max(len(v) for v in d.values()) for d in excel.values() if d),
        default=1
    )
    log(f"Max strings per MPPT: {global_max_spm}")

    row_step_raw = cfg.get('row_spacing', '').strip()
    if row_step_raw and row_step_raw.lower() != 'auto':
        row_step = float(row_step_raw)
    else:
        row_step = compute_row_step(global_max_mppt, global_max_spm)
    log(f"Row step: {row_step:.0f} units")

    col_step_raw = cfg.get('col_spacing', '').strip()
    col_step = float(col_step_raw) if col_step_raw and col_step_raw.lower() != 'auto' else COL_STEP
    log(f"Column step: {col_step:.0f} units")

    # -- 3. Create DXF --------------------------------------------------------
    log("Creating DXF document from scratch ...")
    doc = _ez.new('R2010')
    msp = doc.modelspace()
    _setup_doc(doc, log)
    _define_p00_block(doc, log)

    for lname in [l.name for l in doc.layouts if not l.is_modelspace]:
        try:
            doc.layouts.delete(lname)
        except Exception:
            pass

    # -- 4. Position mapping --------------------------------------------------
    tx_row  = {T: i for i, T in enumerate(transformer_list)}
    inv_col = {}
    for T, invs in transformers.items():
        for idx, I in enumerate(invs):
            inv_col[(T, I)] = idx

    def get_inv_x(T, I):
        return ORIGIN_X + inv_col[(T, I)] * col_step

    def get_band_top(T):
        return ORIGIN_Y - tx_row[T] * row_step

    # -- 5. Draw inverters ----------------------------------------------------
    log(f"Drawing {len(inv_list)} inverter columns ...")
    drawn_cabin = set()

    for idx, (T, I) in enumerate(inv_list):
        inv_x    = get_inv_x(T, I)
        band_top = get_band_top(T)

        _draw_inverter(msp, inv_x, band_top, T, I,
                       excel.get((T, I), {}),
                       global_max_mppt, global_max_spm, cfg, log)

        if T not in drawn_cabin:
            drawn_cabin.add(T)
            n_inv_t  = len(transformers[T])
            cabin_cx = ORIGIN_X + n_inv_t * col_step / 2
            _draw_cabin_header(msp, cabin_cx, band_top, T, n_inv_t, col_step)

        if (idx + 1) % 10 == 0 or (idx + 1) == len(inv_list):
            log(f"  {idx + 1}/{len(inv_list)} done")

    # -- 6. Paper-space viewports (A3 landscape) ------------------------------
    log(f"Creating {len(transformer_list)} paper-space layouts ...")
    for T in transformer_list:
        lname = f"Tx{T}"
        try:
            layout = doc.layouts.new(lname)
        except Exception:
            layout = doc.layouts.get(lname)

        n_inv   = len(transformers[T])
        row     = tx_row[T]
        bt      = get_band_top(T)
        row_cx  = ORIGIN_X + (n_inv - 1) * col_step / 2 + COL_STEP / 2
        row_cy  = bt - row_step / 2
        row_w   = n_inv * col_step
        view_h  = max(row_step * 0.90, row_w / (420.0 / 297.0) * 1.05)
        layout.add_viewport(
            center=(210, 148.5), size=(420, 297),
            view_center_point=(row_cx, row_cy),
            view_height=view_h,
        )
        log(f"  '{lname}': {n_inv} inverters")

    # -- 7. Save --------------------------------------------------------------
    log(f"Saving -> {output}")
    doc.saveas(output)
    log("Done!  SLD generated successfully.")


# -------------------------------------------------- history store ---------
_DEFAULTS = {
    'panel_model': [
        'JA Solar JAM66D50-645/GB',
        'JA Solar JAM72D42-625/LB',
        'JA Solar JAM72S20-460/MR',
        'Longi Solar LR5-72HBD-580M',
        'Canadian Solar HiKu7 CS7N-655MB',
        'Jinko Solar JKM660M-78HL4-V',
        'Trina Solar TSM-670NEG21C.20',
        'REC Alpha Pure-R 430AA',
    ],
    'panels_per_string': ['16', '18', '20', '22', '24', '26', '28', '30'],
    'inverter_model': [
        'Sungrow SGX350HX',
        'Sungrow SG350HX',
        'Sungrow SG250HX',
        'Sungrow SG125HX',
        'Huawei SUN2000-330KTL',
        'Huawei SUN2000-275KTL-H1',
        'Huawei SUN2000-215KTL-H3',
        'ABB PVS-250-TL',
        'SMA Sunny Tripower Core2 150',
        'Fronius Symo GEN24 25.0 Plus',
    ],
}


class _HistoryStore:
    def __init__(self, path):
        self._path = path
        self._data: dict = {}
        self._load()

    def _load(self):
        try:
            if os.path.isfile(self._path):
                with open(self._path, encoding='utf-8') as f:
                    self._data = json.load(f)
        except Exception:
            self._data = {}

    def _save(self):
        try:
            with open(self._path, 'w', encoding='utf-8') as f:
                json.dump(self._data, f, indent=2, ensure_ascii=False)
        except Exception as ex:
            print(f"[warn] Cannot save history: {ex}", file=sys.stderr)

    def values(self, key):
        entries = sorted(self._data.get(key, []),
                         key=lambda e: e.get('ts', ''), reverse=True)
        result  = [e['value'] for e in entries]
        for v in _DEFAULTS.get(key, []):
            if v not in result:
                result.append(v)
        return result

    def record(self, key, value):
        value = value.strip()
        if not value:
            return
        ts      = datetime.now().isoformat(timespec='seconds')
        entries = self._data.setdefault(key, [])
        for e in entries:
            if e['value'] == value:
                e['ts'] = ts
                break
        else:
            entries.append({'value': value, 'ts': ts})
        self._save()

    def all_entries(self, key):
        return list(self._data.get(key, []))

    def add_manual(self, key, value, note=''):
        value = value.strip()
        if not value:
            return
        entries = self._data.setdefault(key, [])
        for e in entries:
            if e['value'] == value:
                if note.strip():
                    e['note'] = note.strip()
                break
        else:
            entry = {'value': value}
            if note.strip():
                entry['note'] = note.strip()
            entries.append(entry)
        self._save()

    def delete_entry(self, key, value):
        entries = self._data.get(key, [])
        self._data[key] = [e for e in entries if e['value'] != value]
        self._save()


_history = _HistoryStore(_HIST_PATH)

_FIELD_LABELS = {
    'panel_model':       'Panel Model',
    'panels_per_string': 'Panels per String',
    'inverter_model':    'Inverter Model',
}


# ------------------------------------------------------- GUI widgets ------
_PAD = {'padx': 6, 'pady': 3}


def _load_logo_tk(size=(90, 90)):
    try:
        from PIL import Image, ImageTk
        img = Image.open(_LOGO_PATH).convert('RGBA')
        img.thumbnail(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None


class _HistoryCombo(ttk.Frame):
    def __init__(self, parent, label, history_key, default='', unit='',
                 width=28, refresh_callback=None, on_select=None, **kw):
        super().__init__(parent, **kw)
        self._key        = history_key
        self._refresh_cb = refresh_callback
        self._select_cb  = on_select
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=default)
        self._combo = ttk.Combobox(self, textvariable=self.var, width=width,
                                   values=_history.values(history_key))
        self._combo.pack(side='left', **_PAD)
        if unit:
            ttk.Label(self, text=unit, foreground='gray').pack(side='left')
        ttk.Button(self, text='...', width=3,
                   command=self._open_manager).pack(side='left', padx=(6, 0))
        self._combo.bind('<<ComboboxSelected>>', self._on_combo_select)

    def _on_combo_select(self, _=None):
        if self._select_cb:
            self._select_cb(self.get())

    def _open_manager(self):
        _PresetManagerDialog(self.winfo_toplevel(), initial_key=self._key,
                             on_close=self._refresh_cb)

    def get(self):
        return self.var.get().strip()

    def refresh(self):
        self._combo['values'] = _history.values(self._key)

    def record(self):
        v = self.get()
        if v:
            _history.record(self._key, v)
            self._combo['values'] = _history.values(self._key)


class _FileRow(ttk.Frame):
    def __init__(self, parent, label, default='', save=False,
                 filetypes=None, **kw):
        super().__init__(parent, **kw)
        self._save      = save
        self._filetypes = filetypes or [('All files', '*.*')]
        ttk.Label(self, text=label, width=18, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=default)
        ttk.Entry(self, textvariable=self.var, width=52).pack(
            side='left', padx=(0, 4), fill='x', expand=True)
        ttk.Button(self, text='Browse...', width=9,
                   command=self._browse).pack(side='left')

    def _browse(self):
        cur      = self.var.get()
        init_dir = os.path.dirname(cur) if cur else os.path.expanduser('~')
        if self._save:
            p = filedialog.asksaveasfilename(
                initialdir=init_dir, defaultextension='.dxf',
                filetypes=self._filetypes)
        else:
            p = filedialog.askopenfilename(
                initialdir=init_dir, filetypes=self._filetypes)
        if p:
            self.var.set(p)

    def get(self):
        return self.var.get().strip()


class _FieldRow(ttk.Frame):
    def __init__(self, parent, label, default='', unit='', width=20, **kw):
        super().__init__(parent, **kw)
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.StringVar(value=str(default))
        ttk.Entry(self, textvariable=self.var, width=width).pack(side='left', **_PAD)
        if unit:
            ttk.Label(self, text=unit, foreground='gray').pack(side='left')

    def get(self):
        return self.var.get().strip()


class _CheckboxRow(ttk.Frame):
    def __init__(self, parent, label, default=True, **kw):
        super().__init__(parent, **kw)
        ttk.Label(self, text=label, width=26, anchor='e').pack(side='left', **_PAD)
        self.var = tk.BooleanVar(value=default)
        ttk.Checkbutton(self, variable=self.var).pack(side='left', **_PAD)

    def get(self):
        return self.var.get()


class _PresetManagerDialog(tk.Toplevel):
    def __init__(self, parent, initial_key=None, on_close=None):
        super().__init__(parent)
        self.title("Manage Presets")
        self.geometry("720x520")
        self.minsize(580, 380)
        self._on_close = on_close
        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._close)
        self._build(initial_key)

    def _build(self, initial_key):
        top = ttk.Frame(self, padding=(12, 10, 12, 4))
        top.pack(fill='x')
        ttk.Label(top, text="Field:").pack(side='left', padx=(0, 6))
        self._field_var = tk.StringVar()
        self._keys      = list(_FIELD_LABELS.keys())
        labels          = [_FIELD_LABELS[k] for k in self._keys]
        self._field_cb  = ttk.Combobox(top, textvariable=self._field_var,
                                       values=labels, state='readonly', width=32)
        self._field_cb.pack(side='left')
        self._field_cb.bind('<<ComboboxSelected>>', lambda _: self._refresh_tree())

        mid = ttk.Frame(self, padding=(12, 4))
        mid.pack(fill='both', expand=True)
        cols = ('value', 'note', 'last_used', 'kind')
        self._tree = ttk.Treeview(mid, columns=cols, show='headings', selectmode='browse')
        for col, w, txt in [('value', 210, 'Value'), ('note', 210, 'Note'),
                             ('last_used', 140, 'Last Used'), ('kind', 70, 'Type')]:
            self._tree.heading(col, text=txt)
            self._tree.column(col, width=w, minwidth=60)
        vsb = ttk.Scrollbar(mid, orient='vertical', command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='left', fill='y')
        self._tree.bind('<<TreeviewSelect>>', self._on_select)

        form = ttk.LabelFrame(self, text='Add / Update Entry', padding=(10, 6))
        form.pack(fill='x', padx=12, pady=(4, 0))
        for lbl, attr in (('Value:', '_val_var'), ('Note:', '_note_var')):
            r = ttk.Frame(form)
            r.pack(fill='x', pady=2)
            ttk.Label(r, text=lbl, width=8, anchor='e').pack(side='left')
            setattr(self, attr, tk.StringVar())
            ttk.Entry(r, textvariable=getattr(self, attr), width=50).pack(side='left', padx=4)
        btns = ttk.Frame(form)
        btns.pack(fill='x', pady=(6, 2))
        ttk.Button(btns, text='Add / Update', command=self._add_entry).pack(side='left', padx=(0, 8))
        self._del_btn = ttk.Button(btns, text='Delete Selected',
                                   command=self._delete_entry, state='disabled')
        self._del_btn.pack(side='left')

        foot = ttk.Frame(self, padding=(12, 6))
        foot.pack(fill='x')
        ttk.Button(foot, text='Close', command=self._close).pack(side='right')

        if initial_key and initial_key in self._keys:
            self._field_cb.current(self._keys.index(initial_key))
        else:
            self._field_cb.current(0)
        self._refresh_tree()

    def _current_key(self):
        label = self._field_var.get()
        return next((k for k, v in _FIELD_LABELS.items() if v == label), None)

    def _refresh_tree(self):
        self._tree.delete(*self._tree.get_children())
        key = self._current_key()
        if not key:
            return
        user_entries = _history.all_entries(key)
        user_values  = {e['value'] for e in user_entries}
        for e in sorted(user_entries, key=lambda x: x.get('ts', ''), reverse=True):
            self._tree.insert('', 'end', iid=f'u|{e["value"]}',
                              values=(e['value'], e.get('note', ''),
                                      e.get('ts', ''), 'User'), tags=('user',))
        for v in _DEFAULTS.get(key, []):
            if v not in user_values:
                self._tree.insert('', 'end', iid=f'd|{v}',
                                  values=(v, '', '', 'Default'), tags=('default',))
        self._tree.tag_configure('default', foreground='gray')
        self._del_btn.configure(state='disabled')

    def _on_select(self, _):
        sel = self._tree.selection()
        if not sel:
            return
        row = self._tree.item(sel[0], 'values')
        self._val_var.set(row[0])
        self._note_var.set(row[1])
        self._del_btn.configure(state='normal' if sel[0].startswith('u|') else 'disabled')

    def _add_entry(self):
        key   = self._current_key()
        value = self._val_var.get().strip()
        if not key or not value:
            messagebox.showwarning("Input Required", "Please enter a value.", parent=self)
            return
        _history.add_manual(key, value, self._note_var.get())
        self._val_var.set('')
        self._note_var.set('')
        self._refresh_tree()

    def _delete_entry(self):
        sel = self._tree.selection()
        if not sel or not sel[0].startswith('u|'):
            return
        value = self._tree.item(sel[0], 'values')[0]
        if messagebox.askyesno("Confirm Delete",
                                f"Delete preset:\n\n  {value}\n\nCannot be undone.",
                                parent=self):
            _history.delete_entry(self._current_key(), value)
            self._refresh_tree()

    def _close(self):
        self.grab_release()
        self.destroy()
        if self._on_close:
            self._on_close()


# ----------------------------------------------- Scan preview dialog ------

class _ScanDialog(tk.Toplevel):
    def __init__(self, parent, xlsx_path):
        super().__init__(parent)
        self.title("Excel Scan Results")
        self.geometry("780x520")
        self.minsize(600, 380)
        self.transient(parent)
        self.grab_set()
        self._build()
        self._run(xlsx_path)

    def _build(self):
        ttk.Label(self, text="Autonomous Excel Parse Preview",
                  font=('Segoe UI', 11, 'bold')).pack(anchor='w', padx=12, pady=(10, 4))
        self._log_box = scrolledtext.ScrolledText(
            self, state='disabled', font=('Consolas', 9),
            bg='#1e1e1e', fg='#d4d4d4', relief='flat', borderwidth=1)
        self._log_box.pack(fill='both', expand=True, padx=12, pady=(0, 8))
        ttk.Button(self, text='Close',
                   command=self.destroy).pack(side='right', padx=12, pady=(0, 10))

    def _append(self, msg):
        self._log_box.configure(state='normal')
        self._log_box.insert('end', msg + '\n')
        self._log_box.see('end')
        self._log_box.configure(state='disabled')

    def _run(self, xlsx_path):
        def worker():
            try:
                excel, inv_list, transformers, transformer_list = parse_excel(xlsx_path, self._append)
                self._append("")
                self._append("-" * 60)
                if inv_list:
                    gm  = max(max(d.keys()) for d in excel.values() if d)
                    gsm = max((max(len(v) for v in d.values()) for d in excel.values() if d), default=1)
                    rs  = compute_row_step(gm, gsm)
                    self._append(f"Inverters         : {len(inv_list)}")
                    self._append(f"Transformers      : {len(transformer_list)}")
                    self._append(f"Max MPPT channels : {gm}")
                    self._append(f"Max strings/MPPT  : {gsm}")
                    self._append(f"Computed row step : {rs:.0f} units")
                self._append("-" * 60)
                self._append("Scan complete.  Ready to generate.")
            except Exception as ex:
                self._append(f"\n[ERROR] {ex}")
        threading.Thread(target=worker, daemon=True).start()


# ------------------------------------------- main application window ------

class SLDAppV5(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("A176LAB  -  DC Single Line Diagram Generator  v5  (Autonomous)")
        self.geometry("960x600")
        self.minsize(740, 520)
        self.resizable(True, True)
        style = ttk.Style(self)
        for theme in ('vista', 'winnative', 'clam', 'default'):
            try:
                style.theme_use(theme)
                break
            except Exception:
                pass
        self._build_ui()
        self._set_icon()
        self.bind('<F5>', lambda _: self._on_generate())

    # ---------------------------------------------------------------- UI ---
    def _build_ui(self):
        self._add_header()
        ttk.Separator(self, orient='horizontal').pack(fill='x', padx=8, pady=(0, 4))
        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        self._build_files_tab(nb)
        self._build_equip_tab(nb)
        self._build_layout_tab(nb)
        self._build_run_tab(nb)

    def _add_header(self):
        hdr = ttk.Frame(self, padding=(10, 8, 10, 6))
        hdr.pack(fill='x')
        logo = _load_logo_tk((72, 72))
        if logo:
            self._logo_ref = logo
            ttk.Label(hdr, image=logo).pack(side='left', padx=(0, 14))
        txt = ttk.Frame(hdr)
        txt.pack(side='left', fill='y', pady=2)
        ttk.Label(txt, text="DC Single Line Diagram Generator  v5  -  Autonomous",
                  font=('Segoe UI', 14, 'bold')).pack(anchor='w')
        ttk.Label(txt,
                  text="A176 LAB  -  Template-free  -  Auto column detection  -  Press F5 to generate",
                  foreground='gray', font=('Segoe UI', 9)).pack(anchor='w')

    def _set_icon(self):
        try:
            from PIL import Image, ImageTk
            img = Image.open(_LOGO_PATH).convert('RGBA')
            img.thumbnail((256, 256), Image.LANCZOS)
            self._icon_ref = ImageTk.PhotoImage(img)
            self.iconphoto(True, self._icon_ref)
        except Exception:
            pass

    def _build_files_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Files  ')
        ttk.Label(tab, text="File Paths", font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))
        self.fe_xlsx = _FileRow(tab, "Excel Cable List:",
                                filetypes=[('Excel files', '*.xlsx *.xls'), ('All files', '*.*')])
        self.fe_xlsx.pack(fill='x', pady=2)
        self.fe_xlsx.var.trace_add('write', self._sync_output_path)
        self.fe_out = _FileRow(tab, "Output DXF:", save=True,
                               filetypes=[('DXF files', '*.dxf'), ('All files', '*.*')])
        self.fe_out.pack(fill='x', pady=2)
        btn_row = ttk.Frame(tab)
        btn_row.pack(fill='x', pady=(10, 0))
        ttk.Button(btn_row, text='  Scan Excel (Preview Parse)  ',
                   command=self._on_scan).pack(side='left')
        ttk.Label(btn_row, text="  <- Auto-detects sheets and columns without generating",
                  foreground='gray').pack(side='left')
        ttk.Separator(tab).pack(fill='x', pady=12)
        hint = (
            "Excel Cable List  --  Any schedule with inverter IDs (e.g. '1.2'), string names,\n"
            "                      and MPPT numbers.  Sheet name is auto-detected.\n\n"
            "Output DXF        --  Auto-fills as <excel_name>_SLD_v5.dxf in the same folder.\n\n"
            "No DXF template or template_data.json required.  Geometry matches Yanel reference.\n"
            "P00 fuse blocks, DC SWITCH labels, MPP headers, and paper-space viewports included.\n\n"
            "Press F5 or click Generate."
        )
        ttk.Label(tab, text=hint, foreground='gray', wraplength=740, justify='left').pack(anchor='w')

    def _build_equip_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Equipment  ')

        self._sec(tab, "Solar Panel")
        cols = ttk.Frame(tab); cols.pack(fill='x', pady=2)
        lp = ttk.Frame(cols); lp.pack(side='left', fill='x', expand=True)
        rp = ttk.Frame(cols); rp.pack(side='left', fill='x', expand=True)
        self.f_panel = _HistoryCombo(lp, "Panel Model:", 'panel_model',
                                     refresh_callback=self._refresh_combos)
        self.f_panel.pack(fill='x', pady=2)
        self.f_pps = _HistoryCombo(rp, "Panels per String:", 'panels_per_string',
                                   default='26', unit='panels', width=14,
                                   refresh_callback=self._refresh_combos)
        self.f_pps.pack(fill='x', pady=2)
        ttk.Label(tab, text="  -> Module Wp is read automatically from the Excel file.",
                  foreground='gray', font=('Segoe UI', 8, 'italic')).pack(
            anchor='w', padx=(196, 6), pady=(0, 6))
        ttk.Separator(tab).pack(fill='x', pady=8)

        self._sec(tab, "Inverter")
        cols2 = ttk.Frame(tab); cols2.pack(fill='x', pady=2)
        li = ttk.Frame(cols2); li.pack(side='left', fill='x', expand=True)
        ri = ttk.Frame(cols2); ri.pack(side='left', fill='x', expand=True)
        self.f_inv = _HistoryCombo(li, "Inverter Model:", 'inverter_model',
                                   refresh_callback=self._refresh_combos)
        self.f_inv.pack(fill='x', pady=2)
        self.f_max_mppts = _FieldRow(ri, "Inverter Max MPPTs:", "Auto",
                                     unit="(Auto = from Excel)", width=14)
        self.f_max_mppts.pack(fill='x', pady=2)
        self.f_dc_switch = _CheckboxRow(li, "Include DC Switch labels:", default=True)
        self.f_dc_switch.pack(fill='x', pady=2)
        self.f_show_cable = _CheckboxRow(ri, "Show Cable Lengths:", default=False)
        self.f_show_cable.pack(fill='x', pady=2)
        self.f_inline_panel = _CheckboxRow(li, "Inline Panel Info per String:", default=False)
        self.f_inline_panel.pack(fill='x', pady=2)

        ttk.Separator(tab).pack(fill='x', pady=10)
        hints = (
            "  [...]  Opens the Preset Manager — add, edit or delete saved values for that field.\n"
            "  Include DC Switch labels  — when checked, adds 'DC SWITCH 1/2/3...' text labels\n"
            "                              between every 4 MPPTs.  Uncheck to omit them.\n"
            "  Inverter Max MPPTs        — leave as 'Auto' to detect the maximum from the Excel.\n"
            "                              Override with a number to force a fixed layout height.\n"
            "  Show Cable Lengths        — appends L+/L- cable lengths to each string label.\n"
            "  Inline Panel Info         — appends panel count and Wp to every string label."
        )
        ttk.Label(tab, text=hints, foreground='gray',
                  font=('Segoe UI', 8, 'italic'),
                  justify='left').pack(anchor='w', padx=4)

    def _build_layout_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Layout  ')

        self._sec(tab, "Grid Spacing")
        cols = ttk.Frame(tab); cols.pack(fill='x', pady=2)
        lc = ttk.Frame(cols); lc.pack(side='left', fill='x', expand=True)
        rc = ttk.Frame(cols); rc.pack(side='left', fill='x', expand=True)
        self.f_col_sp = _FieldRow(lc, "Column Step:", str(COL_STEP),
                                   unit="units", width=14)
        self.f_col_sp.pack(fill='x', pady=2)
        ttk.Label(lc, text=f"  -> Horizontal step between inverter columns (default {COL_STEP}).",
                  foreground='gray', font=('Segoe UI', 8, 'italic')).pack(
            anchor='w', padx=(196, 6), pady=(0, 4))
        self.f_row_sp = _FieldRow(rc, "Row Step:", "Auto",
                                   unit="(Auto = computed)", width=14)
        self.f_row_sp.pack(fill='x', pady=2)
        ttk.Label(rc, text="  -> Vertical step between transformer rows.",
                  foreground='gray', font=('Segoe UI', 8, 'italic')).pack(
            anchor='w', padx=(196, 6), pady=(0, 4))

        ttk.Separator(tab).pack(fill='x', pady=8)
        self._sec(tab, "Heavy Cable Style")
        cols2 = ttk.Frame(tab); cols2.pack(fill='x', pady=2)
        lh = ttk.Frame(cols2); lh.pack(side='left', fill='x', expand=True)
        rh = ttk.Frame(cols2); rh.pack(side='left', fill='x', expand=True)
        self.f_h_sec = _FieldRow(lh, "Heavy Section:", "1x10", unit="mm2", width=12)
        self.f_h_sec.pack(fill='x', pady=2)
        self.f_h_col = _FieldRow(rh, "Heavy Color (ACI):", "40", unit="index", width=12)
        self.f_h_col.pack(fill='x', pady=2)

    def _build_run_tab(self, nb):
        tab = ttk.Frame(nb, padding=14)
        nb.add(tab, text='  Generate  ')
        ctrl = ttk.Frame(tab); ctrl.pack(fill='x', pady=(0, 6))
        self.gen_btn = ttk.Button(ctrl, text='  Generate SLD  (F5)  ',
                                   command=self._on_generate)
        self.gen_btn.pack(side='left', ipadx=10, ipady=4)
        ttk.Button(ctrl, text='Clear Log',
                   command=self._clear_log).pack(side='left', padx=8)
        self.progress = ttk.Progressbar(ctrl, mode='indeterminate', length=160)
        self.progress.pack(side='left', padx=8)
        self.status_var = tk.StringVar(value='Ready.')
        ttk.Label(tab, textvariable=self.status_var, foreground='gray').pack(anchor='w', pady=(0, 4))
        ttk.Label(tab, text='Execution Log', font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(4, 2))
        self.log_box = scrolledtext.ScrolledText(
            tab, state='disabled', font=('Consolas', 9),
            bg='#1e1e1e', fg='#d4d4d4', insertbackground='white',
            relief='flat', borderwidth=1)
        self.log_box.pack(fill='both', expand=True)

    def _sec(self, parent, title):
        ttk.Label(parent, text=title, font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(4, 2))

    # ---------------------------------------------------------------- actions
    def _log(self, msg):
        def _do():
            self.log_box.configure(state='normal')
            self.log_box.insert('end', msg + '\n')
            self.log_box.see('end')
            self.log_box.configure(state='disabled')
        self.after(0, _do)

    def _clear_log(self):
        self.log_box.configure(state='normal')
        self.log_box.delete('1.0', 'end')
        self.log_box.configure(state='disabled')

    def _sync_output_path(self, *_):
        xlsx = self.fe_xlsx.get()
        if xlsx:
            self.fe_out.var.set(os.path.splitext(xlsx)[0] + '_SLD_v5.dxf')

    def _refresh_combos(self):
        for w in (self.f_panel, self.f_pps, self.f_inv):
            w.refresh()

    def _collect(self):
        return {
            'xlsx_path':         self.fe_xlsx.get(),
            'output_path':       self.fe_out.get(),
            'panel_model':       self.f_panel.get(),
            'panels_per_string': self.f_pps.get(),
            'inverter_model':    self.f_inv.get(),
            'max_mppts':         self.f_max_mppts.get(),
            'dc_switch':         self.f_dc_switch.get(),
            'show_cable_info':   self.f_show_cable.get(),
            'inline_panel_info': self.f_inline_panel.get(),
            'col_spacing':       self.f_col_sp.get(),
            'row_spacing':       self.f_row_sp.get(),
            'heavy_section':     self.f_h_sec.get(),
            'heavy_color':       self.f_h_col.get(),
        }

    def _validate(self, cfg):
        errs = []
        if not cfg['xlsx_path']:
            errs.append("Excel file path is required.")
        elif not os.path.isfile(cfg['xlsx_path']):
            errs.append(f"Excel file not found:\n  {cfg['xlsx_path']}")
        if not cfg['output_path']:
            errs.append("Output DXF path is required.")
        else:
            d = os.path.dirname(cfg['output_path'])
            if d and not os.path.isdir(d):
                errs.append(f"Output directory does not exist:\n  {d}")
        for key, lbl in (('heavy_color', 'Heavy Color ACI'),):
            val = cfg.get(key, '')
            if val:
                try:
                    float(val)
                except ValueError:
                    errs.append(f"{lbl}: must be a number (got '{val}').")
        mm = cfg.get('max_mppts', '').strip().lower()
        if mm and mm != 'auto':
            try:
                if int(float(mm)) < 1:
                    errs.append("Max MPPTs must be >= 1.")
            except ValueError:
                errs.append(f"Max MPPTs: number or 'Auto' expected (got '{mm}').")
        return errs

    def _on_scan(self):
        xlsx = self.fe_xlsx.get()
        if not xlsx or not os.path.isfile(xlsx):
            messagebox.showwarning("No Excel File",
                                   "Please select an Excel file first.", parent=self)
            return
        _ScanDialog(self, xlsx)

    def _on_generate(self):
        cfg  = self._collect()
        errs = self._validate(cfg)
        if errs:
            messagebox.showerror("Input Error", "\n\n".join(errs))
            return
        self.gen_btn.configure(state='disabled')
        self.progress.start(12)
        self.status_var.set("Generating ...")
        self._log("=" * 64)
        self._log("Starting autonomous SLD generation (v5 / template-free) ...")

        def worker():
            try:
                _generate(cfg, self._log)
                self.after(0, self._on_success, cfg['output_path'])
            except Exception as ex:
                self.after(0, self._on_error, str(ex))

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self, path):
        self.progress.stop()
        self.gen_btn.configure(state='normal')
        self.status_var.set("Done - file saved.")
        for w in (self.f_panel, self.f_pps, self.f_inv):
            w.record()
        messagebox.showinfo("Success", f"SLD generated successfully!\n\nSaved at:\n{path}")

    def _on_error(self, msg):
        self.progress.stop()
        self.gen_btn.configure(state='normal')
        self.status_var.set("Error - see log.")
        self._log(f"\n[ERROR] {msg}")
        messagebox.showerror("Generation Error", msg)


# ------------------------------------------------------------- CLI --------

def _build_cli():
    p = argparse.ArgumentParser(
        prog='sld_gui_v5.py',
        description='Autonomous DC SLD Generator v5 (template-free, Yanel-geometry)',
    )
    p.add_argument('--excel',           metavar='PATH')
    p.add_argument('--output',          metavar='PATH')
    p.add_argument('--inv-model',       metavar='NAME',  default='')
    p.add_argument('--panel-model',     metavar='NAME',  default='')
    p.add_argument('--panels',          metavar='N',     default='0')
    p.add_argument('--max-mppts',       metavar='N',     default='Auto')
    p.add_argument('--no-dc-switch',    action='store_true',
                   help='Omit DC SWITCH labels; connect MPPT cables directly')
    p.add_argument('--show-cable',      action='store_true')
    p.add_argument('--inline-panel',    action='store_true')
    p.add_argument('--col-spacing',     metavar='N',     default=str(COL_STEP))
    p.add_argument('--row-spacing',     metavar='N',     default='Auto')
    p.add_argument('--heavy-section',   metavar='SEC',   default='1x10')
    p.add_argument('--heavy-color',     metavar='ACI',   default='40')
    return p


def _run_cli(args):
    if not args.excel:
        print("[ERROR] --excel is required in CLI mode.", file=sys.stderr)
        sys.exit(1)
    if not args.output:
        args.output = os.path.splitext(args.excel)[0] + '_SLD_v5.dxf'
        print(f"Output: {args.output}")
    cfg = {
        'xlsx_path':         args.excel,
        'output_path':       args.output,
        'inverter_model':    args.inv_model,
        'panel_model':       args.panel_model,
        'panels_per_string': args.panels,
        'max_mppts':         args.max_mppts,
        'dc_switch':         not args.no_dc_switch,
        'show_cable_info':   args.show_cable,
        'inline_panel_info': args.inline_panel,
        'col_spacing':       args.col_spacing,
        'row_spacing':       args.row_spacing,
        'heavy_section':     args.heavy_section,
        'heavy_color':       args.heavy_color,
    }
    _generate(cfg, print)


# ------------------------------------------------------------ entry point ---

def main():
    parser = _build_cli()
    args, _ = parser.parse_known_args()
    if args.excel:
        _run_cli(args)
        return
    if not _HAS_TK:
        print("[ERROR] tkinter not available. Use CLI: --excel path.xlsx", file=sys.stderr)
        sys.exit(1)
    SLDAppV5().mainloop()


if __name__ == '__main__':
    main()
