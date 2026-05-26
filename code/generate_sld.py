# -*- coding: utf-8 -*-
"""
generate_sld.py
===============
Generate YANEL DC Single Line Diagram DXF from Excel cable list.

LAYOUT (matches original DXF convention):
  - Each TRANSFORMER occupies one horizontal row.
  - INVERTERS within a transformer are placed side-by-side (columns).
  - Rows stack downward; columns grow to the right.
  - Column step ~11,740 units, row step configurable.

HANDLES:
  - Variable number of inverters per transformer
  - Variable number of active MPPTs per inverter
  - 1 or 2 strings per MPPT port (or none → shows "reserve")
  - Completely unused MPPT ports

USAGE:
  python generate_sld.py
  Output:  YANEL_SLD_Generated.dxf  (same folder as DXF_PATH)

REQUIREMENTS:
  pip install ezdxf openpyxl
"""

import importlib

try:
    ezdxf = importlib.import_module('ezdxf')
except ModuleNotFoundError as e:
    raise ModuleNotFoundError(
        "Missing dependency 'ezdxf'. Install it with 'pip install ezdxf'."
    ) from e

import openpyxl
import re
from collections import defaultdict

# ─── FILE PATHS ────────────────────────────────────────────────────────────────
DXF_PATH    = r'C:/Users/user/Desktop/SLD Diagram/YANEL/26S001_2E103 - DC Single Line Diagram.dxf'  #remain same, just read from this path
XLSX_PATH   = r'C:/Users/user/Desktop/SLD Diagram/2025.017 - PV WYMONDLEY/Priory Farm - Lista Cavi - Cavi LV-DC.xlsx'       #change this to your Excel file path
OUTPUT_PATH = r'C:/Users/user/Desktop/SLD Diagram/2025.017 - PV WYMONDLEY/WYMONDLEY_SLD_Generated.dxf'                    #change this to your desired output path

# ─── TEMPLATE SETTINGS ─────────────────────────────────────────────────────────
# Y band that identifies Inverter 1.1 as the template in the source DXF
TEMPLATE_Y_MIN = 159400
TEMPLATE_Y_MAX = 168000

# Horizontal column step between inverters within the same transformer.
# Measured from the original DXF (Tx2 inverter titles spaced ~11,740 units).
COL_STEP = 11740

# Vertical row step between transformer rows.
# Slightly larger than the template height (~8,600) to leave a visible gap.
ROW_STEP = 10200

# X cutoff: only use the narrow inverter-box portion of the template.
# The template inverter 1.1 also contains a wide cable-bus (x up to 257,000)
# that is row-level geometry; we exclude it so inverter columns don't overlap.
# All MPPT labels, port labels, and string labels sit well within this limit.
X_CUTOFF_RELATIVE = COL_STEP        # keep entities within one column-width of left edge


# ══════════════════════════════════════════════════════════════════════════════
# 1.  READ EXCEL
#     Sheet "2E802-3"  Col1=Inverter ID (e.g. "1.2")  Col3=String name  Col4=MPPT#
#     Result: excel[(T, I)][mppt_no] = ['T.I.N', ...]
# ══════════════════════════════════════════════════════════════════════════════
def read_excel():
    wb   = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws   = wb['2E802-3']
    data = {}
    cur  = None
    for i in range(1, ws.max_row + 1):
        inv   = ws.cell(row=i, column=1).value
        sname = ws.cell(row=i, column=3).value
        mppt  = ws.cell(row=i, column=4).value
        if inv is not None and '.' in str(inv):
            try:
                t, v = str(inv).split('.')[:2]
                cur  = (int(t), int(v))
                data.setdefault(cur, defaultdict(list))
            except Exception:
                pass
        if sname and cur and mppt:
            try:
                data[cur][int(float(mppt))].append(str(sname))
            except Exception:
                pass
    return data


excel    = read_excel()
inv_list = sorted(excel.keys())          # [(T,I), ...]

# Group by transformer and sort inverters within each transformer
transformers     = {}
for (T, I) in inv_list:
    transformers.setdefault(T, []).append(I)
for T in transformers:
    transformers[T] = sorted(transformers[T])
transformer_list = sorted(transformers.keys())

print(f"Inverters from Excel : {len(inv_list)}")
print(f"Transformers         : {len(transformer_list)}")
for T in transformer_list:
    inv_ids = transformers[T]
    mppt_counts = [len(excel.get((T, I), {})) for I in inv_ids]
    str_counts  = [sum(len(v) for v in excel.get((T, I), {}).values()) for I in inv_ids]
    print(f"  Tx{T:>2}: {len(inv_ids):>3} inverters  "
          f"MPPTs used: {min(mppt_counts)}-{max(mppt_counts)}  "
          f"Strings: {min(str_counts)}-{max(str_counts)}")


# ══════════════════════════════════════════════════════════════════════════════
# 2.  LOAD SOURCE DXF & IDENTIFY TEMPLATE ENTITIES
# ══════════════════════════════════════════════════════════════════════════════
doc = ezdxf.readfile(DXF_PATH)
msp = doc.modelspace()


def entity_y(e):
    """Return a representative Y coordinate for an entity, or None."""
    try:
        t = e.dxftype()
        if t in ('TEXT', 'MTEXT', 'INSERT'):  return e.dxf.insert.y
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'): return e.dxf.center.y
        if t == 'LINE':                         return e.dxf.start.y
        if t == 'LWPOLYLINE':
            pts = list(e.get_points())
            return pts[0][1] if pts else None
        if t == 'POLYLINE':
            vs = list(e.vertices)
            return vs[0].dxf.location.y if vs else None
    except Exception:
        pass
    return None


tmpl_live = [e for e in msp
             if (y := entity_y(e)) is not None
             and TEMPLATE_Y_MIN <= y <= TEMPLATE_Y_MAX]
print(f"\nTemplate entities (live in source DXF) : {len(tmpl_live)}")


# ══════════════════════════════════════════════════════════════════════════════
# 3.  EXTRACT ENTITY DATA INTO PLAIN DICTS
#     Must be done BEFORE clearing model space, because ezdxf invalidates
#     deleted-entity references immediately after deletion.
# ══════════════════════════════════════════════════════════════════════════════
def common_attrs(e):
    d = {}
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if e.dxf.hasattr(a):
                d[a] = getattr(e.dxf, a)
        except Exception:
            pass
    return d


def extract_entity(e):
    """Return a plain dict with all drawing data, or None for unsupported types."""
    t = e.dxftype()
    d = {'type': t}
    d.update(common_attrs(e))
    try:
        if t == 'LWPOLYLINE':
            d['pts']    = list(e.get_points())
            d['closed'] = e.closed
            if e.dxf.hasattr('const_width'):
                d['const_width'] = e.dxf.const_width

        elif t == 'MTEXT':
            pos = e.dxf.insert
            d['x']    = pos.x
            d['y']    = pos.y
            d['text'] = e.text
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'style'):
                try:
                    if e.dxf.hasattr(a):
                        d[a] = getattr(e.dxf, a)
                except Exception:
                    pass

        elif t == 'ARC':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'radius':      e.dxf.radius,
                      'start_angle': e.dxf.start_angle,
                      'end_angle':   e.dxf.end_angle})

        elif t == 'CIRCLE':
            c = e.dxf.center
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z, 'radius': e.dxf.radius})

        elif t == 'LINE':
            s, en = e.dxf.start, e.dxf.end
            d.update({'sx': s.x, 'sy': s.y, 'sz': s.z,
                      'ex': en.x, 'ey': en.y, 'ez': en.z})

        elif t == 'ELLIPSE':
            c  = e.dxf.center
            ma = e.dxf.major_axis
            d.update({'cx': c.x, 'cy': c.y, 'cz': c.z,
                      'major_axis':  (ma.x, ma.y, ma.z),
                      'ratio':       e.dxf.ratio,
                      'start_param': e.dxf.start_param,
                      'end_param':   e.dxf.end_param})

        elif t == 'INSERT':
            ins = e.dxf.insert
            d.update({'name': e.dxf.name,
                      'ix': ins.x, 'iy': ins.y,
                      'iz': ins.z if hasattr(ins, 'z') else 0})
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if e.dxf.hasattr(a):
                        d[a] = getattr(e.dxf, a)
                except Exception:
                    pass

        elif t == 'POLYLINE':
            pts3d     = [v.dxf.location for v in e.vertices]
            d['pts3d'] = [(p.x, p.y, p.z) for p in pts3d]

        else:
            return None   # skip unsupported entity types

    except Exception as ex:
        print(f"  [warn] extract_entity {t}: {ex}")
        return None

    return d


# ── Measure template left edge so we can apply the X cutoff ──────────────────
def entity_min_x(d):
    t = d['type']
    try:
        if t == 'LWPOLYLINE':  return min(p[0] for p in d['pts'])
        if t == 'MTEXT':       return d['x']
        if t in ('ARC', 'CIRCLE', 'ELLIPSE'):
                               return d['cx'] - d.get('radius', 0)
        if t == 'LINE':        return min(d['sx'], d['ex'])
        if t == 'INSERT':      return d['ix']
        if t == 'POLYLINE':    return min(p[0] for p in d['pts3d'])
    except Exception:
        pass
    return 0


raw_dicts = [d for e in tmpl_live if (d := extract_entity(e)) is not None]

# Determine the left edge of the template (minimum X across all entities)
TMPL_X_MIN = min((entity_min_x(d) for d in raw_dicts), default=0)
X_CUTOFF   = TMPL_X_MIN + X_CUTOFF_RELATIVE

# Filter to narrow inverter box only (exclude wide cable bus)
tmpl_data  = [d for d in raw_dicts if entity_min_x(d) <= X_CUTOFF]

print(f"Template X left edge : {TMPL_X_MIN:.0f}")
print(f"X cutoff applied     : {X_CUTOFF:.0f}  "
      f"(kept {len(tmpl_data)}/{len(raw_dicts)} entities)")


# ══════════════════════════════════════════════════════════════════════════════
# 4.  CLASSIFY MTEXT ENTITIES
#     Categories:
#       'title'        – "INVERTER 1.1 - P= 350 KWp …"  (one per section)
#       'cabin_header' – "CABIN 1"
#       'cabin_label'  – "Cabin Tx.1 / Inverter 1.1"
#       'string_label' – "String 1.1.1"  or  "reserve"
#       'fixed'        – everything else (copied verbatim)
# ══════════════════════════════════════════════════════════════════════════════
PORT_RE   = re.compile(r'^\d+-\d+$')
STRING_RE = re.compile(r'String \d+\.\d+\.\d+')


def classify(txt):
    c = re.sub(r'\{[^}]*\}', '', re.sub(r'\\[A-Za-z][^;]*;', '', txt))
    c = c.strip().replace('\n', ' ')
    if re.search(r'INVERTER 1\.1', txt, re.I) and 'P=' in txt:
        return 'title'
    if re.search(r'Cabin Tx\.\d+.*Inverter', txt):
        return 'cabin_label'
    if re.match(r'^CABIN \d+$', c.strip()):
        return 'cabin_header'
    if STRING_RE.search(txt) or c.strip() == 'reserve':
        return 'string_label'
    return 'fixed'


tmpl_texts = []
for d in tmpl_data:
    if d['type'] == 'MTEXT':
        d['cls'] = classify(d['text'])
        tmpl_texts.append(d)

print(f"\nMTEXT totals: {len(tmpl_texts)}  |  "
      + "  ".join(f"{k}={sum(1 for m in tmpl_texts if m['cls']==k)}"
                  for k in ('title','cabin_header','cabin_label','string_label','fixed')))


# ══════════════════════════════════════════════════════════════════════════════
# 5.  BUILD MPPT-PORT → STRING-LABEL POSITION MAP
#
#     Port labels ("N-M") sit at fixed positions in the template.
#     For each port label, find the nearest string-label MTEXT to its right
#     (same Y, ±80 units) – that is the slot where we write the string name.
#
#     Special case: MPPT 1, port 1 is the "panel string" at the top,
#     whose string-label has no matching port label.
# ══════════════════════════════════════════════════════════════════════════════
port_labels   = [(m['x'], m['y'], m['text'].strip())
                 for m in tmpl_texts if PORT_RE.match(m['text'].strip())]
string_labels = [m for m in tmpl_texts if m['cls'] == 'string_label']

mppt_port_sl = {}   # (mppt, port) → string-label dict

for px, py, ptxt in port_labels:
    best, best_d = None, 9999
    for sl in string_labels:
        if abs(sl['y'] - py) < 80 and sl['x'] > px:
            dd = abs(sl['y'] - py)
            if dd < best_d:
                best_d, best = dd, sl
    if best:
        m2 = re.match(r'^(\d+)-(\d+)$', ptxt)
        if m2:
            mppt_port_sl[(int(m2.group(1)), int(m2.group(2)))] = best

# Panel string: the one string label not matched by any port label
panel_sl = next(
    (sl for sl in string_labels
     if not any(abs(sl['y'] - py) < 80 for _, py, _ in port_labels)),
    None
)
if panel_sl:
    mppt_port_sl[(1, 1)] = panel_sl

print(f"MPPT/port slots found in template: {len(mppt_port_sl)}")
mppt_nums = sorted(set(mppt for mppt, _ in mppt_port_sl))
print(f"MPPTs represented: {mppt_nums}")


# ══════════════════════════════════════════════════════════════════════════════
# 6.  STRING LOOKUP  (variable MPPTs & string counts per inverter)
#
#     Returns "String T.I.N" if a string is assigned to this slot,
#     or "reserve" if the slot is empty (MPPT unused or port unused).
# ══════════════════════════════════════════════════════════════════════════════
def get_string(T, I, mppt, port):
    """
    Args:
        T    - transformer number
        I    - inverter number within transformer
        mppt - MPPT number (1-based)
        port - port within MPPT (1 or 2)

    Returns:
        "String T.I.N"  if assigned
        "reserve"       if the MPPT is absent or the port index is out of range
    """
    inv_strings = excel.get((T, I), {})
    mppt_list   = inv_strings.get(mppt, [])   # [] if MPPT not present for this inverter
    idx         = port - 1
    if idx < len(mppt_list):
        return f"String {mppt_list[idx]}"
    return "reserve"


# ══════════════════════════════════════════════════════════════════════════════
# 7.  POSITION HELPER
#
#     Inverters in the same transformer share the same row (same dy).
#     Inverter index within the transformer determines the column (dx).
# ══════════════════════════════════════════════════════════════════════════════
def inv_offsets(T, I):
    """
    Return (dx, dy) for inverter (T, I).
    All template entities are shifted by this offset when stamping copies.
    """
    row = transformer_list.index(T)          # 0-based transformer row
    col = transformers[T].index(I)           # 0-based inverter column within row
    dx  = col * COL_STEP
    dy  = -row * ROW_STEP
    return dx, dy


# ══════════════════════════════════════════════════════════════════════════════
# 8.  ENTITY PLACEMENT HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def apply_common(ne, d):
    for a in ('layer', 'color', 'linetype', 'lineweight', 'ltscale'):
        try:
            if a in d:
                setattr(ne.dxf, a, d[a])
        except Exception:
            pass


def add_entity(layout, d, dx, dy):
    """Stamp one entity from its dict, shifted by (dx, dy)."""
    t = d['type']
    try:
        if t == 'LWPOLYLINE':
            pts = [(p[0] + dx, p[1] + dy) + tuple(p[2:]) for p in d['pts']]
            ne  = layout.add_lwpolyline(pts)
            ne.closed = d.get('closed', False)
            if 'const_width' in d:
                ne.dxf.const_width = d['const_width']
            apply_common(ne, d)

        elif t == 'MTEXT':
            attribs = {'insert': (d['x'] + dx, d['y'] + dy)}
            for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                      'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
                if a in d:
                    attribs[a] = d[a]
            ne = layout.add_mtext(d['text'], dxfattribs=attribs)
            apply_common(ne, d)

        elif t == 'ARC':
            ne = layout.add_arc(
                (d['cx'] + dx, d['cy'] + dy, d['cz']),
                d['radius'], d['start_angle'], d['end_angle'])
            apply_common(ne, d)

        elif t == 'CIRCLE':
            ne = layout.add_circle(
                (d['cx'] + dx, d['cy'] + dy, d['cz']),
                d['radius'])
            apply_common(ne, d)

        elif t == 'LINE':
            ne = layout.add_line(
                (d['sx'] + dx, d['sy'] + dy, d['sz']),
                (d['ex'] + dx, d['ey'] + dy, d['ez']))
            apply_common(ne, d)

        elif t == 'ELLIPSE':
            ne = layout.add_ellipse(
                center=(d['cx'] + dx, d['cy'] + dy, d['cz']),
                major_axis=d['major_axis'],
                ratio=d['ratio'],
                start_param=d['start_param'],
                end_param=d['end_param'])
            apply_common(ne, d)

        elif t == 'INSERT':
            ne = layout.add_blockref(
                d['name'], (d['ix'] + dx, d['iy'] + dy, d['iz']))
            for a in ('xscale', 'yscale', 'rotation'):
                try:
                    if a in d:
                        setattr(ne.dxf, a, d[a])
                except Exception:
                    pass
            apply_common(ne, d)

        elif t == 'POLYLINE':
            pts = [(p[0] + dx, p[1] + dy, p[2]) for p in d['pts3d']]
            if pts:
                ne = layout.add_polyline3d(pts)
                apply_common(ne, d)

    except Exception as ex:
        print(f"  [warn] add_entity {t}: {ex}")


def add_mtext_var(layout, d, dx, dy, text):
    """Stamp an MTEXT entity at (dx, dy) offset with a custom text string."""
    try:
        attribs = {'insert': (d['x'] + dx, d['y'] + dy)}
        for a in ('char_height', 'width', 'attachment_point', 'flow_direction',
                  'line_spacing_style', 'line_spacing_factor', 'layer', 'style'):
            if a in d:
                attribs[a] = d[a]
        ne = layout.add_mtext(text, dxfattribs=attribs)
        apply_common(ne, d)
    except Exception as ex:
        print(f"  [warn] add_mtext_var: {ex}")


# ══════════════════════════════════════════════════════════════════════════════
# 9.  CLEAR MODEL SPACE
# ══════════════════════════════════════════════════════════════════════════════
print("\nClearing existing model space …")
for e in list(msp):
    try:
        msp.delete_entity(e)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════════════════
# 10. GENERATE ALL INVERTER SECTIONS
# ══════════════════════════════════════════════════════════════════════════════
title_d      = next((m for m in tmpl_texts if m['cls'] == 'title'),        None)
cabin_hdr_d  = next((m for m in tmpl_texts if m['cls'] == 'cabin_header'), None)
cabin_lbl_d  = next((m for m in tmpl_texts if m['cls'] == 'cabin_label'),  None)

print(f"Generating {len(inv_list)} inverter sections …")
for idx, (T, I) in enumerate(inv_list):
    dx, dy = inv_offsets(T, I)

    # ── 10a. Geometry and fixed MTEXT (copied verbatim) ──
    for d in tmpl_data:
        if d['type'] == 'MTEXT':
            if d.get('cls') == 'fixed':
                add_entity(msp, d, dx, dy)
            # variable MTEXTs handled below
        else:
            add_entity(msp, d, dx, dy)

    # ── 10b. Variable MTEXT: inverter title ──
    if title_d:
        add_mtext_var(msp, title_d, dx, dy,
                      f"INVERTER {T}.{I} - P= 350 KWp - P= 320 KWac @40°C")

    # ── 10c. Variable MTEXT: cabin header ──
    if cabin_hdr_d:
        add_mtext_var(msp, cabin_hdr_d, dx, dy, f"CABIN {T}")

    # ── 10d. Variable MTEXT: cabin label ──
    if cabin_lbl_d:
        add_mtext_var(msp, cabin_lbl_d, dx, dy,
                      f"\\pxqc;Cabin Tx.{T}\\PInverter {T}.{I}")

    # ── 10e. Variable MTEXT: string labels per MPPT port ──
    #     get_string() returns "reserve" if the MPPT is unused or
    #     if the port index exceeds the number of assigned strings.
    for (mppt, port), sl in mppt_port_sl.items():
        label = get_string(T, I, mppt, port)
        add_mtext_var(msp, sl, dx, dy, label)

    if (idx + 1) % 10 == 0:
        print(f"  {idx + 1:>3}/{len(inv_list)} done")

print("All inverter sections generated.")


# ══════════════════════════════════════════════════════════════════════════════
# 11. PAPER SPACE LAYOUTS  (one per transformer)
#     Each layout shows the full horizontal row for that transformer,
#     scaled to fit on A3 landscape paper.
# ══════════════════════════════════════════════════════════════════════════════
existing = [l.name for l in doc.layouts if l.name != 'Model']
for lname in existing:
    try:
        doc.layouts.delete(lname)
    except Exception:
        pass

TMPL_X_CENTER = TMPL_X_MIN + COL_STEP / 2
TMPL_HEIGHT   = TEMPLATE_Y_MAX - TEMPLATE_Y_MIN
TMPL_Y_CENTER = (TEMPLATE_Y_MIN + TEMPLATE_Y_MAX) / 2

for T in transformer_list:
    lname = f"Tx{T}"
    try:
        layout = doc.layouts.new(lname)
    except Exception:
        layout = doc.layouts.get(lname)

    n_inv = len(transformers[T])
    row   = transformer_list.index(T)

    # Model-space centre of this transformer's row
    row_center_x = TMPL_X_CENTER + (n_inv - 1) * COL_STEP / 2
    row_center_y = TMPL_Y_CENTER - row * ROW_STEP

    # Full width of the row in model space
    row_width = COL_STEP * n_inv

    # A3 landscape: 420 × 297 mm
    # view_height chosen so the full row width fits horizontally
    view_height = max(
        TMPL_HEIGHT * 1.05,
        row_width / (420.0 / 297.0) * 1.05
    )

    layout.add_viewport(
        center=(210, 148.5),
        size=(420, 297),
        view_center_point=(row_center_x, row_center_y),
        view_height=view_height,
    )

print(f"Created {len(transformer_list)} paper-space layouts "
      f"({', '.join(f'Tx{T}' for T in transformer_list)}).")


# ══════════════════════════════════════════════════════════════════════════════
# 12. SAVE
# ══════════════════════════════════════════════════════════════════════════════
print(f"\nSaving → {OUTPUT_PATH}")
doc.saveas(OUTPUT_PATH)
print("Done!")
