# -*- coding: utf-8 -*-
"""
app.py - Flask web application for the A176LAB SLD Generator (sld-auto).

Routes:
    GET  /                  - main form
    POST /generate          - start generation job, returns {job_id}
    GET  /stream/<job_id>   - SSE log stream
    GET  /download/<job_id> - download the generated DXF
    GET  /status/<job_id>   - JSON job status
"""
import json
import os
import queue
import re
import shutil
import sys
import tempfile
import threading
import time
import uuid
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sld_core import (
    COL_SPACING_DEFAULT,
    ROW_SPACING_DEFAULT,
    _INVERTER_POWERS,
    _PANELS_PRESETS,
    generate,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 150 * 1024 * 1024  # 150 MB

_JOBS: dict = {}
_JOBS_LOCK = threading.Lock()

_UPLOAD_ROOT = Path(tempfile.gettempdir()) / "sld_auto_jobs"
_UPLOAD_ROOT.mkdir(exist_ok=True)

# Template path: set TEMPLATE_DXF env var or place file at webapp/template/template.dxf
_TEMPLATE_DXF = os.environ.get(
    "TEMPLATE_DXF",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "template", "template.dxf"),
)

# ── Input validation constants ────────────────────────────────────────────────

_ALLOWED_EXCEL_EXT = {".xlsx", ".xls"}

# Strict UUID v4 pattern - rejects any non-UUID job_id in URL params
_UUID4_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-4[0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}$"
)


def _valid_job_id(job_id: str) -> bool:
    return bool(_UUID4_RE.match(job_id))


def _allowed_excel(filename: str) -> bool:
    return Path(filename).suffix.lower() in _ALLOWED_EXCEL_EXT


def _json_safe(data) -> str:
    # Embed JSON in <script> tags safely: replace HTML-sensitive chars with
    # their JSON Unicode escapes so a stray </script> cannot break the page.
    return (
        json.dumps(data)
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
        .replace("'", "\\u0027")
    )


# ── Security response headers ─────────────────────────────────────────────────

@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    return response


# ── Job helpers ───────────────────────────────────────────────────────────────

def _cleanup_job(job_id: str) -> None:
    with _JOBS_LOCK:
        job = _JOBS.pop(job_id, None)
    if not job:
        return
    job_dir = job.get("job_dir")
    if job_dir:
        shutil.rmtree(job_dir, ignore_errors=True)  # removes all files even if rmdir would fail


def _cleanup_stale_jobs() -> None:
    """Background daemon: evict jobs older than 2 hours to prevent memory/disk accumulation."""
    while True:
        time.sleep(3600)
        cutoff = time.time() - 7200
        with _JOBS_LOCK:
            stale = [jid for jid, j in _JOBS.items() if j.get("created", 0) < cutoff]
        for jid in stale:
            _cleanup_job(jid)


threading.Thread(target=_cleanup_stale_jobs, daemon=True).start()


def _run_generation(job_id: str, cfg: dict) -> None:
    with _JOBS_LOCK:
        log_q = _JOBS[job_id]["log_queue"]

    def log_cb(msg: str) -> None:
        log_q.put({"type": "log", "msg": str(msg)})

    try:
        generate(cfg, log_cb=log_cb)
        log_q.put({"type": "done", "msg": "SLD generated successfully."})
        with _JOBS_LOCK:
            _JOBS[job_id]["status"] = "done"
    except Exception as exc:
        err = str(exc)
        log_q.put({"type": "error", "msg": err})
        with _JOBS_LOCK:
            _JOBS[job_id]["status"] = "error"
            _JOBS[job_id]["error"] = err
    finally:
        log_q.put(None)  # sentinel - closes the SSE stream


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template(
        "index.html",
        inverter_powers=_json_safe(_INVERTER_POWERS),
        panel_presets=_json_safe(_PANELS_PRESETS),
        col_spacing_default=COL_SPACING_DEFAULT,
        row_spacing_default=ROW_SPACING_DEFAULT,
    )


@app.route("/parse-excel", methods=["POST"])
def parse_excel_wp():
    """Quick parse: detect distinct Wp values from the cable-schedule Excel."""
    if "excel_file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["excel_file"]
    filename = f.filename or ""
    if not _allowed_excel(filename):
        return jsonify({"error": "Only .xlsx and .xls files are accepted"}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)

        # Locate master sheet '2E802-3' (same logic as generate())
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == "2e802-3":
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Find header row (search rows 26-32 for 'string name')
        header_row = 30
        col_wp = 21  # default column U
        for r_idx in (30, 29, 28, 31, 32, 27, 26):
            row_vals = [
                str(ws.cell(row=r_idx, column=c).value or "").strip().lower()
                for c in range(1, 35)
            ]
            if any("string name" in v for v in row_vals):
                header_row = r_idx
                # Two-tier Wp column detection (mirrors engine logic)
                _wp_strong = False
                for c in range(1, 35):
                    val = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
                    if any(kw in val for kw in ("module type", "module power", "watt peak")):
                        col_wp = c
                        _wp_strong = True
                    elif not _wp_strong and (
                        "potenza" in val or re.search(r"(?<![a-z])wp(?![a-z])", val)
                    ):
                        col_wp = c
                break

        # Collect distinct Wp values from data rows
        wp_set: set = set()
        for row in ws.iter_rows(min_row=header_row + 1, min_col=col_wp, max_col=col_wp):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                try:
                    w = int(float(str(v).replace(",", ".")))
                    if 50 < w < 2000:
                        wp_set.add(w)
                except (ValueError, TypeError):
                    pass

        wb.close()
        return jsonify({"wp_values": sorted(wp_set)})

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/generate", methods=["POST"])
def start_generation():
    # ── File presence & extension ─────────────────────────────────────────────
    if "excel_file" not in request.files:
        return jsonify({"error": "An Excel cable schedule file is required."}), 400

    excel_file = request.files["excel_file"]
    filename = excel_file.filename or ""

    if not filename:
        return jsonify({"error": "Please select an Excel file before submitting."}), 400

    if not _allowed_excel(filename):
        return jsonify({"error": "Only .xlsx and .xls files are accepted."}), 400

    # ── Template availability ─────────────────────────────────────────────────
    if not os.path.isfile(_TEMPLATE_DXF):
        return jsonify({
            "error": (
                "Template DXF not found on the server. "
                "Place your template at /app/template/template.dxf "
                "or set the TEMPLATE_DXF environment variable."
            )
        }), 500

    # ── Save upload ───────────────────────────────────────────────────────────
    job_id = str(uuid.uuid4())
    job_dir = _UPLOAD_ROOT / job_id
    job_dir.mkdir(exist_ok=True)

    safe_name = secure_filename(filename) or "input.xlsx"
    excel_path = str(job_dir / safe_name)
    output_path = str(job_dir / "output_SLD.dxf")

    excel_file.save(excel_path)

    # ── Build generation config ───────────────────────────────────────────────
    form = request.form
    cfg = {
        "xlsx_path":           excel_path,
        "template_dxf":        _TEMPLATE_DXF,
        "output_path":         output_path,
        "panel_model":         form.get("panel_model", ""),
        "panel_model_map":     form.get("panel_model_map", ""),
        "panels_per_string":   form.get("panels_per_string", "0"),
        "strings_per_mppt":    form.get("strings_per_mppt", "2"),
        "mppts_per_switch":    form.get("mppts_per_switch", "4"),
        "num_mppts_total":     form.get("num_mppts_total", "0"),
        "mppt_layout":         form.get("mppt_layout", ""),
        "inverter_model":      form.get("inverter_model", ""),
        "dc_power_kwp":        form.get("dc_power_kwp", "0"),
        "ac_power_kwac":       form.get("ac_power_kwac", "0"),
        "temp_rating":         form.get("temp_rating", "40"),
        "transformer_power":   form.get("transformer_power", ""),
        "show_cable_info":     form.get("show_cable_info") == "true",
        "hide_string_details": form.get("hide_string_details") == "true",
        "show_annot_circle":   form.get("show_annot_circle", "true") == "true",
        "col_spacing":         form.get("col_spacing", str(COL_SPACING_DEFAULT)),
        "row_spacing":         form.get("row_spacing", str(ROW_SPACING_DEFAULT)),
        "circle_radius":       form.get("circle_radius", "24.59"),
        "text_size":           form.get("text_size", "60.44"),
        "heavy_section":       form.get("heavy_section", "1x10"),
        "heavy_linetype":      form.get("heavy_linetype", "TRATTEGGIATA"),
        "heavy_color":         form.get("heavy_color", "40"),
        "heavy_layer":         form.get("heavy_layer", "TRATTEGGIATA"),
    }

    log_queue: queue.Queue = queue.Queue()
    with _JOBS_LOCK:
        _JOBS[job_id] = {
            "status":      "running",
            "log_queue":   log_queue,
            "output_path": output_path,
            "job_dir":     str(job_dir),
            "excel_path":  excel_path,
            "created":     time.time(),
        }

    threading.Thread(target=_run_generation, args=(job_id, cfg), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/stream/<job_id>")
def stream_logs(job_id: str):
    if not _valid_job_id(job_id):
        return Response(status=400)

    with _JOBS_LOCK:
        job = _JOBS.get(job_id)

    if not job:
        def _not_found():
            yield f"data: {json.dumps({'type': 'error', 'msg': 'Job not found'})}\n\n"
        return Response(_not_found(), mimetype="text/event-stream")

    log_q: queue.Queue = job["log_queue"]

    def event_stream():
        while True:
            try:
                msg = log_q.get(timeout=30)
                if msg is None:  # sentinel
                    yield f"data: {json.dumps({'type': 'sentinel'})}\n\n"
                    break
                yield f"data: {json.dumps(msg)}\n\n"
            except queue.Empty:
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.route("/download/<job_id>")
def download_result(job_id: str):
    if not _valid_job_id(job_id):
        return "Invalid job ID", 400

    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        return "Job not found", 404
    if job["status"] != "done":
        return "Job not complete yet", 400
    output_path = job["output_path"]
    if not os.path.exists(output_path):
        return "Output file missing", 404

    response = send_file(
        output_path,
        as_attachment=True,
        download_name="SLD_Generated.dxf",
        mimetype="application/octet-stream",
    )
    threading.Timer(10.0, _cleanup_job, args=(job_id,)).start()
    return response


@app.route("/status/<job_id>")
def job_status(job_id: str):
    if not _valid_job_id(job_id):
        return jsonify({"status": "invalid"}), 400

    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        return jsonify({"status": "not_found"}), 404
    return jsonify({"status": job["status"], "error": job.get("error", "")})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
